#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IntelliBI - Consolidated Master Lead Report builder
===================================================

Merges leads from FIVE sources (IntelliBI Lead Information, Walk-In, Website,
WhatsApp/Interakt, Direct Calling/Exotel) into ONE de-duplicated master dataset
that serves as the single source of truth for all downstream reporting.

Implements, per the specification:
  * Basic data cleansing (reusable utility functions) applied to every field
  * Duplicate detection with priority  Phone > Email  (Full Name NOT used)
  * Source-priority field merge         IntelliBI > Walk-In > Direct Calling > WhatsApp > Website
  * Source tracking columns             IsWalk-In / IsWebsite / IsWhatsapp / IsCall
  * "Counselling By" priority           IntelliBI > Call > WhatsApp > Walk-In > Website
  * Lead Interaction History            full chronological enquiry journey
  * UPSERT logic                        Insert / Update / Skip vs. existing master
  * Execution summary logging

The script can run in two INPUT modes:
    INPUT_MODE = "csv"     -> read the four exported CSVs from LOCAL_DIR   (offline / testing)
    INPUT_MODE = "gsheet"  -> read the four source Google Sheets live via gspread (production)

and two OUTPUT modes:
    write_csv_xlsx()       -> always produces consolidated_master.csv / .xlsx + summary
    push_to_gsheet()       -> UPSERT the master into the target Google Sheet (production)

Author: generated for IntelliBI Innovations Technologies.
"""

import os
import re
import sys
import json
import difflib
import unicodedata
from datetime import datetime, timedelta, timezone
from collections import defaultdict, OrderedDict

import pandas as pd

try:
    from dateutil import parser as dtparser
except Exception:                                   # pragma: no cover
    dtparser = None

# ---------------------------------------------------------------------------
# 1. CONFIGURATION
# ---------------------------------------------------------------------------

# --- IntelliBI Sales Automation portability bootstrap (auto-inserted) ---
import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "common"))
import _bootstrap  # noqa: E402  (sys.path + env defaults + config.yaml)
from paths import CREDENTIALS_DIR, CONFIG_DIR, LOGS_DIR  # noqa: E402
# --- end bootstrap ---

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Default is PRODUCTION: read the 4 source Google Sheets live and write the
# master into the target sheet. Set INTELLIBI_INPUT_MODE=csv only for offline
# testing against exported CSVs in LOCAL_DIR.
INPUT_MODE = os.environ.get("INTELLIBI_INPUT_MODE", "gsheet")   # "gsheet" | "csv"
LOCAL_DIR  = os.environ.get("INTELLIBI_LOCAL_DIR", _SCRIPT_DIR)
OUT_DIR    = os.environ.get("INTELLIBI_OUT_DIR", _SCRIPT_DIR)

# Source Google Sheet IDs (used in gsheet mode). First tab is read from each.
SOURCE_SHEETS = {
    # Highest-priority source: the IntelliBI Lead Information form backend
    # ("PhoneNumber_FiledMapping_Data"). First tab (gid=0) is read.
    "IntelliBI": {"id": "1ReJVPl_Y8WnOl_P2sui_uC1jjZXVk0dWqNWRcXGVHCw", "csv": "intellibi.csv"},
    "Walk-In":  {"id": "19Ecal2JpOL1FbzGKWlno4ZywG3HsXsiK-BmMzew5TqQ", "csv": "walkin.csv"},
    "Website":  {"id": "1prW3GKMnGJZ2U5b0gKjLqTJfczfFTYxUwmWImneDtnE", "csv": "website.csv"},
    "WhatsApp": {"id": "1s6fscV531_zozRqT2sZQzJP7LxyqzySRNizMsaTUAq8", "csv": "whatsapp.csv"},
    "Call":     {"id": "1L-Ew4-GF7MzzAnnIhVBafOmN_DJlMBTaRI6048PUo4I", "csv": "calling.csv"},
}

# Target consolidated master sheet ("IntelliBI Consolidate Sales Tracking Report").
TARGET_SHEET_FULL = os.environ.get(
    "INTELLIBI_TARGET_SHEET_ID", "1zZQjXnMJD96Ca0MNyfSt4-XS0z5w3rT7WPdb9qsP1Gs")

# Configuration sheet that maps 'Current Status' values -> 'Lead Type' category.
# Layout is WIDE: each column header is a Lead Type and the cells below it list the
# Current Status values that belong to it. The mapping is read LIVE every run, so
# editing the sheet (adding/removing values) automatically updates classification
# with no code change. The sheet must be shared with the service account (Viewer).
LEAD_TYPE_MAP_SHEET_ID = os.environ.get(
    "INTELLIBI_LEAD_TYPE_MAP_SHEET_ID", "1b7KbkJ3a8QvL2RVDyEgNdcduzGBbZ0_bsY-gZGK0UMo")
LEAD_TYPE_UNIDENTIFIED = "Unidentified"

# Company / virtual (Exotel) phone lines that must NEVER be treated as leads.
# These are auto-augmented at runtime with every distinct 'To' business line
# seen in the calling log. Add any further office/test numbers here.
BUSINESS_NUMBERS_SEED = {
    "8047286087",   # main Exotel business line (To / PhoneNumber)
    "7070734242",   # virtual / IVR line
    "7387028359",   # virtual / IVR line
    "7070544949",   # internal / office origination line (24 self-originated calls)
}

# Google service-account credentials (production mode) — same stack as the rest
# of the repo: google-api-python-client + google-auth, Sheets v4 API.
SERVICE_ACCOUNT_FILE = os.environ.get(
    "GOOGLE_SERVICE_ACCOUNT_FILE",
    os.environ.get("GOOGLE_APPLICATION_CREDENTIALS",
                   os.path.join(CREDENTIALS_DIR, "service_account.json")))
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

# ---------------------------------------------------------------------------
# 2. TARGET SCHEMA
# ---------------------------------------------------------------------------

# "Final Sales Lead Tracker Fields" from the mapping document, in order.
TARGET_FIELDS = [
    "LeadInitalTimestamp",
    "Full Name",
    "Email Address",
    "Mobile Number",
    "Current City",
    "Current Area / Locality",
    "Preferred Contact Method",
    "Highest Qualification",
    "Graduation / Passing Year",
    "College / University Name",
    "Grade / Percentage / CGPA",
    "Current Status",
    "Lead Type",
    "Current Company Name",
    "Total Years of Experience",
    "Current Domain / Technology",
    "Which technology are you interested in learning?",
    "Course Advised",
    "What is your primary goal?",
    "How did you hear about IntelliBI?",
    "IsReferral",
    "Referrer's Name",
    "Preferred Learning Mode",
    "Preferred Batch Timing",
    "When are you planning to take admission?",
    "Remarks",
    "IsGoogleMeetSchedule",
    "IsGoogleMeetScheduleDate",
    "IsWalkInSchedule",
    "IsWalkInScheduleDate",
    "Lead Status",
    "Admission Status",
    "Backout Reason",
    "Counselling By",
    "ScheduledOrDirectWalkIn",
    "Follow-Up Type",
    "Next Follow-Up Date",
]

TRACKING_FIELDS = ["IsWalk-In", "IsWebsite", "IsWhatsapp", "IsCall"]

# Explicitly requested history column + a few analytics helpers for Lead Journey.
HISTORY_FIELD   = "Lead Interaction History"
ANALYTICS_FIELDS = [
    "First Enquiry Date",
    "Latest Enquiry Date",
    "Number of Interactions",
    "Platforms Used",
]

# Record-quality flags (recomputed every run; invalid/irrelevant records are
# KEPT but flagged for review):
#   IsPhoneNumberValid : Yes/No on the normalised mobile (with "invalid" remark
#                        exception)
#   IsLeadRelevant     : No if the remark indicates an irrelevant lead, else Yes
VALIDATION_FIELD = "IsPhoneNumberValid"
RELEVANCE_FIELD  = "IsLeadRelevant"

MASTER_COLUMNS = (TARGET_FIELDS + TRACKING_FIELDS + [VALIDATION_FIELD, RELEVANCE_FIELD]
                  + [HISTORY_FIELD] + ANALYTICS_FIELDS)

# Source-priority for filling a field (lower rank wins). IntelliBI Lead
# Information is the HIGHEST priority (rank 0); the rest keep their order.
SOURCE_MERGE_RANK = {"IntelliBI": 0, "Walk-In": 1, "Call": 2, "WhatsApp": 3, "Website": 4}
# Priority for the dedicated "Counselling By" column (IntelliBI first).
COUNSELLING_PRIORITY = ["IntelliBI", "Call", "WhatsApp", "Walk-In", "Website"]
# Human label for the interaction history "Lead Source" line.
SOURCE_LABEL = {"IntelliBI": "IntelliBI", "Walk-In": "Walk-In", "Website": "Website",
                "WhatsApp": "WhatsApp", "Call": "Call"}


# ---------------------------------------------------------------------------
# 3. REUSABLE CLEANSING UTILITIES
#    (single implementation, applied consistently to all sources)
# ---------------------------------------------------------------------------

_NULL_TOKENS = {"", "null", "n/a", "na", "n.a", "n.a.", "-", "--", "none", "nil",
                "#n/a", "#ref!", "#value!", "nan", "not sure", "notsure"}
_WS_RE       = re.compile(r"\s+")
_EMAIL_RE    = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_CTRL_RE     = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def clean_text(value, treat_null_tokens=True):
    """General field cleanser: strip, collapse whitespace, drop control chars,
    remove invisible characters, and convert NULL-like tokens to ''. Preserves
    the meaning of valid business text."""
    if value is None:
        return ""
    # Defensive: if a duplicate-named column ever slips through as a Series/list,
    # collapse it to the first non-empty scalar rather than stringifying the whole
    # object (which produced 'Total Experience Total Experience Name: 57, dtype: str').
    if isinstance(value, (pd.Series, list, tuple)):
        seq = value.tolist() if isinstance(value, pd.Series) else list(value)
        value = ""
        for v in seq:
            sv = "" if v is None else str(v).strip()
            if sv and sv.lower() != "nan":
                value = v
                break
    s = str(value)
    if s.lower() == "nan":
        return ""
    # Normalise unicode, strip zero-width / invisible chars.
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("​", "").replace("‌", "").replace("‍", "").replace("﻿", "")
    s = _CTRL_RE.sub(" ", s)
    s = s.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    s = _WS_RE.sub(" ", s).strip()
    if treat_null_tokens and s.lower() in _NULL_TOKENS:
        return ""
    return s


# emoji / pictographs / symbols we strip from display names & status values
_EMOJI_RE = re.compile(
    "["
    "\U0001F000-\U0001FAFF"
    "\U00002600-\U000027BF"
    "\U0001F1E6-\U0001F1FF"
    "\U00002190-\U000021FF"
    "\U00002B00-\U00002BFF"
    "️♀♂⚕"
    "]+", flags=re.UNICODE)


def strip_emoji(s):
    return _WS_RE.sub(" ", _EMOJI_RE.sub("", s)).strip()


def clean_name(value):
    """Return a clean DISPLAY name (spaces collapsed, emoji removed, trimmed)."""
    s = clean_text(value)
    s = strip_emoji(s)
    return s.strip(" .,-_/")


def name_key(value):
    """Case-insensitive, punctuation/emoji-free matching key for a name.
    Returns '' for junk (single chars, digit-only 'names', empty)."""
    s = clean_name(value).lower()
    s = re.sub(r"[^0-9a-zÀ-ɏ\s]", " ", s)   # keep latin letters/accents + digits
    s = _WS_RE.sub(" ", s).strip()
    if not s:
        return ""
    if s.replace(" ", "").isdigit():                   # a phone masquerading as a name
        return ""
    if len(s.replace(" ", "")) < 2:                    # single-char junk
        return ""
    if is_status_name(value):                          # 'irrelevant'/'invalid'/'dnp'… are NOT names
        return ""
    return s


# call-status / disposition words that agents sometimes type into a NAME field.
# These must never become a display name or a dedupe key (else every record with
# the same word merges into one lead).
_NAME_STATUS_TOKENS = {
    "invalid", "irrelevant", "dnp", "did not pick", "test", "testing", "unknown",
    "wrong number", "wrong no", "not interested", "not reachable", "switch off",
    "switched off", "busy", "no answer", "ringing", "call back", "callback",
    "follow up", "followup", "pending", "spam", "hang up", "hangup", "cut",
    "call cut", "no name", "blank", "wn", "ni", "dnd", "not relevant",
}
_STATUS_IRRELEVANT_PREFIXES = ("irrelev", "irrelav", "irreval", "irrevel", "irelev", "irrelvant")


def is_status_name(value):
    """True if a 'name' is really a call-status/disposition token (irrelevant,
    invalid, dnp, wrong number, …) rather than a person's name. Fuzzy on the
    'irrelevant' family; substring on 'invalid'."""
    s = re.sub(r"[^a-z\s]", " ", clean_name(value).lower())
    s = _WS_RE.sub(" ", s).strip()
    if not s:
        return False
    if s in _NAME_STATUS_TOKENS:
        return True
    if "invalid" in s or "not relevant" in s or "not interested" in s or "wrong number" in s:
        return True
    for tok in s.split():
        if len(tok) >= 6 and tok[0] == "i" and (
                tok.startswith(_STATUS_IRRELEVANT_PREFIXES)
                or difflib.SequenceMatcher(None, tok, "irrelevant").ratio() >= 0.8):
            return True
    return False


def clean_lead_name(value):
    """Display name for a lead; blanks call-status/junk tokens so they never
    become a name or a merge key."""
    name = clean_name(value)
    return "" if is_status_name(name) else name


def clean_email(value):
    """Lowercase, remove stray spaces (incl. around @), validate. Invalid -> ''."""
    s = clean_text(value).lower()
    s = re.sub(r"\s*@\s*", "@", s)
    s = s.replace(" ", "")
    return s if _EMAIL_RE.match(s) else ""


def norm_phone(value):
    """Normalise to a bare national 10-digit number for matching.
    Handles +91 / 91 / leading-zero / spaces / brackets / hyphens."""
    if value is None:
        return ""
    d = re.sub(r"\D", "", str(value))
    if not d:
        return ""
    d = d.lstrip("0")
    if len(d) == 12 and d.startswith("91"):
        d = d[2:]
    if len(d) == 11 and d.startswith("0"):
        d = d[1:]
    if len(d) > 10:
        d = d[-10:]
    return d


def is_valid_mobile(nphone):
    return len(nphone) == 10 and nphone[0] in "6789"


def display_phone(nphone):
    """Consistent display format for a normalised number."""
    return nphone


_YES = {"yes", "y", "true", "1", "scheduled", "done", "completed"}
_NO  = {"no", "n", "false", "0", "not scheduled", "declined"}


def std_yes_no(value):
    """Standardise a Yes/No flag column to exactly 'Yes' / 'No' / '' (blank).

    Every caller of this is a Yes/No flag (IsReferral, IsGoogleMeetSchedule,
    IsWalkInSchedule). Any other value — a date, timestamp or free text — is a
    data-quality issue for such a column and must NOT silently flow into the
    report (that is how dates were leaking into IsWalkInSchedule), so it is
    dropped to blank rather than passed through. Date/time values continue to be
    populated only in their own *Date columns via fmt_date_field()."""
    s = clean_text(value).lower()
    if s == "":
        return ""
    if s in _YES:
        return "Yes"
    if s in _NO:
        return "No"
    return ""                         # unexpected (date / garbage) -> drop to blank


_COUNSEL_JUNK = {"active", "inactive", "0", "1", "true", "false", "none", "null",
                 "new lead", "cold", "hot", "warm", "-"}


def clean_counsellor(value):
    """Normalise a counsellor / agent name for consistent reporting.
    Title-cases (collapsing case variants like 'arshkhan pathan' vs
    'ArshKhan Pathan') and drops known junk / status tokens."""
    s = clean_name(value)
    if not s or s.lower() in _COUNSEL_JUNK:
        return ""
    # keep separators like '/' but title-case the words
    return " ".join(w if (len(w) > 1 and w.isupper()) else w.title() for w in s.split())


# WhatsApp 'Conversation Label' -> canonical counsellor overrides.
CONVERSATION_LABEL_OVERRIDES = {"arsh": "ArshKhan Pathan", "harish": "Harish Rathod"}


def map_conversation_label(label):
    """Map a WhatsApp 'Conversation Label' to a counsellor. Only assignment
    labels ('assign to X' / 'asign to X') yield a counsellor; anything else
    returns '' so the COALESCE falls through to Assigned Agent / enr_contact_owner.
      'assign to Arsh'   -> 'ArshKhan Pathan'
      'asign to harish'  -> 'Harish Rathod'"""
    s = clean_text(label).lower()
    m = re.match(r"^(?:assign(?:ed)?|asign)\s+to\s+(.+)$", s)
    if not m:
        return ""
    who = m.group(1).strip()
    for kw, name in CONVERSATION_LABEL_OVERRIDES.items():
        if kw in who:
            return name
    return clean_counsellor(who)


def parse_note_field(notes, key):
    """Extract a 'Key: Value' line from a free-text Notes blob (Direct Calling
    stores structured values inside Notes). Punctuation/space-insensitive on the
    key; returns the cleaned value or ''."""
    if not notes:
        return ""
    tgt = re.sub(r"[^a-z0-9]", "", str(key).lower())
    for line in str(notes).splitlines():
        if ":" in line:
            k, _, v = line.partition(":")
            if re.sub(r"[^a-z0-9]", "", k.lower()) == tgt:
                return clean_text(v)
    return ""


# ---------------------------------------------------------------------------
# NOTES (Call) key -> master-column mapping.
# Source of truth: the "Lead Tracker Field Mapping Details" document. This is the
# ONLY place the mapping lives (not scattered through the loader). It can be
# overridden WITHOUT touching code by placing a
# config_files/notes_field_mapping.json of {"Notes Key": "Master Column"} — that
# file, if present, is merged over these defaults, so the mapping stays config-
# driven. Keys are matched case/space/punctuation-insensitively.
# ---------------------------------------------------------------------------
NOTES_FIELD_MAP_DEFAULT = {
    "Full Name":                "Full Name",
    "Current City":             "Current City",
    "Highest Qualification":    "Highest Qualification",
    "Candidate Type":           "Current Status",
    "Current Role":             "Current Domain / Technology",
    "Current Company Name":     "Current Company Name",
    "Total Experience":         "Total Years of Experience",
    "Course Interested In":     "Which technology are you interested in learning?",
    "Course Advised":           "Course Advised",
    "Career Goal":              "What is your primary goal?",
    "IsGoogleMeetSchedule":     "IsGoogleMeetSchedule",
    "IsGoogleMeetScheduleDate": "IsGoogleMeetScheduleDate",
    "IsWalkInSchedule":         "IsWalkInSchedule",
    "IsWalkInScheduleDate":     "IsWalkInScheduleDate",
    "AdmissionStatus":          "Admission Status",
    "Next Follow-upDate":       "Next Follow-Up Date",
    "Counsellor Notes":         "Remarks",
    "IsReferral":               "IsReferral",
    "Referrer's Name":          "Referrer's Name",
    "BackOutReason":            "Backout Reason",
}


def _normkey(k):
    """Normalise a key for tolerant matching: lowercase alphanumerics only, so
    'Next Follow-upDate', 'next follow up date', 'NextFollowUpDate' all match."""
    return re.sub(r"[^a-z0-9]", "", str(k).lower())


def load_notes_field_map():
    """Return {normalised Notes key -> master column}. Starts from the documented
    default and merges an optional config_files/notes_field_mapping.json override
    (mapping is config-driven, never hardcoded in the loader). Targets that are
    not real master columns are dropped so a typo can't create a bad column."""
    mapping = dict(NOTES_FIELD_MAP_DEFAULT)
    cfg = os.path.join(CONFIG_DIR, "notes_field_mapping.json")
    try:
        if os.path.isfile(cfg):
            with open(cfg, encoding="utf-8") as f:
                override = json.load(f) or {}
            if isinstance(override, dict):
                mapping.update({str(k): str(v) for k, v in override.items()})
    except Exception:
        pass                                    # a bad override never breaks the run
    valid = set(MASTER_COLUMNS)
    return {_normkey(k): v for k, v in mapping.items() if v in valid}


# Built once at import; targets validated against the master schema.
NOTES_FIELD_MAP = load_notes_field_map()

# How each mapped target is standardised after extraction.
_NOTES_YESNO_TARGETS = {"IsReferral", "IsGoogleMeetSchedule", "IsWalkInSchedule"}
_NOTES_DATE_TARGETS = {"Next Follow-Up Date", "IsGoogleMeetScheduleDate",
                       "IsWalkInScheduleDate"}


def parse_notes_kv(notes):
    """Robustly parse a structured 'Key: Value' Notes blob into
    {normalised_key: cleaned_value}. Tolerant of wrapping/embedded single or
    double quotes, extra spaces, blank values, blank & multiple line breaks,
    multi-line values, leading/trailing spaces and inconsistent formatting.
    Malformed lines are skipped, never fatal. Returns {} for empty/plain notes."""
    text = str(notes or "")
    if not text.strip():
        return {}
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    stripped = text.strip().strip('"').strip("'").strip()     # unwrap the whole blob
    out, last_key = {}, None
    for raw_line in stripped.split("\n"):
        line = raw_line.strip().strip('"').strip("'").strip()
        if not line:
            continue
        if ":" in line:
            k, _, v = line.partition(":")
            nk = _normkey(k)
            if not nk:
                continue
            # tolerate double colons (Key::Value) and wrapping quotes/spaces
            v = v.strip().lstrip(":").strip().strip('"').strip("'").strip()
            out[nk] = clean_text(v)
            last_key = nk
        elif last_key is not None:                            # continuation line
            extra = clean_text(line.strip('"').strip("'"))
            if extra:
                out[last_key] = (out[last_key] + " " + extra).strip()
    return out


def map_notes_to_targets(kv):
    """Map parsed Notes {normkey: value} onto master columns via NOTES_FIELD_MAP,
    applying the right Yes/No / date / status / name standardisation. A blank
    value maps to a blank target (never a placeholder)."""
    out = {}
    for nk, val in kv.items():
        col = NOTES_FIELD_MAP.get(nk)
        if not col:
            continue
        if col in _NOTES_YESNO_TARGETS:
            out[col] = std_yes_no(val)
        elif col in _NOTES_DATE_TARGETS:
            out[col] = fmt_date_field(val)
        elif col == "Admission Status":
            out[col] = std_status(val)
        elif col == "Full Name":
            out[col] = clean_lead_name(val)
        else:
            out[col] = val
    return out


# ---------------------------------------------------------------------------
# WEBSITE ENQUIRY  source-header -> master-column mapping.
# Source of truth: the "Lead Tracker Field Mapping Details" doc, column
# "2. Website Enquiry", aligned to the CURRENT IntelliBi_Website_Leads headers.
# The mapping is HEADER-BASED (never positional), so reordering the Website
# columns does not affect loading. Add/rename a mapping WITHOUT touching code by
# dropping config_files/website_field_mapping.json of {"Sheet Header": "Master
# Column"} — merged over these defaults. Dict order = precedence when two source
# headers target the same master field (first non-blank wins). Legacy header
# aliases are included so older exports keep working too.
# ---------------------------------------------------------------------------
WEBSITE_FIELD_MAP_DEFAULT = OrderedDict([
    ("Enquriy Date",             "LeadInitalTimestamp"),   # current header (sic)
    ("Enquiry Date",             "LeadInitalTimestamp"),   # corrected-spelling alias
    ("Date",                     "LeadInitalTimestamp"),   # legacy alias
    ("Timestamp",                "LeadInitalTimestamp"),   # legacy alias
    ("Full Name",                "Full Name"),
    ("Email Address",            "Email Address"),
    ("Mobile Number",            "Mobile Number"),
    ("Current City",             "Current City"),
    ("Highest Qualification",    "Highest Qualification"),
    ("Candidate Type",           "Current Status"),        # doc: Candidate Type -> Current Status
    ("Current Role",             "Current Status"),         # sheet holds status/role text here
    ("Total Experience",         "Total Years of Experience"),
    ("Course Interested In",     "Which technology are you interested in learning?"),
    ("Career Goal",              "What is your primary goal?"),
    ("Preferred Time",           "Preferred Batch Timing"),  # time-slot preference
    ("IsReferral",               "IsReferral"),
    ("Referrer's Name",          "Referrer's Name"),
    ("Counsellor Notes",         "Remarks"),               # doc: Counsellor Notes -> Remarks
    ("Remark",                   "Remarks"),               # legacy alias
    ("IsGoogleMeetSchedule",     "IsGoogleMeetSchedule"),
    ("IsGoogleMeetScheduleDate", "IsGoogleMeetScheduleDate"),
    ("IsWalkInSchedule",         "IsWalkInSchedule"),
    ("IsWalkInScheduleDate",     "IsWalkInScheduleDate"),
    ("AdmissionStatus",          "Lead Status"),           # doc: Website AdmissionStatus -> Lead Status
    ("Counselling By",           "Counselling By"),
    ("call taken by",            "Counselling By"),         # legacy alias
    ("Next Follow-upDate",       "Next Follow-Up Date"),
    ("BackOutReason",            "Backout Reason"),
])


def load_website_field_map():
    """Return OrderedDict {master target -> [source headers, in precedence order]}
    from the documented default merged with an optional
    config_files/website_field_mapping.json override. Targets not in the master
    schema are dropped (a typo can't create a bad column)."""
    ordered = OrderedDict(WEBSITE_FIELD_MAP_DEFAULT)
    cfg = os.path.join(CONFIG_DIR, "website_field_mapping.json")
    try:
        if os.path.isfile(cfg):
            ov = json.load(open(cfg, encoding="utf-8")) or {}
            if isinstance(ov, dict):
                for k, v in ov.items():
                    ordered[str(k)] = str(v)      # override / add
    except Exception:
        pass                                       # bad override never breaks the run
    valid = set(MASTER_COLUMNS)
    target_sources = OrderedDict()
    for src, tgt in ordered.items():
        if tgt not in valid:
            continue
        target_sources.setdefault(tgt, [])
        if src not in target_sources[tgt]:
            target_sources[tgt].append(src)
    return target_sources


# {master target -> [source headers]} built once at import.
WEBSITE_TARGET_SOURCES = load_website_field_map()

# Targets the Website loader fills explicitly (dt parse / phone / name / etc.);
# the generic pass skips these.
_WEBSITE_EXPLICIT = {"LeadInitalTimestamp", "Mobile Number", "Full Name",
                     "Email Address", "Counselling By"}
_WEB_YESNO = {"IsReferral", "IsGoogleMeetSchedule", "IsWalkInSchedule"}
_WEB_DATES = {"IsGoogleMeetScheduleDate", "IsWalkInScheduleDate", "Next Follow-Up Date"}
_WEB_STATUS = {"Lead Status", "Admission Status"}


def _website_source_for(row, target):
    """First non-blank source value mapped to a master target (header-based,
    tolerant to column reordering)."""
    for src in WEBSITE_TARGET_SOURCES.get(target, []):
        v = g(row, src)
        if v:
            return v
    return ""


def _std_target_value(target, val):
    """Standardise a value for its master target (Yes/No / date / status)."""
    if not val:
        return ""
    if target in _WEB_YESNO:
        return std_yes_no(val)
    if target in _WEB_DATES:
        return fmt_date_field(val)
    if target in _WEB_STATUS:
        return std_status(val)
    return val


# ---------------------------------------------------------------------------
# 5b. IntelliBI Lead Information — HIGHEST-PRIORITY source (header-based)
# ---------------------------------------------------------------------------
# Mapping per the "Lead Tracker Field Mapping Details" doc, column
# "5. IntelliBI Lead Information". Keyed master target -> source header (the
# column name in the IntelliBI form backend sheet). One source column may feed
# more than one target (e.g. "Course Interested In"). Position-independent:
# fields are matched by header name, so column reordering never breaks the load.
# An optional config_files/intellibi_field_mapping.json (target -> source
# header) is merged over these defaults for future changes without code edits.
# ---------------------------------------------------------------------------
INTELLIBI_FIELD_MAP_DEFAULT = OrderedDict([
    ("LeadInitalTimestamp",                              "RecordTimeStamp"),
    ("Full Name",                                        "Full Name"),
    ("Email Address",                                    "Email Address"),
    ("Mobile Number",                                    "Mobile Number"),
    ("Current City",                                     "Current City"),
    ("Current Area / Locality",                          "Current Area / Locality"),
    ("Highest Qualification",                            "Highest Qualification"),
    ("Graduation / Passing Year",                        "Graduation / Passing Year"),
    ("Current Status",                                   "Candidate Type"),
    ("Current Company Name",                             "Current Company Name"),
    ("Total Years of Experience",                        "Total Years of Experience"),
    ("Current Domain / Technology",                      "Current Domain / Technology"),
    ("Which technology are you interested in learning?", "Course Interested In"),
    ("Course Advised",                                   "Course Interested In"),
    ("What is your primary goal?",                       "Career Goal"),
    ("IsReferral",                                       "Is Referral"),
    ("Referrer's Name",                                  "Referrer's Name"),
    ("Remarks",                                          "Counsellor Notes"),
    ("IsGoogleMeetSchedule",                             "IsGoogleMeetSchedule"),
    ("IsGoogleMeetScheduleDate",                         "IsGoogleMeetScheduleDate"),
    ("IsWalkInSchedule",                                 "IsWalkInSchedule"),
    ("IsWalkInScheduleDate",                             "IsWalkInScheduleDate"),
    ("Admission Status",                                 "Admission Status"),
    ("Counselling By",                                   "Counselling By"),
    ("Follow-Up Type",                                   "Follow-Up Type"),
    ("Next Follow-Up Date",                              "Next Follow-Up Date"),
    ("Backout Reason",                                   "BackOutReason"),
])


def load_intellibi_field_map():
    """Return OrderedDict {master target -> [source headers]} from the documented
    default merged with an optional config_files/intellibi_field_mapping.json
    (target -> source header). Targets not in the master schema are dropped."""
    ordered = OrderedDict(INTELLIBI_FIELD_MAP_DEFAULT)     # target -> source header
    cfg = os.path.join(CONFIG_DIR, "intellibi_field_mapping.json")
    try:
        if os.path.isfile(cfg):
            ov = json.load(open(cfg, encoding="utf-8")) or {}
            if isinstance(ov, dict):
                for k, v in ov.items():
                    ordered[str(k)] = str(v)              # override / add
    except Exception:
        pass                                              # bad override never breaks the run
    valid = set(MASTER_COLUMNS)
    target_sources = OrderedDict()
    for tgt, src in ordered.items():
        if tgt not in valid:
            continue
        target_sources.setdefault(tgt, [])
        if src not in target_sources[tgt]:
            target_sources[tgt].append(src)
    return target_sources


# {master target -> [source headers]} built once at import.
INTELLIBI_TARGET_SOURCES = load_intellibi_field_map()

# Targets the IntelliBI loader fills explicitly (dt parse / phone / name / etc.);
# the generic pass skips these.
_INTELLIBI_EXPLICIT = {"LeadInitalTimestamp", "Mobile Number", "Full Name",
                       "Email Address", "Counselling By"}


def _intellibi_source_for(row, target):
    """First non-blank source value mapped to a master target (header-based,
    tolerant to column reordering)."""
    for src in INTELLIBI_TARGET_SOURCES.get(target, []):
        v = g(row, src)
        if v:
            return v
    return ""


_PICKED_UP_RE = re.compile(
    r"who\s+(?:picked\s*up|answered|attended|took|received|handled)"
    r"|(?:picked\s*up|answered|attended)\s+the\s+call", re.IGNORECASE)


def extract_call_handler(transcription):
    """From an Exotel call Transcription, return the name of the FIRST person who
    picked up / answered the call, ignoring the bracketed phone number. Generic
    across Exotel transcription formats. Returns '' if none can be extracted.
    Example line: 'Harish Rathod (09022344821), who picked up the call' -> 'Harish Rathod'."""
    if not transcription:
        return ""
    for line in str(transcription).splitlines():          # keep line structure
        if not _PICKED_UP_RE.search(line):
            continue
        seg = re.split(r"\(", line, 1)[0]                  # drop '(phone) ...'
        seg = re.split(r",?\s*who\b", seg, 1, flags=re.IGNORECASE)[0]  # drop ', who ...'
        seg = _PICKED_UP_RE.split(seg, 1)[0]               # drop an inline phrase
        name = clean_text(seg)
        toks = name.split()
        if len(toks) > 4:                                  # strip any leading context
            name = " ".join(toks[-3:])
        if name and any(ch.isalpha() for ch in name) and not name.replace(" ", "").isdigit():
            return name
    return ""


def _is_sushma(name):
    return "sushma kutal" in clean_text(name).lower()


def std_status(value):
    """Standardise Lead Status / Admission Status: strip emoji, tidy spacing,
    Title-case a known controlled vocabulary."""
    s = strip_emoji(clean_text(value))
    if not s:
        return ""
    low = s.lower()
    canon = {
        "hot": "Hot", "warm": "Warm", "cold": "Cold", "casual": "Casual",
        "not interested": "Not Interested",
        "admission confirmed": "Admission Confirmed",
        "follow-up pending": "Follow-up Pending", "followup pending": "Follow-up Pending",
        "interested": "Interested",
    }
    return canon.get(low, s)


# ---------------------------------------------------------------------------
# 4. DATE PARSING / FORMATTING
# ---------------------------------------------------------------------------

OUT_DT_FMT   = "%d-%b-%Y %I:%M %p"     # dd-MMM-yyyy hh:mm a
OUT_DATE_FMT = "%d-%b-%Y"              # dd-MMM-yyyy


def parse_dt(value, dayfirst=False):
    s = clean_text(value)
    if not s or dtparser is None:
        return None
    try:
        return dtparser.parse(s, dayfirst=dayfirst, fuzzy=True)
    except Exception:
        return None


# The team (and every other source: Walk-In / Website / Exotel) reads IST. The
# WhatsApp/Interakt export gives created_at_utc in UTC, so it must be shifted by
# +5:30 to line up. All other sources are already IST and are NOT shifted.
_IST_OFFSET = timedelta(hours=5, minutes=30)


def parse_dt_ist(value, dayfirst=False):
    """Parse a UTC timestamp and return it as IST (naive). A tz-aware value is
    converted to IST; a naive value is assumed UTC and shifted +5:30."""
    dt = parse_dt(value, dayfirst=dayfirst)
    if not dt:
        return None
    if getattr(dt, "tzinfo", None) is not None:
        try:
            return dt.astimezone(timezone(_IST_OFFSET)).replace(tzinfo=None)
        except Exception:
            return dt.replace(tzinfo=None) + _IST_OFFSET
    return dt + _IST_OFFSET


def fmt_dt(dt):
    if not dt:
        return ""
    try:
        return dt.strftime(OUT_DT_FMT).replace(" 0", " ").lstrip("0") \
            if False else dt.strftime(OUT_DT_FMT)
    except Exception:
        return ""


def fmt_date_field(value, dayfirst=False):
    """Standardise an arbitrary date-ish field value to dd-MMM-yyyy hh:mm a
    (falls back to cleaned original text if it cannot be parsed)."""
    raw = clean_text(value)
    if not raw:
        return ""
    dt = parse_dt(raw, dayfirst=dayfirst)
    return fmt_dt(dt) if dt else raw


# ---------------------------------------------------------------------------
# 5. SOURCE LOADERS  ->  normalised "interaction" records
# ---------------------------------------------------------------------------
#
# Each interaction dict carries:
#   _source, _ts (datetime|None), _name_disp, _name_key, _email, _phone (norm),
#   _counsel (source-specific Counselling-By value), plus target-field values.
# ---------------------------------------------------------------------------

def _blank_targets():
    return {f: "" for f in TARGET_FIELDS}


def _coalesce_duplicate_columns(df):
    """Merge duplicate-named columns into ONE, keeping the first non-empty value
    per row.

    A source sheet/export sometimes carries the SAME header twice (e.g. a form
    question that was duplicated, or an extra hidden column). pandas then makes
    `df[col]` / `row[col]` return a *Series* instead of a scalar, and stringifying
    that Series produced garbage values such as
    'Total Experience Total Experience Name: 57, dtype: str' in the output. Folding
    duplicates down to a single clean column at load time fixes this at the root —
    for EVERY field and every future lead, not just Total Years of Experience."""
    if df is None or getattr(df, "empty", True):
        return df
    if df.columns.is_unique:
        return df

    def _first_non_empty(values):
        for v in values:
            sv = "" if v is None else str(v).strip()
            if sv and sv.lower() != "nan":
                return v
        return ""

    merged, order = {}, []
    for name in df.columns:
        if name in merged:
            continue
        order.append(name)
        block = df.loc[:, df.columns == name]
        merged[name] = (block.iloc[:, 0] if block.shape[1] == 1
                        else block.apply(_first_non_empty, axis=1))
    dup = [n for n in order if (df.columns == n).sum() > 1]
    if dup:
        print(f"  [load] coalesced duplicate column(s): {', '.join(map(str, dup))}")
    return pd.DataFrame({n: merged[n] for n in order}, index=df.index)


def load_dataframe(source_key):
    """Load one source into a DataFrame (all columns as strings)."""
    if INPUT_MODE == "gsheet":
        df = _read_gsheet_df(SOURCE_SHEETS[source_key]["id"])
    else:
        path = os.path.join(LOCAL_DIR, SOURCE_SHEETS[source_key]["csv"])
        df = pd.read_csv(path, dtype=str, keep_default_na=False,
                         on_bad_lines="skip", engine="python").fillna("")
    # Never let duplicate-named columns leak downstream as a Series (root cause of
    # the corrupted 'Total Years of Experience' values).
    return _coalesce_duplicate_columns(df)


def g(row, *names):
    """Fetch first present, cleaned column value from a row (tolerant to header
    whitespace / minor naming differences)."""
    for n in names:
        if n in row and clean_text(row[n]):
            return clean_text(row[n])
    # tolerant match ignoring surrounding whitespace in header
    for n in names:
        for col in row.index:
            if col.strip() == n.strip() and clean_text(row[col]):
                return clean_text(row[col])
    return ""


def load_walkin():
    df = load_dataframe("Walk-In")
    out = []
    for _, r in df.iterrows():
        rec = _blank_targets()
        ts_raw = g(r, "Timestamp")
        dt = parse_dt(ts_raw, dayfirst=False)          # M/D/YYYY
        name  = clean_lead_name(g(r, "Full Name"))
        email = clean_email(g(r, "Email Address"))
        phone = norm_phone(g(r, "Mobile Number"))
        rec.update({
            "LeadInitalTimestamp": fmt_dt(dt) if dt else ts_raw,
            "Full Name": name,
            "Email Address": email,
            "Mobile Number": display_phone(phone) or g(r, "Mobile Number"),
            "Current City": g(r, "Current City"),
            "Current Area / Locality": g(r, "Current Area / Locality"),
            "Preferred Contact Method": g(r, "Preferred Contact Method"),
            "Highest Qualification": g(r, "Highest Qualification"),
            "Graduation / Passing Year": g(r, "Graduation / Passing Year"),
            "College / University Name": g(r, "College / University Name"),
            "Grade / Percentage / CGPA": g(r, "Grade / Percentage / CGPA"),
            "Current Status": g(r, "Current Status"),
            "Current Company Name": g(r, "Current Company Name"),
            "Total Years of Experience": g(r, "Total Years of Experience"),
            "Current Domain / Technology": g(r, "Current Domain / Technology"),
            "Which technology are you interested in learning?":
                g(r, "Which technology are you interested in learning?"),
            "Course Advised": g(r, "Course Advised", "Course Interested In",
                                "Course Interested"),
            "What is your primary goal?": g(r, "What is your primary goal?"),
            "How did you hear about IntelliBI?": g(r, "How did you hear about IntelliBI?"),
            "IsReferral": std_yes_no(g(r, "IsReferral", "Is Referral")),
            "Referrer's Name": g(r, "Referrer's Name"),
            "Preferred Learning Mode": g(r, "Preferred Learning Mode"),
            "Preferred Batch Timing": g(r, "Preferred Batch Timing"),
            "When are you planning to take admission?":
                g(r, "When are you planning to take admission?"),
            "Remarks": g(r, "Remarks"),
            "Lead Status": std_status(g(r, "Lead Status")),
            "Admission Status": std_status(g(r, "Admission Status")),
            "Backout Reason": g(r, "BackOutReason"),
            "Counselling By": clean_counsellor(g(r, "Counselling By")),
            "ScheduledOrDirectWalkIn": g(r, "ScheduledOrDirectWalkIn"),
            "Next Follow-Up Date": fmt_date_field(g(r, "Next Follow-Up Date"), dayfirst=False),
        })
        counsel = clean_counsellor(g(r, "Counselling By"))
        yield _finalise(rec, "Walk-In", dt, name, email, phone, counsel)


def load_website():
    """Website Enquiry loader — fully header-based and config-driven.

    Every source field is resolved by its *column header name* via
    WEBSITE_TARGET_SOURCES (documented default merged with an optional
    config_files/website_field_mapping.json). No fixed column positions or
    indexes are used, so the sheet can be reordered freely without breaking
    the load. See WEBSITE_FIELD_MAP_DEFAULT for the documented mapping."""
    df = load_dataframe("Website")
    for _, r in df.iterrows():
        rec = _blank_targets()

        # Generic pass: every mapped target except the ones handled explicitly
        # below (timestamp / name / email / phone / counsellor).
        for target in WEBSITE_TARGET_SOURCES:
            if target in _WEBSITE_EXPLICIT:
                continue
            val = _website_source_for(r, target)
            if val:
                rec[target] = _std_target_value(target, val)

        # LeadInitalTimestamp — parse the enquiry date (DD/MM/YYYY first).
        ts_raw = _website_source_for(r, "LeadInitalTimestamp")
        dt = parse_dt(ts_raw, dayfirst=True)
        rec["LeadInitalTimestamp"] = fmt_dt(dt) if dt else ts_raw

        # Name.
        name = clean_lead_name(_website_source_for(r, "Full Name"))
        rec["Full Name"] = name

        # Email + phone (column-shift tolerant recovery kept for safety).
        email_field = _website_source_for(r, "Email Address")
        email = clean_email(email_field)
        rec["Email Address"] = email
        phone_raw = _website_source_for(r, "Mobile Number")
        phone = norm_phone(phone_raw)
        if not is_valid_mobile(phone):
            for cand in (email_field, phone_raw):
                p2 = norm_phone(cand)
                if is_valid_mobile(p2):
                    phone = p2
                    break
        rec["Mobile Number"] = display_phone(phone) or phone_raw

        # Counsellor.
        counsel = clean_counsellor(_website_source_for(r, "Counselling By"))
        rec["Counselling By"] = counsel

        yield _finalise(rec, "Website", dt, name, email, phone, counsel)


def load_intellibi():
    """IntelliBI Lead Information loader — the HIGHEST-priority source.

    Fully header-based and config-driven, mapped per the 'Lead Tracker Field
    Mapping Details' document, column '5. IntelliBI Lead Information'. Every
    field is resolved by its column header name via INTELLIBI_TARGET_SOURCES, so
    the source sheet can be reordered without breaking the load. Its timestamp
    (RecordTimeStamp) is already IST, so no timezone shift is applied."""
    df = load_dataframe("IntelliBI")
    for _, r in df.iterrows():
        rec = _blank_targets()

        # Generic pass: every mapped target except the explicit ones below.
        for target in INTELLIBI_TARGET_SOURCES:
            if target in _INTELLIBI_EXPLICIT:
                continue
            val = _intellibi_source_for(r, target)
            if val:
                rec[target] = _std_target_value(target, val)

        # LeadInitalTimestamp <- RecordTimeStamp (e.g. "06-Aug-2026 17:27:33", IST).
        ts_raw = _intellibi_source_for(r, "LeadInitalTimestamp")
        dt = parse_dt(ts_raw, dayfirst=True)
        rec["LeadInitalTimestamp"] = fmt_dt(dt) if dt else ts_raw

        # Name.
        name = clean_lead_name(_intellibi_source_for(r, "Full Name"))
        rec["Full Name"] = name

        # Email + phone.
        email = clean_email(_intellibi_source_for(r, "Email Address"))
        rec["Email Address"] = email
        phone_raw = _intellibi_source_for(r, "Mobile Number")
        phone = norm_phone(phone_raw)
        rec["Mobile Number"] = display_phone(phone) or phone_raw

        # Counsellor.
        counsel = clean_counsellor(_intellibi_source_for(r, "Counselling By"))
        rec["Counselling By"] = counsel

        yield _finalise(rec, "IntelliBI", dt, name, email, phone, counsel)


def load_whatsapp():
    df = load_dataframe("WhatsApp")
    for _, r in df.iterrows():
        rec = _blank_targets()
        ts_raw = g(r, "created_at_utc")
        dt = parse_dt_ist(ts_raw)          # created_at_utc is UTC -> shift to IST
        name  = clean_lead_name(g(r, "trait_name") or g(r, "trait_lead_name"))
        phone = norm_phone(g(r, "phone_number"))
        remarks = g(r, "Counsellor Notes") or g(r, "enr_note_latest")
        # Counselling By (WhatsApp): COALESCE(Conversation Label, Assigned Agent,
        # enr_contact_owner). Conversation Label 'assign to X' maps to the counsellor.
        counsel = (map_conversation_label(g(r, "Conversation Label"))
                   or clean_counsellor(g(r, "Assigned Agent"))
                   or clean_counsellor(g(r, "enr_contact_owner")))
        rec.update({
            "LeadInitalTimestamp": fmt_dt(dt) if dt else ts_raw,
            "Full Name": name,
            "Mobile Number": display_phone(phone) or g(r, "phone_number"),
            "Current City": g(r, "Current Location"),
            "Current Status": g(r, "Candidate Type"),
            "Current Company Name": g(r, "Current Company"),
            "Total Years of Experience": g(r, "Total Experience"),
            "Current Domain / Technology": g(r, "Current Role"),
            "Which technology are you interested in learning?": g(r, "Course Interested In"),
            "Course Advised": g(r, "Course Advised"),
            "What is your primary goal?": g(r, "Career Goal / Requirement"),
            "IsReferral": std_yes_no(g(r, "IsReferral")),
            "Referrer's Name": g(r, "Referrer's Name"),
            "Remarks": remarks,
            "IsGoogleMeetSchedule": std_yes_no(g(r, "Google Meet Scheduled")),
            "IsGoogleMeetScheduleDate": fmt_date_field(g(r, "Google Meet Date & Time")),
            "IsWalkInSchedule": std_yes_no(g(r, "Walk-in Scheduled")),
            "IsWalkInScheduleDate": fmt_date_field(g(r, "Walk-in Date & Time")),
            "Admission Status": std_status(g(r, "Admission Status")),
            "Backout Reason": g(r, "BackOutReason"),
            "Counselling By": counsel,
            "Next Follow-Up Date": fmt_date_field(g(r, "Next Follow-up Date")),
        })
        yield _finalise(rec, "WhatsApp", dt, name, None, phone, counsel)


def load_calling(business_numbers):
    df = load_dataframe("Call")
    # Build the business / virtual-line exclusion set WITHOUT swallowing real
    # customers. A company line is: the PhoneNumber column (always an ExoPhone),
    # the 'To' number on INBOUND calls (the line the lead dialed), and the 'From'
    # number on OUTBOUND calls (the line the agent dialed out on). Crucially we do
    # NOT add inbound-'From' or outbound-'To' — those ARE the leads. (Previously
    # every 'To' was blacklisted, so any customer who was later called back showed
    # up as an outbound 'To' and got wrongly excluded.)
    for _, r in df.iterrows():
        outbound = g(r, "Direction").lower().startswith("outbound")
        for v in (norm_phone(g(r, "PhoneNumber")),
                  norm_phone(g(r, "From")) if outbound else norm_phone(g(r, "To"))):
            if v:
                business_numbers.add(v)

    # Pass 2: for each external number, collect every NON-'Sushma Kutal' handler
    # seen across all of that number's call rows (ToName / Assign To / who picked
    # up in the Transcription). Used to resolve the counsellor when a row's own
    # ToName & Assign To only give 'Sushma Kutal'.
    number_handlers = defaultdict(list)
    for _, r in df.iterrows():
        outbound = g(r, "Direction").lower().startswith("outbound")
        ext = norm_phone(g(r, "To") if outbound else g(r, "From"))
        if not ext or ext in business_numbers:
            continue
        for cand in (clean_counsellor(g(r, "ToName")),
                     clean_counsellor(g(r, "Assign To")),
                     clean_counsellor(extract_call_handler(str(r.get("Transcription", "") or "")))):
            if cand and not _is_sushma(cand) and cand not in number_handlers[ext]:
                number_handlers[ext].append(cand)

    for _, r in df.iterrows():
        direction = g(r, "Direction").lower()
        # external party = From on inbound, To on outbound
        if direction.startswith("outbound"):
            ext_raw = g(r, "To")
        else:
            ext_raw = g(r, "From")
        phone = norm_phone(ext_raw)
        if not phone or phone in business_numbers:
            continue                                    # skip internal / business legs
        ts_raw = g(r, "StartTime") or g(r, "DateCreated")
        dt = parse_dt(ts_raw)
        raw_name = g(r, "FromName")
        name = clean_lead_name(raw_name)
        # Counselling By (Call): COALESCE(ToName, Assign To) but NEVER 'Sushma Kutal'.
        _ton = clean_counsellor(g(r, "ToName"))
        _ato = clean_counsellor(g(r, "Assign To"))
        if _ton and not _is_sushma(_ton):
            counsel = _ton                                   # ToName is a real (non-Sushma) name
        elif _ato and not _is_sushma(_ato):
            counsel = _ato                                   # Assign To is a real (non-Sushma) name
        else:
            # both are 'Sushma Kutal' (or blank): find who actually handled it —
            # first from THIS call's Transcription, then from any other call of the
            # same number, else fall back to the existing value (may be Sushma).
            counsel = clean_counsellor(extract_call_handler(str(r.get("Transcription", "") or "")))
            if _is_sushma(counsel):
                counsel = ""
            if not counsel:
                counsel = next(iter(number_handlers.get(phone, [])), "")
            if not counsel:
                counsel = _ton or _ato
        # --- Parse the structured Notes blob and map each key to its target
        #     column (per the Lead Tracker Field Mapping) instead of dumping the
        #     whole blob into a single field. -------------------------------
        raw_notes = str(r.get("Notes", "") or "")
        kv = parse_notes_kv(raw_notes)
        mapped = map_notes_to_targets(kv)

        # Full Name (Call) = COALESCE(FromName, Notes:Full Name).
        if not name:
            name = clean_lead_name(mapped.get("Full Name", ""))

        # Remarks = the free-text "Counsellor Notes" value. If the Notes blob is
        # NOT structured (nothing parsed), keep the whole note so free text isn't
        # lost; also preserve a status word typed into FromName for the flags.
        remarks = mapped.get("Remarks", "")
        if not kv:
            remarks = g(r, "Notes")
        if not remarks and raw_name and is_status_name(raw_name):
            remarks = clean_name(raw_name)

        rec = _blank_targets()
        # Populate every mapped Notes field (blank value -> column stays blank).
        for col, val in mapped.items():
            if val:
                rec[col] = val
        # Fields from the call row / computed above take precedence over Notes.
        rec.update({
            "LeadInitalTimestamp": fmt_dt(dt) if dt else ts_raw,
            "Full Name": name,
            "Mobile Number": display_phone(phone),
            "Remarks": remarks,
            "Lead Status": std_status(g(r, "Lead Status")),
            "Counselling By": counsel,
        })
        yield _finalise(rec, "Call", dt, name, None, phone, counsel)


def _finalise(rec, source, dt, name, email, phone, counsel):
    rec["_source"]    = source
    rec["_ts"]        = dt
    rec["_name_disp"] = name or ""
    rec["_name_key"]  = name_key(name or "")
    rec["_email"]     = email or ""
    # Identity key accepts ANY normalised number with >= 10 digits (mobiles,
    # landlines, non-6-9 numbers), so genuine leads/corrections are never
    # silently dropped for failing the strict mobile pattern.
    rec["_phone"]     = phone if (phone and len(phone) >= 10) else ""
    rec["_phone_any"] = phone or ""                    # any digits (for display fallback)
    rec["_counsel"]   = counsel or ""
    return rec


# ---------------------------------------------------------------------------
# 6. CLUSTERING  (dedupe:  Phone > Name > Email)
# ---------------------------------------------------------------------------

class Stats:
    def __init__(self):
        self.processed = defaultdict(int)      # per-source usable records
        self.ignored_empty = 0
        self.match_phone = 0
        self.match_name = 0
        self.match_email = 0
        self.new_leads = 0
        self.failed_match = 0                  # record with no phone/name/email key


def cluster_interactions(interactions, stats):
    """Group interactions into unique leads. Duplicate matching uses ONLY
    Phone (priority 1) then Email (priority 2). Full Name is NOT used for
    matching (records are never merged or split on the basis of a name)."""
    phone2c, email2c = {}, {}
    clusters = []                              # list[list[record]]

    for rec in interactions:
        ph, em, nk = rec["_phone"], rec["_email"], rec["_name_key"]
        # keep name-bearing records (loaded as standalone leads), drop only rows
        # with no phone, no email AND no name at all
        if not ph and not em and not nk:
            stats.ignored_empty += 1           # "Ignore Empty Records"
            continue

        cid = None
        if ph and ph in phone2c:
            cid = phone2c[ph];  stats.match_phone += 1
        elif em and em in email2c:
            cid = email2c[em];  stats.match_email += 1

        if cid is None:
            cid = len(clusters)
            clusters.append([])
            stats.new_leads += 1

        clusters[cid].append(rec)
        if ph and ph not in phone2c:  phone2c[ph] = cid
        if em and em not in email2c:  email2c[em] = cid

    return clusters


# ---------------------------------------------------------------------------
# 7. MERGE cluster -> master row
# ---------------------------------------------------------------------------

_EPOCH = datetime(1970, 1, 1)


def _ts_key(dt):
    """Windows-safe 'seconds since epoch' for sorting by recency.
    datetime.timestamp() raises OSError [Errno 22] on Windows for out-of-range
    dates (pre-1970 or an extreme year) that a malformed source value can parse
    into. This never raises and keeps the same ordering."""
    if not dt:
        return 0.0
    try:
        d = dt.replace(tzinfo=None) if getattr(dt, "tzinfo", None) else dt
        return (d - _EPOCH).total_seconds()
    except Exception:
        return 0.0


def first_non_blank(records, field):
    for r in sorted(records, key=lambda x: (SOURCE_MERGE_RANK[x["_source"]],
                                            -_ts_key(x["_ts"]))):
        v = clean_text(r.get(field, ""))
        if v:
            return v
    return ""


def latest_remark(records):
    """Return the LATEST non-blank Remark across a customer's interactions.

    Chosen by interaction timestamp (newest first) so a newer remark from ANY
    source — e.g. a fresh Walk-In note — is never overwritten by an older remark
    from a higher-priority source. Records are already the same customer
    (phone/email de-duplicated). Source precedence is used ONLY as a tie-breaker
    when timestamps are equal or missing, keeping the result deterministic."""
    best_key, best_val = None, ""
    for r in records:
        v = clean_text(r.get("Remarks", ""))
        if not v:
            continue
        # newest ts wins; on a tie, the higher-priority (lower-rank) source wins
        key = (_ts_key(r["_ts"]), -SOURCE_MERGE_RANK[r["_source"]])
        if best_key is None or key > best_key:
            best_key, best_val = key, v
    return best_val


def counselling_by(records):
    """Counselling By priority: Call > WhatsApp > Walk-In > Website."""
    for src in COUNSELLING_PRIORITY:
        recs = [r for r in records if r["_source"] == src and r["_counsel"]]
        recs.sort(key=lambda x: -_ts_key(x["_ts"]))
        if recs:
            return recs[0]["_counsel"]
    return ""


def remark_flags_invalid(remark):
    """Phone-validation exception: a remark containing the word 'invalid'
    (case-insensitive, anywhere in the text) forces IsPhoneNumberValid = No."""
    return "invalid" in clean_text(remark).lower()


# common misspellings / forms of "irrelevant" (all start with 'i')
_IRRELEVANT_PREFIXES = ("irrelev", "irrelav", "irreval", "irrevel", "irelev", "irrelvant")


def remark_flags_irrelevant(remark):
    """Fuzzy check for an 'irrelevant' lead in the remark. Case-insensitive,
    punctuation/space-insensitive, tolerant of spelling variants (irrelevent,
    irrevelant, irrevalant, irelevant, ...) and the phrase 'not relevant'.
    Requires an irrelevant-form token (starts with 'i') so the positive word
    'relevant' is NOT matched, minimising false positives."""
    s = clean_text(remark).lower()
    if not s:
        return False
    s = re.sub(r"[^a-z\s]", " ", s)
    s = _WS_RE.sub(" ", s).strip()
    if not s:
        return False
    if "not relevant" in s or "non relevant" in s or "not a relevant" in s:
        return True
    for tok in s.split():
        if len(tok) < 6 or tok[0] != "i":
            continue
        if tok.startswith(_IRRELEVANT_PREFIXES):
            return True
        if difflib.SequenceMatcher(None, tok, "irrelevant").ratio() >= 0.8:
            return True
    return False


# Exact (mis)spelling variants of 'irrelevant' that must force IsLeadRelevant=No
# when they appear anywhere in Remarks. This is an ADDITIONAL check ORed alongside
# the existing relevance logic — it never replaces or weakens it.
_REMARK_IRRELEVANT_WORDS = (
    "irrellevant", "irelevant", "irelevent", "irellevant", "irrelevnt",
    "irrelevan", "irrelevantt", "irrevelant", "irrelavent", "irrelevent",
    "irrlvent",
)


def remark_has_irrelevant_keyword(remark):
    """True if Remarks contains any configured 'irrelevant' keyword variant, after
    full normalisation: trim, collapse extra/edge whitespace, strip punctuation,
    and lower-case (so case, spacing and punctuation differences never block a
    match). Blank / null / missing Remarks return False. The keyword is detected
    even when it appears inside a longer remark (substring match)."""
    s = clean_text(remark).lower()
    if not s:
        return False
    s = re.sub(r"[^a-z]+", " ", s)          # punctuation/digits -> single space
    s = _WS_RE.sub(" ", s).strip()
    if not s:
        return False
    return any(w in s for w in _REMARK_IRRELEVANT_WORDS)


def build_history(records):
    """Chronological enquiry journey. Collapses CONSECUTIVE duplicate
    interactions (same Lead Source + Lead Date & Time + Counselling By), then
    renumbers sequentially. Genuine repeats at different times/platforms are
    preserved."""
    ordered = sorted(records, key=lambda x: (x["_ts"] or datetime.max))
    entries = []                                   # list of (src, date, cby)
    for r in ordered:
        src  = SOURCE_LABEL[r["_source"]]
        date = fmt_dt(r["_ts"]) if r["_ts"] else clean_text(r.get("LeadInitalTimestamp", ""))
        cby  = r["_counsel"] or "-"
        entry = (src, date, cby)
        if entries and entries[-1] == entry:       # consecutive duplicate -> skip
            continue
        entries.append(entry)

    nums  = [str(i) for i in range(1, len(entries) + 1)]
    srcs  = [e[0] for e in entries]
    dates = [e[1] for e in entries]
    cbys  = [e[2] for e in entries]
    block = (
        "Lead Enq Num: " + ", ".join(nums) + "\n" +
        "Lead Source: " + ", ".join(srcs) + "\n" +
        "Lead Date: " + ", ".join(dates) + "\n" +
        "Counselling By: " + ", ".join(cbys)
    )
    return block, ordered, len(entries)


# ---------------------------------------------------------------------------
# 7b. LEAD TYPE — derived from the live 'Lead Type/Current Status Value Mapping'
#     configuration sheet (see LEAD_TYPE_MAP_SHEET_ID). Never hard-coded.
# ---------------------------------------------------------------------------

def _norm_status(value):
    """Normalise a Current Status / mapping value for reliable matching: unicode
    NFKC, strip invisibles, collapse internal whitespace, trim, and lower-case.
    (clean_text already does everything except the case fold.)"""
    return clean_text(value).lower()


def load_lead_type_map():
    """Read the mapping sheet (wide: header = Lead Type, cells = Current Status
    values) and return (exact_map, ordered_pairs):
      exact_map     : {normalised value -> Lead Type}  (first column wins on ties)
      ordered_pairs : [(normalised value, Lead Type)]  in sheet order, for the
                      'contains' pass.
    The Lead Type label is whitespace-normalised (so a header typed as
    'Fresher -  Passed Out' is emitted as 'Fresher - Passed Out'). Read live every
    run so future edits to the sheet take effect automatically. On any failure the
    map is empty (every lead becomes 'Unidentified') and the run continues."""
    exact, pairs = {}, []
    try:
        df = _read_gsheet_df(LEAD_TYPE_MAP_SHEET_ID)
    except Exception as e:
        print(f"  [Lead Type] WARN could not read mapping sheet "
              f"{LEAD_TYPE_MAP_SHEET_ID}: {e}\n"
              f"             -> every lead will be '{LEAD_TYPE_UNIDENTIFIED}'. "
              f"Share the sheet with the service account to enable mapping.")
        return exact, pairs
    if df is None or getattr(df, "empty", True):
        print("  [Lead Type] WARN mapping sheet is empty — every lead will be "
              f"'{LEAD_TYPE_UNIDENTIFIED}'.")
        return exact, pairs
    for col in df.columns:
        lead_type = clean_text(col)          # collapses stray double-spaces
        if not lead_type:
            continue
        for cell in df[col].tolist():
            val = clean_text(cell)
            if not val:
                continue
            key = _norm_status(val)
            if not key:
                continue
            if key not in exact:             # first (left-most) category wins ties
                exact[key] = lead_type
            pairs.append((key, lead_type))
    print(f"  [Lead Type] loaded {len(pairs)} mapping value(s) across "
          f"{len({lt for _, lt in pairs})} categor(y/ies)")
    return exact, pairs


_LEAD_TYPE_MAP = None          # lazy singleton: (exact_map, ordered_pairs)


def _lead_type_map():
    global _LEAD_TYPE_MAP
    if _LEAD_TYPE_MAP is None:
        _LEAD_TYPE_MAP = load_lead_type_map()
    return _LEAD_TYPE_MAP


def classify_lead_type(current_status):
    """Map a Current Status value to its Lead Type using the config sheet:
      1) exact match (normalised) against every mapping value;
      2) else 'contains' — the first mapping value (in sheet order) that appears
         inside the Current Status text (values < 3 chars are skipped in this pass
         to avoid noisy substring hits);
      3) else 'Unidentified'. A blank Current Status is 'Unidentified'."""
    s = _norm_status(current_status)
    if not s:
        return LEAD_TYPE_UNIDENTIFIED
    exact, pairs = _lead_type_map()
    if s in exact:                                   # Step 1 — exact
        return exact[s]
    for val, lead_type in pairs:                     # Step 2 — contains
        if len(val) >= 3 and val in s:
            return lead_type
    return LEAD_TYPE_UNIDENTIFIED                     # Step 3 — no match


# ── Course-interest normalization ─────────────────────────────────────────────
# Maps raw "Which technology are you interested in learning?" values (spellings /
# abbreviations / course codes) to the approved STANDARD course names. Applied to
# THIS field only. Matching is case-insensitive and space/punctuation-insensitive,
# with the Gen-AI family (Generative AI / Gen AI / GenAI) collapsed so spacing
# variants match. Values not listed are left cleaned but otherwise unchanged.
COURSE_INTEREST_FIELD = "Which technology are you interested in learning?"
_COURSE_STD = {
    "Azure Data Engineering": [
        "ADE", "Azure Data Engineering with Databricks", "AI 200",
        "azure developer", "Azure cloud developer",
    ],
    "Azure Data Engineering with GenAI": [
        "Data Engineer with Gen AI Course", "Data Engineer with Gen AI",
    ],
    "Data Analytics": [
        "Data Analysis",
    ],
    "Data Analytics with GenAI": [
        "Data Analytics with AI", "DAAI", "Data Analytics with Gen AI",
    ],
    "Advanced Artificial Intelligence (AI) with Generative AI + Agentic AI + Machine Learning(ML)": [
        "Advanced Artificial Intelligence Course", "Advanced Artificial Intelligence",
        "Gen AI", "GEN AI", "GenAI", "Data science and AI ML",
    ],
}


def _course_key(value):
    """Normalized match key: lower-cased, Gen-AI family unified, and stripped of
    spaces/punctuation so 'gen ai', 'Gen AI' and 'GenAI' all collapse together."""
    k = re.sub(r"\s+", " ", clean_text(value).lower()).strip()
    k = re.sub(r"\bgenerative\s*ai\b", "genai", k)
    k = re.sub(r"\bgen\s*ai\b", "genai", k)
    return re.sub(r"[^a-z0-9]", "", k)


_COURSE_LOOKUP = {}
for _std, _srcs in _COURSE_STD.items():
    for _sv in _srcs:
        _COURSE_LOOKUP[_course_key(_sv)] = _std


def normalize_course_interest(value):
    """Map a raw course-interest value to its approved standard name. Unlisted
    values are returned cleaned but unchanged; blanks stay blank."""
    v = clean_text(value)
    if not v:
        return v
    return _COURSE_LOOKUP.get(_course_key(v), v)


# Columns whose consolidated value is OVERWRITTEN by the customer's latest
# Walk-In record once they visit the office (their Walk-In form is the most
# reliable, up-to-date info). Applied only to these columns; every other column
# keeps its existing merge. A blank Walk-In value never overwrites an existing
# one (see latest_walkin_value + the guard in merge_cluster).
WALKIN_OVERWRITE_FIELDS = [
    "Full Name",
    "Email Address",
    "Current City",
    "Current Area / Locality",
    "Preferred Contact Method",
    "Highest Qualification",
    "Graduation / Passing Year",
    "College / University Name",
    "Grade / Percentage / CGPA",
    "Current Status",
    "Current Company Name",
    "Total Years of Experience",
    "Current Domain / Technology",
    "Which technology are you interested in learning?",
    "Course Advised",
    "What is your primary goal?",
    "How did you hear about IntelliBI?",
    "IsReferral",
    "Referrer's Name",
    "When are you planning to take admission?",
    "Lead Status",
]


def latest_walkin_value(records, field):
    """Latest NON-BLANK value of `field` across the customer's Walk-In records,
    newest interaction first (by Timestamp; source precedence is irrelevant here
    because every record is a Walk-In). Returns '' when there is no Walk-In record
    or none of them carries a value — so a blank Walk-In never overwrites an
    existing consolidated value."""
    walkins = sorted((r for r in records if r["_source"] == "Walk-In"),
                     key=lambda x: -_ts_key(x["_ts"]))
    for r in walkins:
        v = clean_text(r.get(field, ""))
        if v:
            return v
    return ""


def merge_cluster(records):
    row = OrderedDict((c, "") for c in MASTER_COLUMNS)

    # source-priority field fill
    for f in TARGET_FIELDS:
        row[f] = first_non_blank(records, f)

    # Walk-In office-visit override: once a customer visits the office, their
    # LATEST Walk-In record is the most reliable/current information. For the
    # specified columns, prefer the latest NON-BLANK Walk-In value over the
    # source-priority merge above. Done BEFORE the IsReferral / Lead Type
    # derivations below so those (which read the referral inputs / Current Status)
    # reflect the Walk-In-primary values. A blank Walk-In value never overwrites.
    if any(r["_source"] == "Walk-In" for r in records):
        for f in WALKIN_OVERWRITE_FIELDS:
            wv = latest_walkin_value(records, f)
            if wv:
                row[f] = wv

    # Normalize the course fields to the approved standard course names, so the
    # consolidated master is the single normalized source every report reads.
    # BOTH course columns are mapped: "Which technology are you interested in
    # learning?" AND "Course Advised" (the reports fall back to Course Advised
    # when the interest field is blank, so a raw value like 'ADE' there would
    # otherwise still surface). Applied to these two fields only.
    row[COURSE_INTEREST_FIELD] = normalize_course_interest(
        row.get(COURSE_INTEREST_FIELD, ""))
    row["Course Advised"] = normalize_course_interest(row.get("Course Advised", ""))

    # Remarks — always the LATEST remark across the customer's interactions.
    # first_non_blank() above picks by SOURCE precedence, which can return an
    # older higher-priority-source remark instead of a newer one (e.g. a fresh
    # Walk-In note). Override it with the newest non-blank remark by interaction
    # timestamp (source precedence is only a tie-breaker). Nothing else changes.
    row["Remarks"] = latest_remark(records)

    # IsReferral — standardised Yes / No, never blank. A lead is a referral when
    # ANY of these hold (existing logic + Walk-In conditions), all evaluated on
    # the merged row (first non-blank across every phone/email-deduped source):
    #   * IsReferral / "Is Referral" = Yes   (existing flag, both spellings)
    #   * a Referrer's Name is present       (existing logic; the Walk-In form
    #                                         carries this but no IsReferral col)
    #   * "How did you hear about IntelliBI?" = "Friend / Referral"
    # Comparisons are case-insensitive and whitespace-trimmed. Anything else = No.
    _isref = std_yes_no(row.get("IsReferral", ""))
    _refname = clean_text(row.get("Referrer's Name", ""))
    _heard_key = re.sub(r"\s+", "", clean_text(
        row.get("How did you hear about IntelliBI?", "")).lower())
    _is_referral = (_isref.lower() == "yes"
                    or bool(_refname)
                    or _heard_key == "friend/referral")
    row["IsReferral"] = "Yes" if _is_referral else "No"

    # Lead Type — derived from the finalised Current Status via the live config
    # sheet mapping (exact -> contains -> Unidentified). Never sourced from a lead.
    row["Lead Type"] = classify_lead_type(row.get("Current Status", ""))
    # Additional override — applied ONLY when the mapping above did not yield a real
    # category, i.e. the current Lead Type is blank / null / "Unidentified"
    # (case-insensitive, after cleaning). In that case, a lead whose Admission Status
    # is "Unable to Connect" (case/space-insensitive) becomes "Unreachable". A lead
    # that already has a real Lead Type keeps it — the existing mapping is untouched.
    if clean_text(row.get("Lead Type", "")).lower() in ("", "unidentified"):
        if clean_text(row.get("Admission Status", "")).lower() == "unable to connect":
            row["Lead Type"] = "Unreachable"

    # dedicated Counselling By logic overrides the plain merge
    row["Counselling By"] = counselling_by(records)

    # timestamps
    ts_list = [r["_ts"] for r in records if r["_ts"]]
    first_dt = min(ts_list) if ts_list else None
    last_dt  = max(ts_list) if ts_list else None
    if first_dt:
        row["LeadInitalTimestamp"] = fmt_dt(first_dt)

    # ensure a display phone / name even if merge picked a blank-normalised one
    if not row["Mobile Number"]:
        for r in records:
            if r["_phone_any"]:
                row["Mobile Number"] = r["_phone_any"]; break
    if not row["Full Name"]:
        for r in sorted(records, key=lambda x: SOURCE_MERGE_RANK[x["_source"]]):
            if r["_name_disp"]:
                row["Full Name"] = r["_name_disp"]; break

    # source tracking flags
    present = {r["_source"] for r in records}
    row["IsWalk-In"]  = "Yes" if "Walk-In"  in present else "No"
    row["IsWebsite"]  = "Yes" if "Website"   in present else "No"
    row["IsWhatsapp"] = "Yes" if "WhatsApp"  in present else "No"
    row["IsCall"]     = "Yes" if "Call"      in present else "No"

    # record-quality flags (recomputed every run; records are still kept)
    _remark = row.get("Remarks", "")
    if remark_flags_invalid(_remark):
        row[VALIDATION_FIELD] = "No"      # remark already marks the number invalid
    else:
        row[VALIDATION_FIELD] = "Yes" if is_valid_mobile(norm_phone(row.get("Mobile Number", ""))) else "No"
    # Existing relevance logic (remark-based) OR Admission Status = "Irrelevant"
    # OR the Remarks contains one of the configured 'irrelevant' keyword variants.
    _admit_status = str(row.get("Admission Status", "")).strip().lower()
    row[RELEVANCE_FIELD] = "No" if (remark_flags_irrelevant(_remark)
                                    or _admit_status == "irrelevant"
                                    or remark_has_irrelevant_keyword(_remark)) else "Yes"

    # interaction history + analytics
    hist, ordered, n_int = build_history(records)
    row[HISTORY_FIELD]        = hist
    row["First Enquiry Date"] = fmt_dt(first_dt) if first_dt else ""
    row["Latest Enquiry Date"]= fmt_dt(last_dt) if last_dt else ""
    row["Number of Interactions"] = str(n_int)
    row["Platforms Used"] = ", ".join(SOURCE_LABEL[s] for s in
                                      ["IntelliBI", "Walk-In", "Call", "WhatsApp", "Website"] if s in present)
    return row


# ---------------------------------------------------------------------------
# 8. UPSERT vs. existing master
# ---------------------------------------------------------------------------

def lead_key(row):
    """Stable identity key for UPSERT — Phone then Email only (Full Name is NOT
    used for duplicate matching)."""
    ph = norm_phone(row.get("Mobile Number", ""))
    if ph and len(ph) >= 10:
        return "P:" + ph
    em = clean_email(row.get("Email Address", ""))
    if em:
        return "E:" + em
    # no phone/email: content key for row uniqueness only (not a name match)
    return "R:" + "|".join([
        clean_text(row.get("Full Name", "")).lower(),
        clean_text(row.get("LeadInitalTimestamp", "")),
        clean_text(row.get("Mobile Number", "")),
    ])


def upsert_stats(new_rows, existing_df):
    """Classify each freshly-computed lead against the existing master for
    logging: insert / update / skip, plus 'removed' (leads that were in the
    master but no longer exist in any source). The master is fully rebuilt from
    the current sources each run, so these counts are informational — the
    written dataset is always the fresh truth (no stale rows survive)."""
    existing = {}
    if existing_df is not None and len(existing_df):
        for _, r in existing_df.iterrows():
            existing[lead_key(r)] = {c: clean_text(r.get(c, "")) for c in MASTER_COLUMNS}

    inserted = updated = skipped = 0
    new_keys = set()
    for row in new_rows:
        k = lead_key(row)
        new_keys.add(k)
        if k not in existing:
            inserted += 1
        elif any(clean_text(row.get(c, "")) != existing[k].get(c, "") for c in MASTER_COLUMNS):
            updated += 1
        else:
            skipped += 1
    removed = sum(1 for k in existing if k not in new_keys)
    return inserted, updated, skipped, removed


# ---------------------------------------------------------------------------
# 9. OUTPUT
# ---------------------------------------------------------------------------

def write_csv_xlsx(df, summary):
    csv_path  = os.path.join(OUT_DIR, "Consolidated_Master_Lead.csv")
    xlsx_path = os.path.join(OUT_DIR, "Consolidated_Master_Lead.xlsx")
    df.to_csv(csv_path, index=False)

    sum_df = pd.DataFrame(list(summary.items()), columns=["Metric", "Value"])
    try:
        import openpyxl  # noqa
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as xw:
            df.to_excel(xw, sheet_name="Consolidated Master", index=False)
            sum_df.to_excel(xw, sheet_name="Execution Summary", index=False)
        _autoformat(xlsx_path)
    except Exception as e:
        print("xlsx write skipped:", e)
    return csv_path, xlsx_path


def _autoformat(xlsx_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.load_workbook(xlsx_path)
    for ws in wb.worksheets:
        head_fill = PatternFill("solid", fgColor="1F4E78")
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = head_fill
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        for col in ws.columns:
            width = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 45)
    wb.save(xlsx_path)


# ---------------------------------------------------------------------------
# 9b. gspread helpers (production mode)
# ---------------------------------------------------------------------------

def _sheets_service():
    """Authenticate with the repo service account and return a Sheets v4 client
    (identical auth to pyGoogleSheetSync.py)."""
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return build("sheets", "v4", credentials=creds, cache_discovery=False)


def _first_tab_title(svc, sheet_id):
    meta = svc.spreadsheets().get(spreadsheetId=sheet_id).execute()
    return meta["sheets"][0]["properties"]["title"]


def _read_gsheet_df(sheet_id):
    svc = _sheets_service()
    title = _first_tab_title(svc, sheet_id)
    resp = svc.spreadsheets().values().get(
        spreadsheetId=sheet_id, range=title,
        valueRenderOption="FORMATTED_VALUE",
        dateTimeRenderOption="FORMATTED_STRING").execute()
    values = resp.get("values", [])
    if not values:
        return pd.DataFrame()
    header = [h if h else f"col_{i}" for i, h in enumerate(values[0])]
    n = len(header)
    rows = [(r + [""] * n)[:n] for r in values[1:]]
    return pd.DataFrame(rows, columns=header).astype(str)


def read_existing_master():
    if INPUT_MODE != "gsheet":
        p = os.path.join(OUT_DIR, "Consolidated_Master_Lead.csv")
        if os.path.exists(p):
            return pd.read_csv(p, dtype=str, keep_default_na=False)
        return None
    try:
        return _read_gsheet_df(TARGET_SHEET_FULL)
    except Exception:
        return None


def push_to_gsheet(df, sheet_id):
    """Write the reconciled master into the target Google Sheet's first tab
    (clears then writes headers + rows; multi-line history stays as in-cell
    line breaks), then refresh the basic filter over the FULL written range and
    unmerge any stray merged cells so the sheet stays fully filterable."""
    svc = _sheets_service()
    meta = svc.spreadsheets().get(spreadsheetId=sheet_id).execute()
    first = meta["sheets"][0]["properties"]
    title, gid = first["title"], first["sheetId"]
    svc.spreadsheets().values().clear(spreadsheetId=sheet_id, range=title).execute()
    data = [list(df.columns)] + df.fillna("").astype(str).values.tolist()
    svc.spreadsheets().values().update(
        spreadsheetId=sheet_id, range=f"{title}!A1",
        valueInputOption="RAW", body={"values": data}).execute()

    # Keep the sheet filterable run-to-run: values().clear()/update() do NOT touch
    # an existing basic filter or merged cells, so a filter range left from a run
    # with fewer rows (or a stray merged cell) makes newer rows un-filterable.
    # Unmerge the whole tab and (re)apply a basic filter over the EXACT written
    # range. setBasicFilter replaces any existing basic filter.
    nrows, ncols = len(df) + 1, len(df.columns)          # +1 header row
    requests = [
        {"unmergeCells": {"range": {"sheetId": gid}}},
        {"setBasicFilter": {"filter": {"range": {
            "sheetId": gid, "startRowIndex": 0, "endRowIndex": nrows,
            "startColumnIndex": 0, "endColumnIndex": ncols}}}},
    ]
    try:
        svc.spreadsheets().batchUpdate(
            spreadsheetId=sheet_id, body={"requests": requests}).execute()
    except Exception as exc:                              # cosmetics never fail the push
        print("  [gsheet] filter/unmerge refresh skipped:", exc)
    print("Pushed %d rows to target sheet %s" % (len(df), sheet_id))


# ---------------------------------------------------------------------------
# 10. MAIN
# ---------------------------------------------------------------------------

def run():
    print(f"IntelliBI lead consolidation | mode={INPUT_MODE}")
    if INPUT_MODE == "gsheet":
        print(f"  service account : {SERVICE_ACCOUNT_FILE}")
        print(f"  target sheet    : {TARGET_SHEET_FULL}")
        if not os.path.exists(SERVICE_ACCOUNT_FILE):
            sys.exit(f"ERROR: service account file not found at {SERVICE_ACCOUNT_FILE}\n"
                     f"Set GOOGLE_SERVICE_ACCOUNT_FILE to your service_account.json path.")

    stats = Stats()
    business = set(BUSINESS_NUMBERS_SEED)

    loaders = [
        ("IntelliBI", load_intellibi),          # highest-priority source
        ("Walk-In",   load_walkin),
        ("Call",      lambda: load_calling(business)),
        ("WhatsApp",  load_whatsapp),
        ("Website",   load_website),
    ]

    interactions = []
    for src, fn in loaders:
        recs = list(fn())
        # usable = has at least one identity key
        usable = [r for r in recs if r["_phone"] or r["_name_key"] or r["_email"]]
        stats.processed[src] = len(usable)
        interactions.extend(recs)

    # --- Diagnostic: flag any record whose enquiry date parsed OUTSIDE a sane
    #     range (a malformed source timestamp). These no longer crash the run,
    #     but are surfaced here so the source row can be located and corrected.
    _yr_max = datetime.now().year + 5
    bad_dates = [r for r in interactions
                 if r.get("_ts") and not (2015 <= r["_ts"].year <= _yr_max)]
    if bad_dates:
        print(f"  [WARN] {len(bad_dates)} record(s) have an out-of-range enquiry "
              f"date (expected year 2015..{_yr_max}) — please correct the source row:")
        for r in bad_dates[:50]:
            phone = r.get("_phone_any") or r.get("_phone") or "(no phone)"
            name = r.get("_name_disp") or "(no name)"
            shown = fmt_dt(r["_ts"]) or r["_ts"].isoformat()
            print(f"         {r['_source']:8} | {phone:14} | {name:24} | parsed date: {shown}")
        if len(bad_dates) > 50:
            print(f"         ... and {len(bad_dates) - 50} more.")

    clusters = cluster_interactions(interactions, stats)
    master_rows = [merge_cluster(recs) for recs in clusters]

    # sort master by First Enquiry Date (chronological), blanks last
    def sort_key(r):
        dt = parse_dt(r.get("First Enquiry Date", ""))
        return (0, dt) if dt else (1, datetime.max)
    master_rows.sort(key=sort_key)

    existing_df = read_existing_master()
    inserted, updated, skipped, removed = upsert_stats(master_rows, existing_df)

    # Always write the freshly rebuilt master (current truth) — no stale rows.
    df = pd.DataFrame(master_rows, columns=MASTER_COLUMNS)

    total_records = sum(stats.processed.values())
    duplicates_merged = total_records - len(master_rows)
    summary = OrderedDict([
        ("Run Timestamp",            RUN_STAMP),
        ("Input Mode",               INPUT_MODE),
        ("IntelliBI Leads Processed", stats.processed["IntelliBI"]),
        ("Walk-In Leads Processed",  stats.processed["Walk-In"]),
        ("Website Leads Processed",  stats.processed["Website"]),
        ("WhatsApp Leads Processed", stats.processed["WhatsApp"]),
        ("Calling Leads Processed",  stats.processed["Call"]),
        ("Total Records Processed",  total_records),
        ("Unique Master Leads",      len(master_rows)),
        ("New Leads Inserted",       inserted),
        ("Existing Leads Updated",   updated),
        ("Existing Leads Skipped",   skipped),
        ("Existing Leads Removed",   removed),
        ("Duplicate Leads Merged",   duplicates_merged),
        ("Matching by Phone",        stats.match_phone),
        ("Matching by Name",         stats.match_name),
        ("Matching by Email",        stats.match_email),
        ("Failed Matches",           stats.failed_match),
        ("Empty Records Ignored",    stats.ignored_empty),
    ])

    csv_path, xlsx_path = write_csv_xlsx(df, summary)

    if INPUT_MODE == "gsheet":
        push_to_gsheet(df, TARGET_SHEET_FULL)

    print("\n================ EXECUTION SUMMARY ================")
    for k, v in summary.items():
        print(f"  {k:<28}: {v}")
    print("==================================================")
    print("CSV :", csv_path)
    print("XLSX:", xlsx_path)
    return df, summary


RUN_STAMP = datetime.now().strftime(OUT_DT_FMT)

if __name__ == "__main__":
    run()
