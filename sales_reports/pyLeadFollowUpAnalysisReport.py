#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IntelliBI - Lead Follow-Up Analysis Report  (second-level analytical report)
===========================================================================

A NEW, standalone report that complements (never modifies)
pyConsolidatedLeadPerformanceReport.py. Where the first report describes lead
*activity*, this one is *decision-oriented*: it tells the counselling team what
to act on next, how each counsellor is performing, and — most importantly — a
**data-driven Conversion Chance %** for every active lead, learned from
IntelliBI's own historical conversions.

WHAT IT PRODUCES  (Daily / Weekly / Monthly / Full)
    Summary            - follow-up, Google-Meet and Walk-In KPIs (counts + %)
    Priority & Actions - active leads ranked by Conversion Chance %, with the
                         single next action (overdue follow-up / meet no-show …)
    Counsellor Perf.   - every metric per counsellor, in %
    Conversion Model   - full transparency: the learned weight of each factor
    (Additional analysis: overdue follow-ups, meet/walk-in vs conversion, and
     where leads are being lost.)

DATA SOURCES
    1. IntelliBI Lead Information Result  (Active + InActive tabs)
         - Active   = current lead record  (the leads we score & act on)
         - InActive = archived historical versions (RecordVersion / ArchivedAt)
                      -> used to reconstruct the real follow-up history.
    2. IntelliBI Consolidate Sales Tracking Report (master)
         - the full journey: Platforms Used, Number of Interactions, First
           Enquiry, Lead Interaction History, candidate background, etc.
    3. IntelliBI Career Guidance Google Meet Form  (LeadAttendanceStatus)
    4. Walk-In form                                (presence == attended)
    Historical files in  intellibi_lead_cycle\\  are read for extra conversion
    signal when present (optional; never required).

CONVERSION TRUTH (chosen by IntelliBI)
    A lead counts as CONVERTED when its Admission Status = "Admission Confirmed".
    The Conversion Chance % model is trained on that label (see MODEL below).

MODEL  (data-driven, transparent — no black-box ML)
    A weight-of-evidence (naive-Bayes log-odds) model. For each factor value we
    measure its real historical conversion rate, convert it to a log-odds shift
    versus the base rate, shrink it toward zero by how much data supports it
    (so tiny groups can't dominate), and sum the shifts. The result is a
    calibrated probability per lead. Hot/Warm/Nurture/Low bands are anchored to
    the measured base rate, not to arbitrary numbers.

AUTH / CONVENTIONS  (identical to pyConsolidatedLeadPerformanceReport.py)
    service account  config_files/service_account.json ; Sheets v4 (SA-direct
    read) + Drive v3 (impersonated upload). Output Google Sheet is uploaded to
    the configured Drive folder, with a styled local .xlsx backup.

RUN
    python Reports/pyLeadFollowUpAnalysisReport.py
OFFLINE TEST (no Google I/O — read every source from local .xlsx/.csv):
    LFA_LOCAL_DIR="/path/to/intellibi_lead_cycle" LFA_DRY_RUN=1 \
        python Reports/pyLeadFollowUpAnalysisReport.py
"""

import os
import re
import sys
import csv
import copy
import math
import glob
import calendar
from datetime import datetime, date, timedelta, time, timezone
from collections import OrderedDict, defaultdict, Counter

# ── team reads IST; stamp "Generated" in IST regardless of machine TZ ─────────
IST = timezone(timedelta(hours=5, minutes=30))


def now_ist():
    return datetime.now(timezone.utc).astimezone(IST).replace(tzinfo=None)


# ═══════════════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════
# GENERATE_AUTO = scheduler-friendly automatic selection based on the execution
# date. When True, the Daily/Weekly/Monthly/Manual flags below are IGNORED for
# that run and the script decides automatically:
#     • Daily            -> every run, for the current day
#     • Weekly           -> only on Monday, for the previous complete Mon–Sun week
#     • Monthly          -> only on the last calendar day of the month, for that
#                           whole month (1st → last day)
# When False, the manual GENERATE_DAILY/WEEKLY/MONTHLY/MANUAL flags are used
# exactly as before.
GENERATE_AUTO    = True

GENERATE_DAILY   = True
GENERATE_WEEKLY  = True
GENERATE_MONTHLY = True
GENERATE_MANUAL  = False         # Manual = a custom start/end date range

# Optional manual periods (None -> use today / current week / current month).
DAILY_DATE            = None #'2026-08-22'     # "YYYY-MM-DD"
WEEKLY_REFERENCE_DATE = None     # any day in the wanted week
MONTHLY_MONTH         = None     # 1-12
MONTHLY_YEAR          = None
MANUAL_START_DATE     = None     # "YYYY-MM-DD"  (required when GENERATE_MANUAL)
MANUAL_END_DATE       = None     # "YYYY-MM-DD"  (required when GENERATE_MANUAL)

# ── Source spreadsheets ──────────────────────────────────────────────────────
RESULT_SHEET_ID  = "1ReJVPl_Y8WnOl_P2sui_uC1jjZXVk0dWqNWRcXGVHCw"   # Active+InActive
RESULT_ACTIVE_TABS   = ["IntelliBI Lead Information Active",
                        "IntelliBI Lead Information Acti",
                        "IntelliBI Lead Information Result"]
RESULT_INACTIVE_TABS = ["IntelliBI Lead Information InActive",
                        "IntelliBI Lead Information InAc"]

MASTER_SHEET_ID  = "1zZQjXnMJD96Ca0MNyfSt4-XS0z5w3rT7WPdb9qsP1Gs"   # Consolidate Sales Tracking
MASTER_TABS      = None          # None -> first tab

# Already-enrolled students. Anyone whose phone appears here is EXCLUDED from all
# Follow-Up Pending calculations/details (they have already enrolled, so there is
# nothing to follow up). Read from the "New Enroll" tab (gid below) of:
#   IntelliBI — Student Admission Responses-New Enroll
ENROLL_SHEET_ID  = "1oaXxg3JdtxFp8lFWijIMZKaMZvS0SiglI1K2JTrN2fs"
ENROLL_GID       = 689069678     # the "New Enroll" tab
ENROLL_TABS      = ["New Enroll", "Enroll", "Enrolled", "Student Admission Responses",
                    "Admission Responses", "Responses"]
ENROLL_PHONE_COLS = ["mobile number", "mobile", "phone number", "phone",
                     "whatsapp number", "whatsapp", "contact number", "contact",
                     "mobile no", "phone no"]

MEET_SHEET_ID    = "1dlWiU5K7kFi014p8pgH4PbMuoy4QbwMHh48YxNOqjpg"   # Google Meet form
MEET_TABS        = None          # None -> first tab
MEET_ATTEND_COL_CANDIDATES = ["LeadAttendanceStatus", "Lead Attendance Status",
                              "Attendance Status", "Attendance"]

WALKIN_SHEET_ID  = "19Ecal2JpOL1FbzGKWlno4ZywG3HsXsiK-BmMzew5TqQ"   # Walk-In form
WALKIN_TABS      = None          # None -> first tab
# Walk-In ATTENDANCE = the "Walk-In New" tab of the Student Inquiry Tracker
# (same spreadsheet). A scheduled walk-in "attended" = its mobile appears here;
# the "Timestamp" column is the actual walk-in date/time.
WALKIN_ATTEND_TABS = ["Walk-In New", "WalkIn New", "Walk In New"]
WALKIN_TS_COL_CANDIDATES = ["Timestamp", "Time Stamp", "Walk-In Date & Time",
                            "Walk-in Date & Time", "Date"]

# ── Output Drive folders (chosen by IntelliBI) ───────────────────────────────
# Each period is saved into its own sub-folder under this parent. Sub-folders
# are resolved by NAME at run time (and created if missing), so the report can
# never land in the wrong folder even if a folder is renamed/re-created.
OUTPUT_PARENT_FOLDER_ID = "1UPsCa-i_KV_ynoRSRULNp6WbHOrwhCMJ"
OUTPUT_SUBFOLDERS = {
    "Daily":   "Daily Lead Follow-up Tracking Report",
    "Weekly":  "Weekly Lead Follow-up Tracking Report",
    "Monthly": "Monthly Lead Follow-up Tracking Report",
    "Manual":  "Manual Lead Follow-up Tracking Report",
}

# ── Email ─────────────────────────────────────────────────────────────────────
# Same configuration/recipients as pyConsolidatedLeadPerformanceReport.py: Gmail
# SMTP from config_files/email_config.py (GMAIL_SENDER / GMAIL_APP_PASS).
SEND_EMAIL       = True
EMAIL_RECIPIENTS = [
    "harishintellibi@gmail.com",
    "salesintellibi01@gmail.com",
    "info@intellibiinnovationstechnologies.in",
    "163manish.sharma@gmail.com",
]
# Recipients who must NOT receive the link to the full (unmasked) report — they
# still get the same aggregate summary email, just without the detailed-report
# link (this report has no masked copy). Mirrors the consolidated report's rule
# of never sending unmasked lead detail to a masking recipient.
MASK_RECIPIENTS = {"163manish.sharma@gmail.com"}

# ── Auth (same pattern as the existing report) ───────────────────────────────
# This script lives in the "Reports" sub-folder, so PROJECT_ROOT is its PARENT
# (the IntelliBI Automation root) — same convention as
# pyConsolidatedLeadPerformanceReport.py — so config_files/, intellibi_lead_cycle/
# and output/ all still resolve to the project root after the move.
# --- IntelliBI Sales Automation portability bootstrap (auto-inserted) ---
import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "common"))
import _bootstrap  # noqa: E402  (sys.path + env defaults + config.yaml)
from paths import CREDENTIALS_DIR, CONFIG_DIR, LOGS_DIR  # noqa: E402
# --- end bootstrap ---

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SERVICE_ACCOUNT_FILE = os.environ.get(
    "GOOGLE_SERVICE_ACCOUNT_FILE",
    os.path.join(CREDENTIALS_DIR, "service_account.json"))
READ_SCOPES  = ["https://www.googleapis.com/auth/spreadsheets"]      # SA-direct read
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive"]            # impersonated upload
IMPERSONATE_USER = os.environ.get(
    "REPORT_IMPERSONATE_USER", "info@intellibiinnovationstechnologies.in")

# ── Offline / testing switches (env-driven; unset in production) ──────────────
# LFA_LOCAL_DIR : read every source from local .xlsx in this folder (offline).
# LFA_DRY_RUN=1 : build + write the local .xlsx only, no Google I/O at all.
LOCAL_DIR = os.environ.get("LFA_LOCAL_DIR")
DRY_RUN   = os.environ.get("LFA_DRY_RUN") == "1"
OUTPUT_DIR = os.environ.get(
    "LFA_OUTPUT_DIR", os.path.join(PROJECT_ROOT, "output"))

# Historical-data folder: extra past exports used ONLY to enlarge the model's
# training set (never for the current active-lead journey). Requirement #2/#11.
HISTORY_DIR = os.environ.get("LFA_HISTORY_DIR") or \
    (LOCAL_DIR if os.environ.get("LFA_LOCAL_DIR") else
     os.path.join(PROJECT_ROOT, "intellibi_lead_cycle"))
HISTORY_MASTER_GLOBS = ["*Consolidate*Sales*Tracking*.xlsx",
                        "*Consolidated*Master*Lead*.xlsx"]

# Map each logical source -> (filename glob, [tab name candidates]) for offline.
LOCAL_SOURCES = {
    "result_active":   ("IntelliBI Lead Information Result*.xlsx", RESULT_ACTIVE_TABS),
    "result_inactive": ("IntelliBI Lead Information Result*.xlsx", RESULT_INACTIVE_TABS),
    "master":          ("IntelliBI Consolidate Sales Tracking Report*.xlsx", None),
    "meet":            ("*Google Meet*ttendance*.xlsx", None),   # optional
    "walkin":          ("*Walk*In*form*.xlsx", None),            # optional
    "walkin_new":      ("*Student Inquiry Tracker*.xlsx", WALKIN_ATTEND_TABS),  # optional
}

# ── Follow-up policy ─────────────────────────────────────────────────────────
MAX_FOLLOW_UPS = 5
MAX_COUNSELLOR_TABS = 40

# ── Conversion-model tunables ────────────────────────────────────────────────
SMOOTH_ALPHA = 5.0    # Laplace pseudo-count toward the base rate per feature cell
SHRINK_K     = 15.0   # support at which a cell's evidence is ~half-trusted
PROB_FLOOR, PROB_CEIL = 0.01, 0.99
# Missing / residual buckets must NOT push the score — a blank field means "we
# don't know", not "won't convert". These values contribute 0 to the log-odds.
NEUTRAL_VALUES = {"Unknown", "Other", "none"}
# Independence-correction: features overlap (interactions/reach/speed correlate),
# so naive-Bayes over-counts evidence. Damp the summed shift toward the base rate
# to keep the probability calibrated and avoid extreme 1%/90% swings.
WOE_DAMP = 0.75

CONVERTED_STATUS_MATCH = "admission confirmed"     # the chosen conversion truth

try:
    from dateutil import parser as _dtparser
except Exception:
    _dtparser = None

DATE_FMTS = ("%d-%b-%Y %I:%M %p", "%d-%b-%Y", "%Y-%m-%d %H:%M:%S",
             "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d",
             "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
             "%m/%d/%Y %H:%M:%S", "%m/%d/%Y")


# ═══════════════════════════════════════════════════════════════════════════
#  SMALL HELPERS
# ═══════════════════════════════════════════════════════════════════════════
def s(v):
    return "" if v is None else str(v).strip()


def yes(v):
    return s(v).lower() in ("yes", "y", "true", "1")


def pct(n, d):
    return "0.0%" if not d else f"{100.0 * n / d:.1f}%"


def digits10(v):
    """Last 10 digits of a phone value, robust to numeric/float-formatted cells
    (e.g. 9112454898.0 from an .xlsx) and to +91 / leading-zero prefixes."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = str(int(v))
    else:
        v = str(v).strip()
        m = re.match(r"^(\d+)\.0+$", v)      # numbers-as-text like "9112454898.0"
        if m:
            v = m.group(1)
    d = re.sub(r"\D", "", v)
    return d[-10:] if len(d) >= 10 else d


# ── Recipient-specific masking (identical rules to
#    pyConsolidatedLeadPerformanceReport.py) ───────────────────────────────────
def mask_mobile(value):
    """Mask a mobile number: keep the first 2 and last 2 digits, replace the
    digits in between with '*'.  e.g. 9876543210 -> 98******10.  Values that are
    blank or have <=4 digits are returned unchanged (nothing meaningful to hide)."""
    v = s(value)
    if not v:
        return v
    digits = re.sub(r"\D", "", v)
    if len(digits) <= 4:
        return v
    return digits[:2] + ("*" * (len(digits) - 4)) + digits[-2:]


def mask_email(value):
    """Mask an email's local part: keep the first 2 and last 2 characters, star
    the rest; the domain stays visible.  e.g. rahulkumar@gmail.com ->
    ra******ar@gmail.com,  abcdef@company.com -> ab**ef@company.com."""
    v = s(value)
    if not v or "@" not in v:
        return v
    local, domain = v.split("@", 1)
    n = len(local)
    if n <= 2:
        masked = local                                   # too short to mask
    elif n <= 4:
        masked = local[0] + ("*" * (n - 2)) + local[-1]  # keep 1 + 1
    else:
        masked = local[:2] + ("*" * (n - 4)) + local[-2:]
    return masked + "@" + domain


def parse_dt(v):
    v = s(v)
    if not v:
        return None
    for f in DATE_FMTS:
        try:
            return datetime.strptime(v, f)
        except ValueError:
            pass
    if _dtparser:
        try:
            return _dtparser.parse(v, dayfirst=True, fuzzy=True)
        except Exception:
            return None
    return None


def parse_ts_mdy(v):
    """Parse a Google-Form 'Timestamp' cell (US month-first locale) -> datetime
    | None. The Walk-In New sheet's Timestamp is m/d/Y, e.g.
    '8/12/2026 17:27:57' == 12-Aug-2026 (NOT 8-Dec). Month-first slash formats
    are tried first so the day/month order is never swapped; real datetime cells
    (already unambiguous) pass through unchanged."""
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime.combine(v, time.min)
    t = s(v)
    if not t:
        return None
    for f in ("%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y %I:%M:%S %p",
              "%m/%d/%Y %I:%M %p", "%m/%d/%Y",
              "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S",
              "%Y-%m-%d", "%d-%b-%Y %I:%M %p", "%d-%b-%Y"):
        try:
            return datetime.strptime(t, f)
        except ValueError:
            pass
    if _dtparser:
        try:
            return _dtparser.parse(t, dayfirst=False, fuzzy=True)
        except Exception:
            return None
    return None


def parse_followup_date(v):
    """Return the DATE portion of a Next-Follow-Up value, tolerant of every
    format seen across both sheets and ignoring any time component:
        '18-Aug-2026 12:00 AM', '18-Aug-2026', '8/18/2026', '2026-08-18',
        datetime / date cell objects (openpyxl), etc.
    Slash dates are read US-style (m/d/Y) — matching the Lead Information sheet
    (e.g. '8/18/2026') — with d/m/Y as a final fallback. Returns None if
    unparseable."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    t = s(v)
    if not t:
        return None
    fmts = ("%d-%b-%Y %I:%M %p", "%d-%b-%Y",
            "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d",
            "%m/%d/%Y %I:%M %p", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y",
            "%d/%m/%Y")
    for f in fmts:
        try:
            return datetime.strptime(t, f).date()
        except ValueError:
            pass
    dt = parse_dt(t)
    return dt.date() if dt else None


def norm_email(v):
    return re.sub(r"\s+", "", s(v)).lower()


def sigmoid(x):
    if x < -60:
        return 0.0
    if x > 60:
        return 1.0
    return 1.0 / (1.0 + math.exp(-x))


def pick(columns, *candidates):
    """Resolve a real column name from candidates: exact (case/space-insensitive)
    first, then 'contains'. Returns None if nothing matches."""
    norm = {re.sub(r"\s+", " ", str(c)).strip().casefold(): c for c in columns}
    for cand in candidates:
        k = re.sub(r"\s+", " ", str(cand)).strip().casefold()
        if k in norm:
            return norm[k]
    for cand in candidates:
        k = re.sub(r"\s+", " ", str(cand)).strip().casefold()
        for nk, orig in norm.items():
            if k and k in nk:
                return orig
    return None


# ═══════════════════════════════════════════════════════════════════════════
#  DATA MODEL (light dict rows keyed by resolved column names)
# ═══════════════════════════════════════════════════════════════════════════
class Table:
    """A simple header + rows container with case-insensitive column resolution."""
    def __init__(self, headers, rows):
        self.headers = list(headers)
        self.rows = rows            # list[dict]

    def col(self, *candidates):
        return pick(self.headers, *candidates)

    def __len__(self):
        return len(self.rows)


# ═══════════════════════════════════════════════════════════════════════════
#  SHEET / DRIVE I/O
# ═══════════════════════════════════════════════════════════════════════════
def _creds(scopes, impersonate=False):
    from google.oauth2 import service_account
    if not os.path.exists(SERVICE_ACCOUNT_FILE):
        sys.exit(f"ERROR: service account not found at {SERVICE_ACCOUNT_FILE}")
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=scopes)
    if impersonate and IMPERSONATE_USER:
        creds = creds.with_subject(IMPERSONATE_USER)
    return creds


def get_read_service():
    from googleapiclient.discovery import build
    return build("sheets", "v4",
                 credentials=_creds(READ_SCOPES, impersonate=False),
                 cache_discovery=False)


def get_drive_service():
    from googleapiclient.discovery import build
    return build("drive", "v3",
                 credentials=_creds(DRIVE_SCOPES, impersonate=True),
                 cache_discovery=False)


def _resolve_local_file(fileglob):
    if not LOCAL_DIR:
        return None
    hits = sorted(glob.glob(os.path.join(LOCAL_DIR, fileglob)))
    return hits[0] if hits else None


def _read_local_table(local_key):
    """Offline reader: pull a source from a local .xlsx (openpyxl)."""
    import openpyxl
    fileglob, tab_candidates = LOCAL_SOURCES[local_key]
    path = _resolve_local_file(fileglob)
    if not path:
        return None
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    tab = None
    if tab_candidates:
        tab = pick(wb.sheetnames, *tab_candidates)
    tab = tab or wb.sheetnames[0]
    ws = wb[tab]
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if not header:
        return Table([], [])
    header = [s(h) or f"col_{i}" for i, h in enumerate(header)]
    n = len(header)
    rows = []
    for r in it:
        r = list(r) + [""] * (n - len(r))
        rows.append({header[i]: ("" if r[i] is None else r[i]) for i in range(n)})
    return Table(header, rows)


def read_source(sheets, spreadsheet_id, tab_candidates, local_key, optional=False):
    """Read a source into a Table. Uses the local file in offline mode, else the
    Sheets API. Returns None if the source is unavailable (e.g. optional form
    missing offline) so callers can degrade gracefully.

    optional=True : if the sheet cannot be read (e.g. the service account has no
    permission, or the sheet is missing), warn and return None instead of
    crashing the whole report. Used for the Meet / Walk-In enrichment forms,
    which are nice-to-have, not required. Required sources (leads / master) keep
    raising so a genuine access problem is never silently ignored."""
    if LOCAL_DIR:
        try:
            return _read_local_table(local_key)
        except Exception as e:
            if optional:
                print(f"  [source] optional '{local_key}' unavailable ({e}); "
                      f"continuing without it.")
                return None
            raise
    if sheets is None:
        return None
    try:
        meta = sheets.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    except Exception as e:
        if optional:
            print(f"  [source] optional '{local_key}' ({spreadsheet_id}) not "
                  f"readable ({getattr(e, 'status_code', '')} permission/"
                  f"availability); continuing without it.")
            return None
        raise
    titles = [sh["properties"]["title"] for sh in meta["sheets"]]
    tab = pick(titles, *tab_candidates) if tab_candidates else titles[0]
    tab = tab or titles[0]
    try:
        resp = sheets.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id, range=tab,
            valueRenderOption="FORMATTED_VALUE").execute()
    except Exception as e:
        if optional:
            print(f"  [source] optional '{local_key}' values not readable "
                  f"({e}); continuing without it.")
            return None
        raise
    values = resp.get("values", [])
    if not values:
        return Table([], [])
    header = [h if h else f"col_{i}" for i, h in enumerate(values[0])]
    n = len(header)
    rows = []
    for r in values[1:]:
        r = (r + [""] * n)[:n]
        rows.append({header[i]: r[i] for i in range(n)})
    return Table(header, rows)


def load_enrolled_phones(sheets):
    """Return a set of normalized (last-10-digit) phone numbers of students who
    are ALREADY ENROLLED — read from the New-Enroll tab (ENROLL_GID) of the
    enrolled-students sheet. These phones are excluded from every Follow-Up
    Pending calculation and detail. Phone matching is via digits10() so
    '+91 98765 43210', '98765 43210' and '9876543210' all collapse to the same
    key. Never raises: on any access problem it returns an empty set and warns,
    so the report still runs (just without the exclusion)."""
    if sheets is None or LOCAL_DIR:
        return set()
    try:
        meta = sheets.spreadsheets().get(spreadsheetId=ENROLL_SHEET_ID).execute()
    except Exception as e:
        print(f"  [enrolled] sheet {ENROLL_SHEET_ID} not readable ({e}); "
              f"no enrolled-student exclusion applied.")
        return set()
    # locate the exact tab by gid; fall back to a named/first tab if gid missing.
    title = None
    for sh in meta.get("sheets", []):
        if sh.get("properties", {}).get("sheetId") == ENROLL_GID:
            title = sh["properties"]["title"]
            break
    if title is None:
        titles = [sh["properties"]["title"] for sh in meta.get("sheets", [])]
        title = pick(titles, *ENROLL_TABS) or (titles[0] if titles else None)
    if not title:
        print("  [enrolled] no readable tab; no enrolled-student exclusion applied.")
        return set()
    try:
        resp = sheets.spreadsheets().values().get(
            spreadsheetId=ENROLL_SHEET_ID, range=title,
            valueRenderOption="FORMATTED_VALUE").execute()
    except Exception as e:
        print(f"  [enrolled] '{title}' values not readable ({e}); "
              f"no enrolled-student exclusion applied.")
        return set()
    values = resp.get("values", [])
    if not values:
        return set()
    header = [h if h else f"col_{i}" for i, h in enumerate(values[0])]
    ph_col = pick(header, *ENROLL_PHONE_COLS)
    if not ph_col:
        print(f"  [enrolled] no phone column found in '{title}' "
              f"(headers: {header[:8]}...); no enrolled-student exclusion applied.")
        return set()
    ci = header.index(ph_col)
    phones = set()
    for r in values[1:]:
        if ci < len(r):
            d = digits10(r[ci])
            if len(d) == 10:          # only real 10-digit phones become match keys
                phones.add(d)
    print(f"  [enrolled] loaded {len(phones)} enrolled phone(s) from '{title}' "
          f"(column '{ph_col}') for Follow-Up Pending exclusion.")
    return phones


def resolve_output_folder(drive, parent_id, subfolder_name):
    """Return the Drive folder id for `subfolder_name` under `parent_id`,
    creating it if it does not exist. Guarantees the report is written to the
    correct period folder (never the parent, never a stale id)."""
    safe = subfolder_name.replace("'", "\\'")
    try:
        res = drive.files().list(
            q=("name = '%s' and '%s' in parents and "
               "mimeType = 'application/vnd.google-apps.folder' and trashed = false"
               % (safe, parent_id)),
            fields="files(id,name)", pageSize=10,
            supportsAllDrives=True, includeItemsFromAllDrives=True).execute()
        hits = res.get("files", [])
        if hits:
            return hits[0]["id"]
    except Exception as e:
        print("  [drive] folder lookup failed:", e)
    # not found -> create it
    meta = drive.files().create(
        body={"name": subfolder_name, "parents": [parent_id],
              "mimeType": "application/vnd.google-apps.folder"},
        fields="id", supportsAllDrives=True).execute()
    print(f"  [drive] created sub-folder '{subfolder_name}'")
    return meta["id"]


def upload_report_to_drive(drive, folder_id, name, xlsx_path):
    """Upload the styled .xlsx into the folder as a Google Sheet, replacing any
    same-named file. Drive scope only (impersonated). Returns (url, created,
    file_id)."""
    from googleapiclient.http import MediaFileUpload
    safe = name.replace("'", "\\'")
    old = []
    try:
        res = drive.files().list(
            q="name = '%s' and '%s' in parents and trashed = false" % (safe, folder_id),
            fields="files(id)", pageSize=20,
            supportsAllDrives=True, includeItemsFromAllDrives=True).execute()
        old = res.get("files", [])
        for f in old:
            drive.files().delete(fileId=f["id"], supportsAllDrives=True).execute()
    except Exception as e:
        print("  [drive] could not remove previous file(s):", e)
    media = MediaFileUpload(
        xlsx_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False)
    body = {"name": name, "parents": [folder_id],
            "mimeType": "application/vnd.google-apps.spreadsheet"}
    try:
        f = drive.files().create(body=body, media_body=media,
                                 fields="id,webViewLink",
                                 supportsAllDrives=True).execute()
    except Exception as e:
        if "storageQuota" in str(e):
            sys.exit("\nERROR: 'storageQuotaExceeded' creating the report file.\n"
                     "IMPERSONATE_USER must own the output folder and the DWD grant\n"
                     "must authorise the `drive` scope for the service account.\n")
        raise
    url = f.get("webViewLink") or \
        f"https://docs.google.com/spreadsheets/d/{f['id']}/edit"
    return url, (len(old) == 0), f["id"]


def share_file_with(drive, file_id, emails, role="writer"):
    """Grant `role` (default Editor/writer) on a Drive file to one or more email
    addresses. Best-effort per address; failures are logged, never fatal. No
    Google notification email is sent (our own report email carries the link)."""
    for em in (emails if isinstance(emails, (list, tuple, set)) else [emails]):
        em = s(em)
        if not em:
            continue
        try:
            drive.permissions().create(
                fileId=file_id,
                body={"type": "user", "role": role, "emailAddress": em},
                sendNotificationEmail=False,
                supportsAllDrives=True).execute()
            print(f"  [drive] shared masked report with {em} ({role})")
        except Exception as e:
            print(f"  [drive] could NOT share masked report with {em}:", e)


# Header names (case-insensitive) whose column cells are masked in the restricted
# copy — the same two fields the consolidated report masks.
_MASK_MOBILE_HDRS = {"mobile", "mobile number", "phone number", "phone"}
_MASK_EMAIL_HDRS  = {"email", "email address", "email id"}


def mask_tabs(tabs):
    """Return a DEEP COPY of the report tabs with every Mobile-Number and
    Email-Address cell masked (identical mask_mobile / mask_email rules as
    pyConsolidatedLeadPerformanceReport.py). The original tabs are never
    modified, so authorised recipients still get the complete data. Masks by
    column HEADER name, per section, so only those two fields are touched."""
    out = OrderedDict()
    for name, tab in tabs.items():
        t = copy.deepcopy(tab)
        for hi in sorted(t.headers):
            hdr = t.rows[hi]
            mob_cols, email_cols = [], []
            for ci, h in enumerate(hdr):
                hl = str(h).strip().lower()
                if hl in _MASK_MOBILE_HDRS:
                    mob_cols.append(ci)
                elif hl in _MASK_EMAIL_HDRS:
                    email_cols.append(ci)
            if not mob_cols and not email_cols:
                continue
            a, b = t.data_span(hi)
            for ri in range(a, b):
                row = t.rows[ri]
                for ci in mob_cols:
                    if ci < len(row):
                        row[ci] = mask_mobile(row[ci])
                for ci in email_cols:
                    if ci < len(row):
                        row[ci] = mask_email(row[ci])
        out[name] = t
    return out


# ═══════════════════════════════════════════════════════════════════════════
#  INTERACTION-HISTORY PARSING
# ═══════════════════════════════════════════════════════════════════════════
def parse_history(text):
    """Parse a master 'Lead Interaction History' cell into a chronological list of
    (source, datetime, counsellor). Returns [] when unparseable."""
    t = s(text)
    if not t:
        return []
    srcs, dates, cbys = [], [], []
    for line in t.splitlines():
        low = line.lower()
        if low.startswith("lead source:"):
            srcs = [x.strip() for x in line.split(":", 1)[1].split(",")]
        elif low.startswith("lead date:"):
            dates = [x.strip() for x in line.split(":", 1)[1].split(",")]
        elif low.startswith("counselling by:"):
            cbys = [x.strip() for x in line.split(":", 1)[1].split(",")]
    out = []
    for i, dstr in enumerate(dates):
        dt = parse_dt(dstr)
        if not dt:
            continue
        out.append((srcs[i].strip() if i < len(srcs) else "",
                    dt,
                    cbys[i].strip() if i < len(cbys) else ""))
    out.sort(key=lambda x: x[1])
    return out


def platforms_in_sequence(history, fallback_csv=""):
    seq = []
    for src, _dt, _cb in history:
        n = s(src)
        if n and n != "—" and n not in seq:
            seq.append(n)
    for p in [p.strip() for p in s(fallback_csv).split(",") if p.strip()]:
        if p not in seq:
            seq.append(p)
    return seq


def response_speed_bucket(history):
    """Classify how promptly the lead engages, from gaps between interactions."""
    if len(history) < 2:
        return "single_touch"
    gaps = [(history[i][1] - history[i - 1][1]).total_seconds() / 86400.0
            for i in range(1, len(history))]
    gaps = [g for g in gaps if g >= 0]
    if not gaps:
        return "single_touch"
    med = sorted(gaps)[len(gaps) // 2]
    if med < 1:
        return "same_day"
    if med <= 3:
        return "fast_3d"
    if med <= 7:
        return "week"
    return "delayed"


# ═══════════════════════════════════════════════════════════════════════════
#  NORMALISERS (feature bucketing — shared by model & report)
# ═══════════════════════════════════════════════════════════════════════════
def norm_platform(name):
    n = s(name).lower()
    if not n:
        return "Unknown"
    if "referr" in n:
        return "Referral"
    if "walk" in n:
        return "Walk-In"
    if "website" in n or "web" == n:
        return "Website"
    if "whats" in n:
        return "WhatsApp"
    if "call" in n:
        return "Call"
    if "meta" in n or "facebook" in n or "insta" in n:
        return "Meta"
    if "intellibi" in n:
        return "IntelliBI"
    if "bot" in n:
        return "Bot"
    if "campaign" in n:
        return "Campaign"
    if "master" in n:
        return "Masterclass"
    if "re-targ" in n or "retarg" in n:
        return "Re-Targeting"
    return s(name)


def norm_work(*vals):
    t = " ".join(s(v).lower() for v in vals)
    if not t.strip():
        return "Unknown"
    if "working" in t or "employ" in t or "job" in t or "professional" in t:
        return "Working"
    if "break" in t:
        return "CareerBreak"
    if "student" in t:
        return "Student"
    if "fresher" in t or "passed out" in t or "pass out" in t:
        return "Fresher"
    return "Other"


def norm_timeline(v):
    t = s(v).lower()
    if not t:
        return "Unknown"
    if "immediat" in t or "within 1 week" in t or "1 week" in t:
        return "Immediate"
    if "15" in t:
        return "Within15Days"
    if "1 month" in t or "one month" in t:
        return "Within1Month"
    if "2 month" in t or "two month" in t:
        return "Within2Months"
    if "explor" in t or "decid" in t:
        return "Exploring"
    return "Other"


def norm_city(v):
    t = s(v).lower()
    if not t:
        return "Unknown"
    if "pune" in t or "pimpri" in t or "chinchwad" in t or "pcmc" in t:
        return "Pune"
    return "OtherCity"


def interactions_bucket(n):
    try:
        n = int(float(n))
    except (TypeError, ValueError):
        n = 0
    if n <= 1:
        return "1"
    if n == 2:
        return "2"
    if n <= 4:
        return "3-4"
    return "5+"


def is_converted_status(admission_status):
    return CONVERTED_STATUS_MATCH in s(admission_status).lower()


def is_lost_status(admission_status, lead_status, backout):
    a = s(admission_status).lower()
    l = s(lead_status).lower()
    if any(k in a for k in ("not interested", "irrelevant", "backed out",
                            "backout", "dropped", "lost")):
        return True
    if l in ("not interested", "lost", "dropped"):
        return True
    if s(backout):
        return True
    return False


# ═══════════════════════════════════════════════════════════════════════════
#  FEATURE EXTRACTION  (one canonical feature dict per lead)
# ═══════════════════════════════════════════════════════════════════════════
FEATURE_ORDER = ["platform_primary", "reach", "interactions", "referral",
                 "work", "timeline", "city", "meet", "walkin", "speed"]

FEATURE_LABEL = {
    "platform_primary": "Primary platform",
    "reach": "Single vs multi-platform",
    "interactions": "Number of interactions",
    "referral": "Referral lead",
    "work": "Working status",
    "timeline": "Admission timeline",
    "city": "Location",
    "meet": "Google-Meet engagement",
    "walkin": "Walk-In engagement",
    "speed": "Response speed",
}


def build_features(*, history, platforms, num_interactions, referral,
                   work, timeline, city, meet_state, walkin_state):
    prim = norm_platform(platforms[0]) if platforms else "Unknown"
    return {
        "platform_primary": prim if not referral else "Referral",
        "reach": "multi" if len(platforms) > 1 else "single",
        "interactions": interactions_bucket(num_interactions),
        "referral": "yes" if referral else "no",
        "work": work,
        "timeline": timeline,
        "city": city,
        "meet": meet_state,     # attended / noshow / scheduled / none
        "walkin": walkin_state,  # attended / scheduled / none
        "speed": response_speed_bucket(history),
    }


# ═══════════════════════════════════════════════════════════════════════════
#  WEIGHT-OF-EVIDENCE CONVERSION MODEL
# ═══════════════════════════════════════════════════════════════════════════
class ConversionModel:
    """Naive-Bayes log-odds over bucketed features, with Laplace smoothing and
    support-based shrinkage so small groups can't dominate the score."""

    def __init__(self):
        self.base_rate = 0.05
        self.base_logodds = math.log(0.05 / 0.95)
        self.woe = {}                # feature -> value -> (woe, n, pos, rate)
        self.n_train = 0
        self.n_pos = 0

    def train(self, samples):
        """samples: list of (features_dict, converted_bool)."""
        self.n_train = len(samples)
        self.n_pos = sum(1 for _f, y in samples if y)
        self.base_rate = (self.n_pos / self.n_train) if self.n_train else 0.05
        br = min(max(self.base_rate, 1e-4), 1 - 1e-4)
        self.base_logodds = math.log(br / (1 - br))

        for feat in FEATURE_ORDER:
            cells = defaultdict(lambda: [0, 0])   # value -> [n, pos]
            for f, y in samples:
                v = f.get(feat, "Unknown")
                cells[v][0] += 1
                cells[v][1] += 1 if y else 0
            table = {}
            for v, (n, pos) in cells.items():
                rate = (pos + SMOOTH_ALPHA * br) / (n + SMOOTH_ALPHA)
                rate = min(max(rate, 1e-4), 1 - 1e-4)
                if v in NEUTRAL_VALUES:
                    # missing / residual: keep the observed rate for display but
                    # contribute nothing to the score.
                    table[v] = (0.0, n, pos, rate)
                    continue
                woe = math.log(rate / (1 - rate)) - self.base_logodds
                shrink = n / (n + SHRINK_K)
                table[v] = (woe * shrink, n, pos, rate)
            self.woe[feat] = table
        return self

    def contributions(self, features):
        """Return [(feature, value, woe_contribution)] for one lead."""
        out = []
        for feat in FEATURE_ORDER:
            v = features.get(feat, "Unknown")
            cell = self.woe.get(feat, {}).get(v)
            woe = cell[0] if cell else 0.0
            out.append((feat, v, woe))
        return out

    def score(self, features):
        shift = sum(w for _f, _v, w in self.contributions(features))
        p = sigmoid(self.base_logodds + WOE_DAMP * shift)
        return min(max(p, PROB_FLOOR), PROB_CEIL)


# ── priority banding, anchored to the measured base rate ─────────────────────
def priority_bands(base_rate):
    """Return ordered [(name, lo, hi)] probability bands derived from base rate."""
    br = max(base_rate, 0.01)
    hot = min(max(2.5 * br, 0.12), 0.90)
    warm = min(max(1.5 * br, 0.07), hot - 0.001)
    nur = min(max(0.75 * br, 0.03), warm - 0.001)
    return [("Hot", hot, 1.01),
            ("Warm", warm, hot),
            ("Nurture", nur, warm),
            ("Low-Intent", 0.0, nur)]


def band_for(prob, bands):
    for name, lo, hi in bands:
        if lo <= prob < hi:
            return name
    return bands[-1][0]


# ═══════════════════════════════════════════════════════════════════════════
#  LOADERS  ->  clean per-lead records keyed by 10-digit mobile
# ═══════════════════════════════════════════════════════════════════════════
def index_master(master):
    """mobile10 -> master row dict with resolved journey fields."""
    if not master or not len(master):
        return {}
    c_mob = master.col("Mobile Number", "Mobile", "Phone Number", "Phone")
    c_hist = master.col("Lead Interaction History")
    c_plat = master.col("Platforms Used")
    c_nint = master.col("Number of Interactions", "No of Interactions")
    c_first = master.col("First Enquiry Date", "First Enquiry")
    c_ref = master.col("IsReferral", "Is Referral")
    c_refn = master.col("Referrer's Name", "Referrer Name")
    c_status = master.col("Current Status", "Candidate Type")
    c_qual = master.col("Highest Qualification")
    c_exp = master.col("Total Years of Experience", "Experience")
    c_city = master.col("Current City", "City")
    c_tl = master.col("When are you planning to take admission?", "Admission Plan Time",
                      "Planning to take admission")
    c_admit = master.col("Admission Status")
    c_lead = master.col("Lead Status")
    c_back = master.col("Backout Reason", "BackOutReason")
    c_rem = master.col("Remarks", "Counsellor Notes")
    c_couns = master.col("Counselling By", "Counsellor")
    c_name = master.col("Full Name", "Name")
    c_email = master.col("Email Address", "Email")
    c_nextfu = master.col("Next Follow-Up Date", "Next Follow Up Date")
    c_gmsch = master.col("IsGoogleMeetSchedule", "Google Meet Schedule")
    c_gmdate = master.col("IsGoogleMeetScheduleDate", "Google Meet Schedule Date")
    c_wksch = master.col("IsWalkInSchedule", "Walk In Schedule")
    c_wkdate = master.col("IsWalkInScheduleDate", "Walk In Schedule Date")
    out = {}
    for r in master.rows:
        mob = digits10(r.get(c_mob)) if c_mob else ""
        if not mob:
            continue
        hist = parse_history(r.get(c_hist)) if c_hist else []
        plats = platforms_in_sequence(hist, r.get(c_plat) if c_plat else "")
        out[mob] = {
            "mobile": mob,
            "name": s(r.get(c_name)) if c_name else "",
            "email": s(r.get(c_email)) if c_email else "",
            "history": hist,
            "platforms": plats,
            "num_interactions": (s(r.get(c_nint)) if c_nint else "") or str(len(hist)),
            "first_enquiry": s(r.get(c_first)) if c_first else "",
            "referral": (yes(r.get(c_ref)) if c_ref else False) or
                        (bool(s(r.get(c_refn))) if c_refn else False),
            "work": norm_work(r.get(c_status) if c_status else "",
                              r.get(c_exp) if c_exp else ""),
            "qualification": s(r.get(c_qual)) if c_qual else "",
            "experience": s(r.get(c_exp)) if c_exp else "",
            "city": s(r.get(c_city)) if c_city else "",
            "timeline_raw": s(r.get(c_tl)) if c_tl else "",
            "admission_status": s(r.get(c_admit)) if c_admit else "",
            "lead_status": s(r.get(c_lead)) if c_lead else "",
            "backout": s(r.get(c_back)) if c_back else "",
            "remarks": s(r.get(c_rem)) if c_rem else "",
            "counsellor": s(r.get(c_couns)) if c_couns else "",
            "next_followup_raw": s(r.get(c_nextfu)) if c_nextfu else "",
            "gmeet_sched_raw": s(r.get(c_gmsch)) if c_gmsch else "",
            "gmeet_date_raw": s(r.get(c_gmdate)) if c_gmdate else "",
            "walk_sched_raw": s(r.get(c_wksch)) if c_wksch else "",
            "walk_date_raw": s(r.get(c_wkdate)) if c_wkdate else "",
        }
    return out


def load_meet_attendance(meet):
    """mobile10 -> attendance status string (from LeadAttendanceStatus)."""
    if not meet or not len(meet):
        return {}
    c_ph = meet.col("Phone Number", "Mobile Number", "Mobile", "Phone", "Contact Number")
    c_at = meet.col(*MEET_ATTEND_COL_CANDIDATES)
    if not c_ph:
        return {}
    out = {}
    for r in meet.rows:
        m = digits10(r.get(c_ph))
        if not m:
            continue
        out[m] = s(r.get(c_at)) if c_at else "Scheduled"
    return out


def load_walkin_attended(walkin):
    """Set of mobile10 present in the Walk-In form (presence == attended)."""
    if not walkin or not len(walkin):
        return set()
    c_ph = walkin.col("Phone Number", "Mobile Number", "Mobile", "Phone",
                      "Contact Number", "Contact")
    if not c_ph:
        return set()
    return {digits10(r.get(c_ph)) for r in walkin.rows if digits10(r.get(c_ph))}


def meet_state(scheduled, attendance_status):
    """attended / noshow / scheduled / none  (model + report use the same states)."""
    a = s(attendance_status).lower()
    if a:
        if any(k in a for k in ("attended", "present", "joined", "done", "completed")):
            return "attended"
        if any(k in a for k in ("no show", "no-show", "noshow", "absent",
                                "not attend", "missed")):
            return "noshow"
        return "scheduled"
    return "scheduled" if scheduled else "none"


def walkin_state(scheduled, attended):
    if attended:
        return "attended"
    return "scheduled" if scheduled else "none"


# ═══════════════════════════════════════════════════════════════════════════
#  GOOGLE MEET & WALK-IN  (scheduled from master; attendance from the forms)
# ═══════════════════════════════════════════════════════════════════════════
def meet_attended_set(meet_map):
    """Mobiles whose Google Meet attendance record (LeadAttendanceStatus) says
    they actually attended."""
    out = set()
    for mob, status in (meet_map or {}).items():
        a = s(status).lower()
        if any(k in a for k in ("attended", "present", "joined", "done", "completed")):
            out.add(mob)
    return out


def load_walkin_timestamps(walkin_new):
    """mobile10 -> latest Walk-In 'Timestamp' (datetime|None) from the Walk-In New
    tab. A mobile appearing here == the scheduled walk-in actually happened."""
    out = {}
    if not walkin_new or not len(walkin_new):
        return out
    c_ph = walkin_new.col("Mobile Number", "Mobile", "Phone Number", "Phone",
                          "Contact Number", "Contact")
    c_ts = walkin_new.col(*WALKIN_TS_COL_CANDIDATES)
    if not c_ph:
        return out
    for r in walkin_new.rows:
        m = digits10(r.get(c_ph))
        if not m:
            continue
        dt = parse_ts_mdy(r.get(c_ts)) if c_ts else None
        if m not in out or (dt and (out[m] is None or dt > out[m])):
            out[m] = dt
    return out


def _sched_classify(date_raw, today, cur_s, cur_e, ov_s, ov_e, attended,
                    attended_dt=None):
    """Classify one scheduled Google Meet / Walk-In (by its scheduled DATE) for
    the current report period. Returns None when there is no valid date or the
    date is outside this period's window (current-period OR overdue window).

      status : 'Overdue' (date in previous-period window)
             | 'Current' (date in the current-period window)
      attend : 'Attended'  (a matching attendance record exists)
             | 'Show Off'  (the meet/walk-in should already have happened but
                            no attendance record was found)
             | 'Scheduled' (upcoming — its date has not passed yet)
    """
    d = parse_followup_date(date_raw)
    if d is None:
        return None
    # Report-Date rule: only consider schedules up to the report date. A date
    # in the future (> today) is excluded completely — never Current/Overdue/
    # Attended/Show Off — even if it falls inside the current period window
    # (e.g. a Sun 22-Aug walk-in in a Mon–Sun week reported on 19-Aug).
    if today is not None and d > today:
        return None
    if ov_s is not None and ov_s <= d <= ov_e:
        status = "Overdue"
    elif cur_s is not None and cur_s <= d <= cur_e:
        status = "Current"
    else:
        return None
    if attended:
        attend = "Attended"
    elif d <= today:
        attend = "Show Off"
    else:
        attend = "Scheduled"
    return {"date": d, "date_disp": d.strftime("%d-%b-%Y"), "status": status,
            "attend": attend,
            "attended_disp": (attended_dt.strftime("%d-%b-%Y %I:%M %p")
                              if attended_dt else ""),
            "attended_date_disp": (attended_dt.strftime("%d-%b-%Y")
                                   if attended_dt else "")}


def build_meet_walk_records(master_idx, lead_by_mob, meet_attended, walk_ts,
                            model, bands, today, cur_s, cur_e, ov_s, ov_e):
    """One UNIQUE record per lead (normalised mobile) that has a Google Meet
    and/or Walk-In scheduled with a date in this report period. Google Meet /
    Walk-In scheduled + dates come from the Consolidate master; attendance from
    the Meet form (meet_attended) and the Walk-In New tab (walk_ts)."""
    walk_attended = set(walk_ts or {})
    recs = []
    for mob, m in master_idx.items():
        gm = wk = None
        if yes(m.get("gmeet_sched_raw")):
            gm = _sched_classify(m.get("gmeet_date_raw"), today, cur_s, cur_e,
                                 ov_s, ov_e, mob in meet_attended)
        if yes(m.get("walk_sched_raw")):
            wk = _sched_classify(m.get("walk_date_raw"), today, cur_s, cur_e,
                                 ov_s, ov_e, mob in walk_attended,
                                 attended_dt=(walk_ts or {}).get(mob))
        if not gm and not wk:
            continue
        lead = lead_by_mob.get(mob)
        if lead:
            rec = {"mobile": mob, "name": lead["name"],
                   "counsellor": lead["counsellor"] or "(Unassigned)",
                   "city": lead["city"], "course": lead.get("course", ""),
                   "current_status": lead.get("current_status", ""),
                   "conversion_chance": lead["conversion_chance"],
                   "priority": lead["priority"]}
        else:                                   # master-only -> score minimally
            feats = build_features(history=m.get("history", []),
                                   platforms=m.get("platforms") or [],
                                   num_interactions=m.get("num_interactions") or 1,
                                   referral=bool(m.get("referral")),
                                   work=norm_work(m.get("work", "")),
                                   timeline=norm_timeline(m.get("timeline_raw", "")),
                                   city=norm_city(m.get("city", "")),
                                   meet_state="none", walkin_state="none")
            p = model.score(feats)
            rec = {"mobile": mob, "name": m.get("name", ""),
                   "counsellor": m.get("counsellor", "") or "(Unassigned)",
                   "city": m.get("city", ""), "course": "",
                   "current_status": m.get("work", ""),
                   "conversion_chance": round(p * 100, 1),
                   "priority": band_for(p, bands)}
        rec["gm"], rec["wk"] = gm, wk
        recs.append(rec)
    return recs


def meet_walk_metrics(records, key):
    """5 Executive-Snapshot metrics for Google Meet ('gm') or Walk-In ('wk').
    total = overdue + current; attended + showoff = past-due scheduled."""
    ev = [r[key] for r in records if r.get(key)]
    return {
        "total": len(ev),
        "overdue": sum(1 for e in ev if e["status"] == "Overdue"),
        "current": sum(1 for e in ev if e["status"] == "Current"),
        "attended": sum(1 for e in ev if e["attend"] == "Attended"),
        "showoff": sum(1 for e in ev if e["attend"] == "Show Off"),
    }


# ═══════════════════════════════════════════════════════════════════════════
#  FOLLOW-UP RECONSTRUCTION  (Active + InActive history)
# ═══════════════════════════════════════════════════════════════════════════
def index_inactive_versions(inactive):
    """mobile10 -> list of archived version dicts (RecordVersion / ArchivedAt /
    RecordTimeStamp / Next Follow-Up Date)."""
    if not inactive or not len(inactive):
        return {}
    c_mob = inactive.col("Mobile Number", "Mobile", "Phone Number")
    c_ver = inactive.col("RecordVersion", "Record Version")
    c_arch = inactive.col("ArchivedAt", "Archived At")
    c_ts = inactive.col("RecordTimeStamp", "Record TimeStamp", "Timestamp")
    c_next = inactive.col("Next Follow-Up Date", "Next Follow Up Date")
    by = defaultdict(list)
    if not c_mob:
        return {}
    for r in inactive.rows:
        m = digits10(r.get(c_mob))
        if not m:
            continue
        by[m].append({
            "version": s(r.get(c_ver)) if c_ver else "",
            "archived_at": s(r.get(c_arch)) if c_arch else "",
            "timestamp": s(r.get(c_ts)) if c_ts else "",
            "next_followup": s(r.get(c_next)) if c_next else "",
        })
    return by


def follow_up_metrics(next_raw, versions, converted, lost, today, cutoff,
                      ov_start=None, ov_end=None, master_next_raw="",
                      cur_start=None, cur_end=None):
    """Reconstruct REAL follow-up completion from the versioned lead history.

    A scheduled follow-up is counted as *actually taken* only when the record
    that carried it was superseded (archived) on or after its scheduled
    'Next Follow-Up Date' — i.e. the counsellor reached the follow-up date and
    then updated the lead. A version archived *before* its scheduled date is a
    same-session correction, NOT a follow-up, and is ignored.

    This is fully data-driven: no fixed assumption about how many follow-ups a
    lead has had. Done + Remaining is kept consistent with the follow-up
    requirement (MAX_FOLLOW_UPS per open lead).

        cutoff : report period end (datetime) or None. Only follow-ups archived
                 on/before the cutoff are counted, so historical periods stay
                 accurate to their moment in time.
    """
    done_dates = set()            # distinct scheduled dates that were fulfilled
    last_done = None              # datetime the most recent one was taken
    for v in versions:
        nfu = parse_dt(v.get("next_followup"))
        arch = parse_dt(v.get("archived_at"))
        if not nfu or not arch:
            continue
        if cutoff is not None and arch > cutoff:
            continue              # happened after this report's period
        if arch.date() >= nfu.date():          # follow-up date reached, then updated
            key = nfu.date()
            if key not in done_dates:
                done_dates.add(key)
                if last_done is None or arch > last_done:
                    last_done = arch

    done = min(len(done_dates), MAX_FOLLOW_UPS)
    open_lead = not (converted or lost)
    # Remaining is dynamic: the follow-up requirement is MAX_FOLLOW_UPS per open
    # lead; closed/converted/lost leads have no remaining requirement.
    remaining = max(0, MAX_FOLLOW_UPS - done) if open_lead else 0

    next_raw = s(next_raw)
    next_dt = parse_dt(next_raw)
    has_next = bool(next_dt)

    pending = bool(open_lead and remaining > 0)
    # Effective current Next Follow-Up Date (active record, else master).
    eff_date = parse_followup_date(next_raw) or parse_followup_date(master_next_raw)
    cur_in_ov = bool(eff_date and ov_start is not None
                     and ov_start <= eff_date <= ov_end)
    cur_in_cur = bool(eff_date and cur_start is not None
                      and cur_start <= eff_date <= cur_end)

    # Actioned-away: a follow-up that was DUE in the period (overdue or current
    # window) but whose record was updated DURING the report timeframe, moving
    # its Next Follow-Up Date out of the window. Detected from the archived
    # (InActive) versions so the lead still counts — as a COMPLETED follow-up —
    # instead of silently dropping when its current date moved forward.
    actioned_ov = actioned_cur = False
    if cur_start is not None and cur_end is not None:
        for v in versions:
            arch = parse_dt(v.get("archived_at"))
            vnfu = parse_followup_date(v.get("next_followup"))
            if not arch or not vnfu:
                continue
            if not (cur_start <= arch.date() <= cur_end):
                continue                       # not updated within this timeframe
            if ov_start is not None and ov_start <= vnfu <= ov_end:
                actioned_ov = True
            elif cur_start <= vnfu <= cur_end:
                actioned_cur = True

    # Precedence: a still-open lead whose CURRENT next date is in-window first,
    # else a follow-up that was due-in-window and actioned away during the run
    # (which may even be closed today — still a completed follow-up attempt).
    if open_lead and cur_in_ov:
        status = "Overdue"
    elif open_lead and cur_in_cur:
        status = "Current"
    elif actioned_ov:
        status = "Overdue"
    elif actioned_cur:
        status = "Current"
    else:
        status = None
    if status is None and ov_start is None and ov_end is None:
        if open_lead and eff_date and eff_date < today:   # legacy no-window fallback
            status = "Overdue"
    overdue = (status == "Overdue")
    period_pending = (status == "Current")
    total_pending = status is not None
    actioned_away = bool((actioned_ov or actioned_cur)
                         and not (open_lead and (cur_in_ov or cur_in_cur)))
    next_disp = next_dt.strftime("%d-%b-%Y") if next_dt else next_raw
    return {
        "done": done,
        "remaining": remaining,
        "completion_pct": done / MAX_FOLLOW_UPS,
        "pending": pending,
        "overdue": overdue,
        "period_pending": period_pending,
        "total_pending": total_pending,
        "actioned_away": actioned_away,
        "has_next": has_next,
        "last_follow": last_done.strftime("%d-%b-%Y") if last_done else "",
        "next_follow": next_disp,
        "open": open_lead,
    }


# ═══════════════════════════════════════════════════════════════════════════
#  LEAD ASSEMBLY  (join every source into one enriched active-lead record)
# ═══════════════════════════════════════════════════════════════════════════
def assemble_leads(active, inactive, master_idx, meet_map, walkin_set, today):
    """Return list of enriched active-lead dicts (one per active record)."""
    if not active or not len(active):
        return []
    A = {
        "mob":   active.col("Mobile Number", "Mobile", "Phone Number"),
        "name":  active.col("Full Name", "Name", "Student Name"),
        "email": active.col("Email Address", "Email", "Email ID"),
        "ctype": active.col("Candidate Type", "Current Status"),
        "exp":   active.col("Total Years of Experience", "Experience"),
        "course": active.col("Course Interested In", "Course Advised",
                             "Which technology are you interested in learning?"),
        "gmeet": active.col("IsGoogleMeetSchedule", "Google Meet Schedule"),
        "gmeetd": active.col("IsGoogleMeetScheduleDate"),
        "walk":  active.col("IsWalkInSchedule", "Walk In Schedule"),
        "walkd": active.col("IsWalkInScheduleDate"),
        "city":  active.col("Current City", "City"),
        "area":  active.col("Current Area / Locality", "Area"),
        "tl":    active.col("Admission Plan Time",
                            "When are you planning to take admission?"),
        "next":  active.col("Next Follow-Up Date", "Next Follow Up Date"),
        "qual":  active.col("Highest Qualification"),
        "notes": active.col("Counsellor Notes", "Remarks"),
        "couns": active.col("Counselling By", "Counsellor"),
        "ref":   active.col("Is Referral", "IsReferral"),
        "refn":  active.col("Referrer's Name"),
        "admit": active.col("Admission Status"),
        "lead":  active.col("Lead Status"),
        "back":  active.col("BackOutReason", "Backout Reason"),
        "ts":    active.col("RecordTimeStamp", "Timestamp"),
        "goal":  active.col("Career Goal", "What is your primary goal?"),
        "company": active.col("Current Company Name"),
        "futype": active.col("Follow-Up Type", "Follow Up Type", "FollowUp Type"),
    }
    versions_idx = index_inactive_versions(inactive)
    leads = []
    for r in active.rows:
        mob = digits10(r.get(A["mob"])) if A["mob"] else ""
        if not mob:
            continue
        m = master_idx.get(mob, {})
        history = m.get("history", [])
        platforms = m.get("platforms") or []
        # referral from either source
        referral = (yes(r.get(A["ref"])) if A["ref"] else False) or \
                   (bool(s(r.get(A["refn"]))) if A["refn"] else False) or \
                   bool(m.get("referral"))
        # attendance
        gmeet_sched = yes(r.get(A["gmeet"])) if A["gmeet"] else False
        walk_sched = yes(r.get(A["walk"])) if A["walk"] else False
        mstate = meet_state(gmeet_sched, meet_map.get(mob, ""))
        wstate = walkin_state(walk_sched, mob in walkin_set)
        # status truth
        admit = s(r.get(A["admit"])) if A["admit"] else m.get("admission_status", "")
        lead_status = s(r.get(A["lead"])) if A["lead"] else m.get("lead_status", "")
        backout = s(r.get(A["back"])) if A["back"] else m.get("backout", "")
        converted = is_converted_status(admit)
        lost = is_lost_status(admit, lead_status, backout)
        # follow-ups (all-time, up to the current system time). Re-derived per
        # report period in build_report() via apply_period_followups().
        lead_versions = versions_idx.get(mob, [])
        next_raw = s(r.get(A["next"])) if A["next"] else ""
        fu = follow_up_metrics(next_raw, lead_versions, converted, lost,
                               today, None)
        # background/timeline/city/work — prefer active, fall back to master
        work = norm_work(r.get(A["ctype"]) if A["ctype"] else m.get("work", ""),
                         r.get(A["exp"]) if A["exp"] else m.get("experience", ""))
        timeline = norm_timeline(s(r.get(A["tl"])) if A["tl"] else m.get("timeline_raw", ""))
        city = norm_city(s(r.get(A["city"])) if A["city"] else m.get("city", ""))
        num_int = m.get("num_interactions") or str(max(len(history), 1))

        feats = build_features(history=history, platforms=platforms,
                               num_interactions=num_int, referral=referral,
                               work=work, timeline=timeline, city=city,
                               meet_state=mstate, walkin_state=wstate)
        # activity date for period filtering
        act_dt = (parse_dt(s(r.get(A["ts"])) if A["ts"] else "") or
                  (history[-1][1] if history else None) or
                  parse_dt(m.get("first_enquiry", "")))
        first_dt = (parse_dt(m.get("first_enquiry", "")) or
                    (history[0][1] if history else act_dt))
        leads.append({
            "mobile": mob,
            "name": (s(r.get(A["name"])) if A["name"] else "") or m.get("name", ""),
            "email": (s(r.get(A["email"])) if A["email"] else "") or m.get("email", ""),
            "counsellor": (s(r.get(A["couns"])) if A["couns"] else "") or
                          m.get("counsellor", "") or "(Unassigned)",
            "course": s(r.get(A["course"])) if A["course"] else "",
            "platforms": platforms,
            "num_interactions": num_int,
            "first_enquiry": m.get("first_enquiry", "") or
                             (first_dt.strftime("%d-%b-%Y %I:%M %p") if first_dt else ""),
            "referral": referral,
            "work": work,
            "qualification": (s(r.get(A["qual"])) if A["qual"] else "") or
                             m.get("qualification", ""),
            "experience": (s(r.get(A["exp"])) if A["exp"] else "") or m.get("experience", ""),
            "company": s(r.get(A["company"])) if A["company"] else "",
            "city": (s(r.get(A["city"])) if A["city"] else "") or m.get("city", ""),
            "area": s(r.get(A["area"])) if A["area"] else "",
            "timeline": timeline,
            "timeline_raw": (s(r.get(A["tl"])) if A["tl"] else "") or m.get("timeline_raw", ""),
            "goal": s(r.get(A["goal"])) if A["goal"] else "",
            "gmeet_sched": gmeet_sched,
            "gmeet_date": s(r.get(A["gmeetd"])) if A["gmeetd"] else "",
            "meet_state": mstate,
            "walk_sched": walk_sched,
            "walk_date": s(r.get(A["walkd"])) if A["walkd"] else "",
            "walk_state": wstate,
            "admission_status": admit,
            "lead_status": lead_status,
            "backout": backout,
            "notes": m.get("remarks", "") or (s(r.get(A["notes"])) if A["notes"] else ""),
            "converted": converted,
            "lost": lost,
            "followup": fu,
            "features": feats,
            "speed": feats["speed"],
            "history": history,
            # raw display fields required by the Priority & Actions / counsellor tabs
            "current_status": (s(r.get(A["ctype"])) if A["ctype"] else "") or
                              m.get("work", ""),
            "referral_status": "Referral" if referral else "Non Referral",
            "admission_plan": (s(r.get(A["tl"])) if A["tl"] else "") or
                              m.get("timeline_raw", ""),
            "followup_type": s(r.get(A["futype"])) if A["futype"] else "",
            "_versions": lead_versions,
            "_next_raw": next_raw,
            "_master_next": m.get("next_followup_raw", ""),
            "_record_ts": parse_dt(s(r.get(A["ts"]))) if A["ts"] else None,
            "_act_dt": act_dt,
            "_first_dt": first_dt,
        })
    return leads


def build_master_pending_leads(master_idx, exclude_mobiles, meet_map, walkin_set,
                               model, bands, today, cutoff,
                               cur_start, cur_end, ov_start, ov_end):
    """Build lead-level records for candidates that appear ONLY in the
    Consolidate master (not in the active Lead Information Result) but still
    contribute to Total Follow-Up Pending — i.e. open leads whose Next
    Follow-Up Date falls in the current-period window or the overdue window.

    This lets every counted pending lead carry full details in the detail tabs.
    Records mirror the assemble_leads() schema so the tabs render identically."""
    out = []
    for mob, m in master_idx.items():
        if not mob or mob in exclude_mobiles:
            continue
        admit = m.get("admission_status", "")
        converted = is_converted_status(admit)
        lost = is_lost_status(admit, m.get("lead_status", ""), m.get("backout", ""))
        if converted or lost:                  # pending = still-open only
            continue
        master_next = m.get("next_followup_raw", "")
        fu = follow_up_metrics("", [], converted, lost, today, cutoff,
                               ov_start, ov_end, master_next, cur_start, cur_end)
        if not fu["total_pending"]:             # not in either pending window
            continue
        history = m.get("history", [])
        platforms = m.get("platforms") or []
        referral = bool(m.get("referral"))
        work = norm_work(m.get("work", ""))
        timeline = norm_timeline(m.get("timeline_raw", ""))
        city = norm_city(m.get("city", ""))
        num_int = m.get("num_interactions") or str(max(len(history), 1))
        feats = build_features(history=history, platforms=platforms,
                               num_interactions=num_int, referral=referral,
                               work=work, timeline=timeline, city=city,
                               meet_state="none", walkin_state="none")
        first_dt = (parse_dt(m.get("first_enquiry", "")) or
                    (history[0][1] if history else None))
        lead = {
            "mobile": mob,
            "name": m.get("name", ""),
            "email": m.get("email", ""),
            "counsellor": m.get("counsellor", "") or "(Unassigned)",
            "course": "",
            "platforms": platforms,
            "num_interactions": num_int,
            "first_enquiry": m.get("first_enquiry", "") or
                             (first_dt.strftime("%d-%b-%Y %I:%M %p") if first_dt else ""),
            "referral": referral,
            "work": work,
            "qualification": m.get("qualification", ""),
            "experience": m.get("experience", ""),
            "company": "",
            "city": m.get("city", ""),
            "area": "",
            "timeline": timeline,
            "timeline_raw": m.get("timeline_raw", ""),
            "goal": "",
            "gmeet_sched": False, "gmeet_date": "", "meet_state": "none",
            "walk_sched": False, "walk_date": "", "walk_state": "none",
            "admission_status": admit,
            "lead_status": m.get("lead_status", ""),
            "backout": m.get("backout", ""),
            "notes": m.get("remarks", ""),
            "converted": converted, "lost": lost,
            "followup": fu, "features": feats, "speed": feats["speed"],
            "history": history,
            "current_status": m.get("work", ""),
            "referral_status": "Referral" if referral else "Non Referral",
            "admission_plan": m.get("timeline_raw", ""),
            "followup_type": "",
            "_versions": [], "_next_raw": "", "_master_next": master_next,
            "_record_ts": None,          # not in the Active tab -> no update signal
            "_act_dt": first_dt, "_first_dt": first_dt,
            "_master_only": True,
        }
        p = model.score(feats)
        lead["conversion_chance"] = round(p * 100, 1)
        lead["priority"] = band_for(p, bands)
        lead["next_action"] = next_best_action(lead)
        out.append(lead)
    return out


# ═══════════════════════════════════════════════════════════════════════════
#  MODEL TRAINING SET  (from master, labelled by Admission Status)
# ═══════════════════════════════════════════════════════════════════════════
def load_history_master_index():
    """Read every past 'Consolidate Sales Tracking' / 'Consolidated Master'
    export in HISTORY_DIR and index it by mobile. Used only to enlarge the
    training set; the live master's rows always take precedence."""
    idx = {}
    if not HISTORY_DIR or not os.path.isdir(HISTORY_DIR):
        return idx
    try:
        import openpyxl
    except Exception:
        return idx
    seen_files = set()
    for pattern in HISTORY_MASTER_GLOBS:
        for path in sorted(glob.glob(os.path.join(HISTORY_DIR, pattern))):
            if path in seen_files:
                continue
            seen_files.add(path)
            try:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]
                it = ws.iter_rows(values_only=True)
                header = next(it, None)
                if not header:
                    continue
                header = [s(h) or f"col_{i}" for i, h in enumerate(header)]
                n = len(header)
                rows = [{header[i]: ("" if (i >= len(r) or r[i] is None) else r[i])
                         for i in range(n)} for r in it]
                sub = index_master(Table(header, rows))
                for mob, row in sub.items():
                    idx.setdefault(mob, row)
            except Exception as e:
                print(f"  [history] skipped {os.path.basename(path)}: {e}")
    return idx


def build_training_samples(master_idx, meet_map, walkin_set):
    """Every historical master lead becomes a (features, converted) sample."""
    samples = []
    for mob, m in master_idx.items():
        history = m.get("history", [])
        platforms = m.get("platforms") or []
        gmeet_sched = False  # master rarely carries the schedule flag reliably
        mstate = meet_state(gmeet_sched, meet_map.get(mob, ""))
        wstate = walkin_state(False, mob in walkin_set)
        feats = build_features(
            history=history, platforms=platforms,
            num_interactions=m.get("num_interactions") or len(history),
            referral=bool(m.get("referral")),
            work=norm_work(m.get("work", "")),      # already normalised
            timeline=norm_timeline(m.get("timeline_raw", "")),
            city=norm_city(m.get("city", "")),
            meet_state=mstate, walkin_state=wstate)
        # work/city/timeline may be double-normalised; re-normalise safely
        feats["work"] = m.get("work", "Unknown") if m.get("work") else "Unknown"
        feats["city"] = norm_city(m.get("city", ""))
        feats["timeline"] = norm_timeline(m.get("timeline_raw", ""))
        samples.append((feats, is_converted_status(m.get("admission_status", ""))))
    return samples


# ═══════════════════════════════════════════════════════════════════════════
#  PERIOD FILTERING
# ═══════════════════════════════════════════════════════════════════════════
def in_period(lead, start, end):
    if start is None and end is None:      # Full
        return True
    d = lead.get("_act_dt") or lead.get("_first_dt")
    if not d:
        return True                        # undated active lead: always include
    return start <= d <= end


def day_bounds(d):
    return datetime.combine(d, time.min), datetime.combine(d, time.max)


def week_bounds(ref):
    mon = ref - timedelta(days=ref.weekday())
    sun = mon + timedelta(days=6)
    return datetime.combine(mon, time.min), datetime.combine(sun, time.max), mon, sun


def month_bounds(y, m):
    last = calendar.monthrange(y, m)[1]
    return (datetime.combine(date(y, m, 1), time.min),
            datetime.combine(date(y, m, last), time.max),
            date(y, m, 1), date(y, m, last))


def prev_completed_month(y, m):
    """First and last DATE of the calendar month before (y, m). Leap-safe."""
    py, pm = (y - 1, 12) if m == 1 else (y, m - 1)
    return date(py, pm, 1), date(py, pm, calendar.monthrange(py, pm)[1])


def overdue_window(label, st, en, today):
    """Overdue window (previous completed period) per report type, as dates:
        Daily   -> previous 3 completed calendar days (excludes report date)
        Weekly  -> previous completed Mon-Sun calendar week
        Monthly -> previous completed calendar month
        Manual  -> the already-passed portion of the selected range
    Returns (ov_start, ov_end) dates, or (None, None) if not resolvable."""
    if label == "Daily":
        d = st.date()
        return d - timedelta(days=3), d - timedelta(days=1)
    if label == "Weekly":
        cur_mon = st.date()                    # week_bounds start = Monday
        return cur_mon - timedelta(days=7), cur_mon - timedelta(days=1)
    if label == "Monthly":
        return prev_completed_month(st.year, st.month)
    if label == "Manual":
        return st.date(), min(en.date(), today - timedelta(days=1))
    return None, None


# ═══════════════════════════════════════════════════════════════════════════
#  AGGREGATIONS
# ═══════════════════════════════════════════════════════════════════════════
def pending_followup_keys(active, master, win_start, win_end):
    """Return the SET of unique candidate keys (normalised mobile, or
    'email:<addr>' when a record has no mobile) whose `Next Follow-Up Date`
    (date part only) falls in [win_start, win_end] inclusive, combining BOTH
    source sheets and de-duplicating by mobile first, then email.

    Fully dynamic — the window is supplied by the caller per report period.
    No hard-coded candidates."""
    keys = set()
    if win_start is None or win_end is None:
        return keys
    seen_mob, seen_email = set(), set()

    def consume(table, mob_cands, email_cands, nf_cands):
        if not table or not len(table):
            return
        c_mob = table.col(*mob_cands)
        c_email = table.col(*email_cands)
        c_nf = table.col(*nf_cands)
        if not c_nf:
            return
        for r in table.rows:
            d = parse_followup_date(r.get(c_nf))
            if not d or not (win_start <= d <= win_end):
                continue
            mob = digits10(r.get(c_mob)) if c_mob else ""
            email = norm_email(r.get(c_email)) if c_email else ""
            if mob:
                if mob in seen_mob:
                    continue
                seen_mob.add(mob)
                if email:
                    seen_email.add(email)       # link email so the other sheet dedupes
                keys.add(mob)
            elif email:
                if email in seen_email:
                    continue
                seen_email.add(email)
                keys.add("email:" + email)
            # no usable identifier -> cannot dedupe, skip

    # Lead Information Result (active) first, then the Consolidate master.
    consume(active,
            ["Mobile Number", "Mobile", "Phone Number"],
            ["Email Address", "Email", "Email ID"],
            ["Next Follow-Up Date", "Next Follow Up Date"])
    consume(master,
            ["Mobile Number", "Mobile", "Phone Number", "Phone"],
            ["Email Address", "Email"],
            ["Next Follow-Up Date", "Next Follow Up Date"])
    return keys


def summarise(leads):
    n = len(leads)
    # Executive Snapshot universe = leads still in the follow-up cycle, i.e.
    # EXCLUDING "Follow-Up Completed / Closed" (converted) and
    # "Closed (not interested / backed out)" (lost).
    incl = [l for l in leads if l["followup"]["open"]]
    # 1) Leads with Follow-Up Pending  = open leads still owing follow-ups
    leads_fu_pending = sum(1 for l in incl if l["followup"]["remaining"] > 0)
    # 2) Overdue Follow-Ups (Next Date Passed)
    fu_overdue = sum(1 for l in incl if l["followup"]["overdue"])
    # 3) Total Follow-Up Pending = open leads with a concrete scheduled next date
    total_fu_pending = sum(1 for l in incl if l["followup"]["has_next"])
    # 4) Total Follow-Up Done  = scheduled follow-ups actually taken (data-driven)
    total_fu_done = sum(l["followup"]["done"] for l in incl)
    # 5) Total Follow-Ups Remaining = remaining requirement across open leads
    total_fu_remaining = sum(l["followup"]["remaining"] for l in incl)

    gm_sched = sum(1 for l in leads if l["meet_state"] != "none")
    gm_att = sum(1 for l in leads if l["meet_state"] == "attended")
    gm_no = sum(1 for l in leads if l["meet_state"] == "noshow")
    wk_sched = sum(1 for l in leads if l["walk_state"] != "none")
    wk_att = sum(1 for l in leads if l["walk_state"] == "attended")
    conv = sum(1 for l in leads if l["converted"])
    return OrderedDict([
        ("total", n), ("incl", len(incl)),
        ("leads_fu_pending", leads_fu_pending),
        ("fu_overdue", fu_overdue),
        ("total_fu_pending", total_fu_pending),
        ("total_fu_done", total_fu_done),
        ("total_fu_remaining", total_fu_remaining),
        ("gm_sched", gm_sched), ("gm_att", gm_att), ("gm_no", gm_no),
        ("wk_sched", wk_sched), ("wk_att", wk_att), ("conv", conv),
    ])


def counsellor_rollup(leads):
    groups = defaultdict(list)
    for l in leads:
        key = re.sub(r"\s+", " ", l["counsellor"]).strip() or "(Unassigned)"
        groups[key].append(l)
    rows = []
    for cb in sorted(groups, key=lambda k: (-len(groups[k]), k)):
        g = groups[cb]
        n = len(g)
        # Follow-Up Pending column = this counsellor's Total Follow-Up Pending
        # leads (same set shown in their individual tab -> counts reconcile).
        fup_pend = sum(1 for l in g if l["followup"].get("total_pending"))
        fup_done_leads = n - fup_pend
        gm_s = sum(1 for l in g if l["meet_state"] != "none")
        gm_a = sum(1 for l in g if l["meet_state"] == "attended")
        wk_s = sum(1 for l in g if l["walk_state"] != "none")
        wk_a = sum(1 for l in g if l["walk_state"] == "attended")
        conv = sum(1 for l in g if l["converted"])
        avg_cc = sum(l["conversion_chance"] for l in g) / n if n else 0
        rows.append([
            cb, n, fup_pend, fup_done_leads, pct(fup_done_leads, n),
            gm_s, gm_a, pct(gm_a, gm_s), wk_s, wk_a, pct(wk_a, wk_s),
            conv, pct(conv, n), f"{avg_cc:.1f}%",
        ])
    return rows


# ═══════════════════════════════════════════════════════════════════════════
#  PRESENTATION  (Tab model + styling — self-contained)
# ═══════════════════════════════════════════════════════════════════════════
class Tab:
    def __init__(self, name):
        self.name = name
        self.rows = []
        self.banners = set()
        self.titles = set()
        self.headers = set()
        self.filter_headers = []
        self.kpi = set()
        # when True, suppress the automatic per-row %-traffic-light colouring so
        # colour can be driven by explicit per-cell overrides only (used by the
        # Counsellor Performance scorecard, coloured by Counsellor Performance %).
        self.no_autocolor = False
        # explicit per-cell overrides for hand-laid blocks (e.g. the side-by-side
        # Google-Meet | Walk-In panel). (row, col) -> {"bg":rgb, "fg":rgb, "bold":bool}
        self.cell_fmt = {}

    def set_fmt(self, ri, ci, bg=None, fg=None, bold=None):
        cur = self.cell_fmt.setdefault((ri, ci), {})
        if bg is not None:
            cur["bg"] = bg
        if fg is not None:
            cur["fg"] = fg
        if bold is not None:
            cur["bold"] = bold

    def banner(self, text):
        self.banners.add(len(self.rows)); self.rows.append([text])

    def title(self, text):
        self.titles.add(len(self.rows)); self.rows.append([text])

    def header(self, cols, filterable=False):
        i = len(self.rows); self.headers.add(i)
        if filterable:
            self.filter_headers.append(i)
        self.rows.append([str(c) for c in cols])

    def row(self, cols, kpi=False):
        if kpi:
            self.kpi.add(len(self.rows))
        self.rows.append(["" if c is None else c for c in cols])

    def blank(self):
        self.rows.append([])

    def width(self):
        return max((len(r) for r in self.rows), default=1)

    def _stops(self):
        return self.banners | self.titles | self.headers

    def data_span(self, hi):
        stops = self._stops(); a = hi + 1; b = a
        while b < len(self.rows):
            if b in stops or not any(str(c).strip() for c in self.rows[b]):
                break
            b += 1
        return a, b

    def fill_width(self, ri):
        if ri in self.headers:
            return len(self.rows[ri])
        for hi in self.headers:
            a, b = self.data_span(hi)
            if a <= ri < b:
                return len(self.rows[hi])
        if ri in self.banners:
            return max(self.width(), 1)
        if ri in self.titles:
            nh = sorted(h for h in self.headers if h > ri)
            return len(self.rows[nh[0]]) if nh else max(self.width(), 1)
        return max(len([c for c in self.rows[ri] if str(c).strip()]), 1)


# palette (navy/blue IntelliBI look) ─ RGB (0-1) for Sheets, HEX for xlsx
CLR_TITLE_BG = (0.11, 0.21, 0.37); CLR_TITLE_FG = (1, 1, 1)
CLR_SEC_BG = (0.17, 0.33, 0.53); CLR_SEC_FG = (1, 1, 1)
CLR_HDR_BG = (0.82, 0.87, 0.95); CLR_HDR_FG = (0.10, 0.16, 0.28)
CLR_ALT_BG = (0.955, 0.970, 0.990)
CLR_GREEN = (0.78, 0.90, 0.74); CLR_ORANGE = (1.00, 0.90, 0.66); CLR_RED = (0.97, 0.80, 0.78)
# ── FIXED lead-priority colour mapping (do NOT let charts auto-assign) ─────────
#   Hot = Green | Warm = Yellow | Nurture = Orange | Low-Intent = Red
#   Mid-tone fills chosen so dark text stays readable on every one of them.
CLR_HOT = (0.663, 0.816, 0.557)   # green   #A9D08E
CLR_WARM = (1.000, 0.851, 0.400)  # yellow  #FFD966
CLR_NUR = (0.957, 0.694, 0.514)   # orange  #F4B183
CLR_LOW = (1.000, 0.600, 0.600)   # red     #FF9999
HEX = {"TITLE": "1B355E", "TITLE_FG": "FFFFFF", "SEC": "2B547E", "SEC_FG": "FFFFFF",
       "HDR": "D2DEF2", "HDR_FG": "1A2A48", "ALT": "F4F8FD",
       "GREEN": "C6EFCE", "ORANGE": "FFEB9C", "RED": "FFC7CE",
       "HOT": "A9D08E", "WARM": "FFD966", "NUR": "F4B183", "LOW": "FF9999"}

# Fixed, order-independent mapping used everywhere a lead-priority appears.
BAND_RGB = {"Hot": CLR_HOT, "Warm": CLR_WARM, "Nurture": CLR_NUR, "Low-Intent": CLR_LOW}
BAND_HEX = {"Hot": HEX["HOT"], "Warm": HEX["WARM"], "Nurture": HEX["NUR"], "Low-Intent": HEX["LOW"]}
# Dark text for readability on the priority fills (dark on yellow, etc.).
BAND_TEXT_RGB = (0.10, 0.16, 0.28)
BAND_TEXT_HEX = "1A2A48"


def rgb_hex(rgb):
    return "".join(f"{int(round(c * 255)):02X}" for c in rgb)


def pos_rgb(pv):
    """Higher % is better (attendance): green / orange / red."""
    return CLR_GREEN if pv >= 60 else (CLR_ORANGE if pv >= 30 else CLR_RED)


def neg_rgb(pv):
    """Higher % is WORSE (no-shows / not-attended): red / orange / green."""
    return CLR_RED if pv >= 60 else (CLR_ORANGE if pv >= 30 else CLR_GREEN)


def _pct_value(cell):
    m = re.match(r"^\s*([\d.]+)\s*%\s*$", str(cell))
    try:
        return float(m.group(1)) if m else None
    except ValueError:
        return None


def _row_rgb(rowvals):
    # priority band tint takes precedence
    for v in rowvals:
        sv = str(v).strip()
        if sv in BAND_RGB:
            return BAND_RGB[sv]
    for v in rowvals:                       # else %-traffic light (higher=better)
        pv = _pct_value(v)
        if pv is not None:
            return CLR_GREEN if pv >= 60 else (CLR_ORANGE if pv >= 30 else CLR_RED)
    return None


def _row_hex(rowvals):
    for v in rowvals:
        sv = str(v).strip()
        if sv in BAND_HEX:
            return BAND_HEX[sv]
    for v in rowvals:
        pv = _pct_value(v)
        if pv is not None:
            return HEX["GREEN"] if pv >= 60 else (HEX["ORANGE"] if pv >= 30 else HEX["RED"])
    return None


# ═══════════════════════════════════════════════════════════════════════════
#  TAB BUILDERS
# ═══════════════════════════════════════════════════════════════════════════
def add_header(tab, title, period_range, gen):
    tab.banner(f"IntelliBI  -  Lead Follow-Up Analysis  |  {title}"
               f"  |  Reporting Period: {period_range}  |  Generated: {gen}")
    tab.blank()


def _p100(nn, dd):
    return (100.0 * nn / dd) if dd else 0.0


def build_summary_tab(period_label, period_range, leads, model, bands, gen,
                      period_pending=None, total_pending=None, overdue_count=None,
                      done_count=None, remaining_count=None, mix_leads=None,
                      gm_metrics=None, wk_metrics=None):
    t = Tab("Summary")
    add_header(t, period_label, period_range, gen)
    sm = summarise(leads)
    n = sm["total"]
    # "Total <period> Follow-Up Pending" = unique candidates whose Next Follow-Up
    # Date falls in the current period, combined & de-duplicated across both
    # source sheets. Falls back to the in-subset count only if not supplied.
    period_val = (period_pending if period_pending is not None
                  else sm["total_fu_pending"])
    # "Overdue Follow-Ups" = unique active/pending leads whose Next Follow-Up
    # Date fell in the previous completed period (both sheets, de-duplicated).
    overdue_val = (overdue_count if overdue_count is not None
                   else sm["fu_overdue"])
    # "Total Follow-Up Pending" = de-duplicated union of the period-pending set
    # and the overdue set (= period pending + overdue, disjoint windows).
    total_val = (total_pending if total_pending is not None
                 else period_val + overdue_val)
    ptype = period_label.split()[0]            # Daily / Weekly / Monthly / Manual
    ov_basis = {"Daily": "Next Follow-Up Date in the previous 3 days",
                "Weekly": "Next Follow-Up Date in the previous completed week",
                "Monthly": "Next Follow-Up Date in the previous completed month",
                "Manual": "Next Follow-Up Date passed within the selected range"
                }.get(ptype, "Next follow-up date passed")
    overdue_label = {"Daily": "Overdue Follow-Ups Last 3 Days",
                     "Weekly": "Overdue Follow-Ups Last Week",
                     "Monthly": "Overdue Follow-Ups Last Month",
                     "Manual": "Overdue Follow-Ups (Selected Range)"
                     }.get(ptype, "Overdue Follow-Ups (Next Date Passed)")
    total_pending_label = f"Total {ptype} Follow-Up Pending"

    # ── 1. Executive Snapshot (5 follow-up metrics; excludes closed/converted) ──
    t.title("Executive Snapshot")
    t.header(["Metric", "Value", "Basis"])
    t.row(["Total Follow-Up Pending", total_val,
           f"Total {ptype} Follow-Up Pending + Overdue "
           "(unique candidates, de-duplicated)"], kpi=True)
    t.row([overdue_label, overdue_val,
           ov_basis + " — still pending, active, de-duplicated"], kpi=True)
    t.row([total_pending_label, period_val,
           "Unique candidates with a Next Follow-Up Date in this period "
           "(both sheets, de-duplicated)"], kpi=True)
    done_val = done_count if done_count is not None else sm["total_fu_done"]
    remaining_val = (remaining_count if remaining_count is not None
                     else max(total_val - done_val, 0))
    t.row(["Total Follow-Up Done", done_val,
           "Pending leads updated in this timeframe (RecordTimeStamp) — "
           "a follow-up attempt was made"], kpi=True)
    t.row(["Total Follow-Ups Remaining", remaining_val,
           "Pending leads with no qualifying follow-up update yet "
           "(Done + Remaining = Total Follow-Up Pending)"], kpi=True)
    t.blank()

    # ── 2. Google Meet & Walk-In — scheduled / attendance for THIS period ───────
    # Sourced from the Consolidate master (IsGoogleMeetSchedule / IsWalkInSchedule
    # + their dates) and the attendance forms. Total = Overdue + Current (period);
    # Attended + Show Off = the past-due scheduled ones. Reconciles 1:1 with the
    # detailed "Google Meet & Walk-In" tab. Each row shows a % and is colour-coded
    # (green = good / red = bad) via explicit per-row fills.
    ov_word = {"Daily": "Last 3 Days", "Weekly": "Last Week",
               "Monthly": "Last Month", "Manual": "(Selected Range)"}.get(ptype, "")

    # Colour theme aligned with the "Lead Priority Mix" section: the SAME vivid
    # band palette (green / yellow / red), dark-navy readable text, non-bold.
    def _mix_fill(pv, negative=False):
        if negative:                       # higher % is worse (overdue / show-off)
            return CLR_LOW if pv >= 60 else (CLR_WARM if pv >= 30 else CLR_HOT)
        return CLR_HOT if pv >= 60 else (CLR_WARM if pv >= 30 else CLR_LOW)

    def _paint_row(ri, pv, negative=False, neutral=False):
        rgb = CLR_ALT_BG if neutral else _mix_fill(pv, negative)
        for ci in range(4):
            t.set_fmt(ri, ci, bg=rgb, fg=BAND_TEXT_RGB, bold=False)

    def meet_walk_section(title, mtr, kind, attended_basis):
        past = mtr["attended"] + mtr["showoff"]          # scheduled that should have happened
        t.title(f"{title} — Scheduled & Attendance")
        t.header(["Metric", "Value", "%", "Basis"])
        r0 = len(t.rows)
        t.row([f"Total {kind} Scheduled", mtr["total"], "",
               "Scheduled in this period (Overdue + Current, de-duplicated)"], kpi=True)
        t.row([f"Overdue {kind} Scheduled {ov_word}".strip(), mtr["overdue"],
               pct(mtr["overdue"], mtr["total"]),
               "Scheduled date in the previous completed period"], kpi=True)
        t.row([f"Total {ptype} {kind} Scheduled", mtr["current"],
               pct(mtr["current"], mtr["total"]),
               "Scheduled date in the current period"], kpi=True)
        t.row([f"Total {kind} Attended", mtr["attended"], pct(mtr["attended"], past),
               attended_basis], kpi=True)
        t.row([f"Total {kind} Show Off", mtr["showoff"], pct(mtr["showoff"], past),
               "Should have happened (date passed) but not attended"], kpi=True)
        # colour: overdue = negative (backlog), current = positive (fresh),
        # attended = positive, show-off = negative. Total row = neutral tint
        # (kept consistent with the coloured rows, not left plain white).
        _paint_row(r0 + 0, 0, neutral=True)
        _paint_row(r0 + 1, _p100(mtr["overdue"], mtr["total"]), negative=True)
        _paint_row(r0 + 2, _p100(mtr["current"], mtr["total"]))
        _paint_row(r0 + 3, _p100(mtr["attended"], past))
        _paint_row(r0 + 4, _p100(mtr["showoff"], past), negative=True)
        t.blank()

    if gm_metrics is not None:
        meet_walk_section("Google Meet", gm_metrics, "Google Meet",
                          "LeadAttendanceStatus = Attended (% of past-due scheduled)")
    if wk_metrics is not None:
        meet_walk_section("Walk-In", wk_metrics, "Walk-In",
                          "Mobile found in the Walk-In New tab (% of past-due scheduled)")

    # ── 3. Lead Priority Mix (by Conversion Chance %) ───────────────────────────
    # Actual count of UNIQUE leads in the FINAL applicable dataset for this report
    # period (the same Total Follow-Up Pending leads used everywhere else), per
    # priority band, with the existing mobile/email de-duplication. Counts are
    # computed directly from the leads — never percentage-derived or cached — and
    # always reconcile: Hot + Warm + Nurture + Low-Intent = total applicable
    # unique leads (= Total Follow-Up Pending).
    mix_src = mix_leads if mix_leads is not None else leads
    t.title("Lead Priority Mix (by Conversion Chance %)")
    t.header(["Priority", "Leads", "% of Active", "Conversion-Chance Range", "Avg Chance"])
    seen_pm, uniq_pm = set(), []
    for l in mix_src:
        k = digits10(l.get("mobile", "")) or ("email:" + norm_email(l.get("email", "")))
        if not k or k in seen_pm:
            continue
        seen_pm.add(k); uniq_pm.append(l)
    n_pm = len(uniq_pm)
    for name, lo, hi in bands:
        grp = [l for l in uniq_pm if l["priority"] == name]
        rng = (f"{lo*100:.0f}%+" if hi > 1 else f"{lo*100:.0f}% – {hi*100:.0f}%")
        avg = (sum(l["conversion_chance"] for l in grp) / len(grp)) if grp else 0
        t.row([name, len(grp), pct(len(grp), n_pm), rng, f"{avg:.1f}%"])
    return t


def clean_history(history):
    """One clean, readable line per interaction, oldest first:
        '01-Aug-2026 03:40 PM  ·  Walk-In  ·  ArshKhan Pathan'."""
    if not history:
        return ""
    out = []
    for src, dt, cby in history:
        parts = [dt.strftime("%d-%b-%Y %I:%M %p"), s(src) or "—"]
        if s(cby) and s(cby) not in ("-", "—"):
            parts.append(s(cby))
        out.append("  ·  ".join(parts))
    return "\n".join(out)


def build_priority_tab(leads, gen, period_label, period_range):
    t = Tab("Priority & Actions")
    add_header(t, period_label + " — what to act on first", period_range, gen)
    t.title(f"Total Follow-Up Pending — leads to act on  ({len(leads)}, "
            "ranked by Conversion Chance %)")
    t.header(["Rank", "Priority", "Conversion Chance %", "Lead Name", "Mobile",
              "Current City", "Referral Status", "Current Status", "Experience",
              "Notes / Remarks", "Lead Interaction History",
              "Number of Interactions", "Follow-Up Type", "Counsellor",
              "Next Best Action", "Follow-Ups (Done/Rem)", "Last Follow-Up",
              "Next Follow-Up", "Follow-Up Status", "Follow-Up Completion Status",
              "Meet", "Walk-In", "Timeline"], filterable=True)
    ranked = sorted(leads, key=lambda l: -l["conversion_chance"])
    for i, l in enumerate(ranked, 1):
        fu = l["followup"]
        rank = l.get("_rank", i)
        t.row([rank, l["priority"], f"{l['conversion_chance']:.1f}%", l["name"],
               l["mobile"], l["city"], l["referral_status"], l["current_status"],
               l["experience"], l["notes"],
               clean_history(l["history"]), l["num_interactions"],
               l["followup_type"], l["counsellor"], l["next_action"],
               f"{fu['done']}/{fu['remaining']}", fu["last_follow"],
               fu["next_follow"], l.get("_followup_status", ""),
               l.get("_completion_status", ""),
               l["meet_state"], l["walk_state"], l["timeline"]])
    return t


def build_meetwalk_tab(records, gen, period_label, period_range):
    """Detailed Google Meet & Walk-In tab — one row per UNIQUE scheduled lead,
    with lead-level columns (like Priority & Actions) plus the Google Meet and
    Walk-In scheduled / date / status / attendance flags. Ranked by Conversion
    Chance %."""
    t = Tab("Google Meet & Walk-In")
    add_header(t, period_label + " — Google Meet & Walk-In", period_range, gen)
    t.title("Google Meet & Walk-In — scheduled leads  "
            f"({len(records)}, ranked by Conversion Chance %)")
    t.header(["Rank", "Priority", "Conversion Chance %", "Lead Name", "Mobile",
              "Current City", "Counsellor", "Current Status", "Course Interested",
              "Google Meet Scheduled", "Google Meet Scheduled Date",
              "Google Meet Status", "Google Meet Attendance",
              "Walk-In Scheduled", "Walk-In Scheduled Date",
              "Walk-In Status", "Walk-In Attendance",
              "Walk-In Attended Date"], filterable=True)
    ranked = sorted(records, key=lambda r: -r["conversion_chance"])
    for i, r in enumerate(ranked, 1):
        gm, wk = r.get("gm"), r.get("wk")
        t.row([i, r["priority"], f"{r['conversion_chance']:.1f}%", r["name"],
               r["mobile"], r["city"], r["counsellor"], r["current_status"],
               r["course"],
               "Yes" if gm else "No", gm["date_disp"] if gm else "",
               gm["status"] if gm else "", gm["attend"] if gm else "",
               "Yes" if wk else "No", wk["date_disp"] if wk else "",
               wk["status"] if wk else "", wk["attend"] if wk else "",
               wk["attended_date_disp"] if wk else ""])
    return t


# ── Counsellor Performance scorecard ──────────────────────────────────────────
# Final weightage (must total 100%). Follow-Up carries the most weight (direct
# execution); Walk-In outweighs Google Meet (stronger physical engagement);
# Conversion Chance contributes but must not dominate (also lead-quality driven).
PERF_WEIGHTS = {"fu": 40.0, "wk": 25.0, "gm": 15.0, "cc": 20.0}

# 4-tier colour scheme for Counsellor Performance % (RGB 0-1 fills, dark text).
#   Green = Excellent | Orange = Good | Yellow = Needs Improvement | Red = Poor
def _perf_tier(pv):
    if pv >= 75:
        return "Excellent", CLR_HOT      # green
    if pv >= 60:
        return "Good", CLR_NUR           # orange
    if pv >= 40:
        return "Needs Improvement", CLR_WARM   # yellow
    return "Poor", CLR_LOW               # red


def _norm_couns(name):
    return re.sub(r"\s+", " ", str(name or "")).strip() or "(Unassigned)"


# ── counsellor-name canonicalisation ─────────────────────────────────────────
# Different spellings/casings of the SAME counsellor (e.g. "ArshKhan Pathan" vs
# "Arshkhan Pathan") must collapse to ONE canonical display name, otherwise the
# per-counsellor grouping produces two groups → two tabs.  Google Sheets treats
# sheet titles case-insensitively, so the second variant surfaced as a bogus
# "Arshkhan Pathan1" tab on upload.  Explicit aliases below force a specific
# canonical spelling; any other case/whitespace variants merge to the most
# common spelling actually seen in the data.
COUNSELLOR_ALIASES = {
    # fold-key (lower-cased, whitespace-collapsed)  ->  canonical display name
    "arshkhan pathan": "ArshKhan Pathan",
}


def _couns_fold(name):
    """Case- and whitespace-insensitive grouping key for a counsellor name."""
    return re.sub(r"\s+", " ", str(name or "")).strip().casefold()


def build_counsellor_canon_map(names):
    """Return {fold_key: canonical_display} built from all counsellor names.

    Explicit COUNSELLOR_ALIASES win; otherwise the canonical display is the most
    frequently occurring original spelling (deterministic tie-break)."""
    forms = defaultdict(Counter)
    for raw in names:
        disp = re.sub(r"\s+", " ", str(raw or "")).strip()
        if not disp or disp == "(Unassigned)":
            continue
        forms[disp.casefold()][disp] += 1
    canon = {}
    for fold, counter in forms.items():
        if fold in COUNSELLOR_ALIASES:
            canon[fold] = COUNSELLOR_ALIASES[fold]
        else:
            # most common spelling; tie -> alphabetically first for stability
            canon[fold] = min(counter.items(), key=lambda kv: (-kv[1], kv[0]))[0]
    # aliases apply even if that spelling never literally appeared in the data
    for fold, disp in COUNSELLOR_ALIASES.items():
        canon.setdefault(fold, disp)
    return canon


def canon_couns_display(name, canon_map):
    """Canonical display name for a single counsellor value."""
    disp = re.sub(r"\s+", " ", str(name or "")).strip()
    if not disp:
        return "(Unassigned)"
    return canon_map.get(disp.casefold(), disp)


def apply_counsellor_canon(leads, canon_map):
    """Rewrite every lead's 'counsellor' field to its canonical display name."""
    for l in leads:
        l["counsellor"] = canon_couns_display(l.get("counsellor"), canon_map)


_STRONG_PHRASE = {"fu": "follow-up completion", "wk": "Walk-In attendance",
                  "gm": "Google Meet attendance", "cc": "lead conversion potential"}
_WEAK_SENTENCE = {
    "fu": "Improve pending follow-up closure",
    "wk": "Focus on converting scheduled Walk-Ins into actual visits",
    "gm": "Google Meet attendance needs improvement",
    "cc": "Lead nurturing needs improvement based on conversion chance"}


def _doing_good(sc):
    """Short, value-driven comment on the counsellor's strongest APPLICABLE KPIs
    (N/A metrics — no Walk-In / Meet scheduled — are never mentioned)."""
    applic = sorted([(k, v) for k, v in sc["metrics"].items() if v is not None],
                    key=lambda kv: -kv[1])
    strong = [(k, v) for k, v in applic if v >= 60][:2]
    if strong:
        parts = [f"{_STRONG_PHRASE[k]} ({v:.0f}%)" for k, v in strong]
        return "Strong " + " and ".join(parts) + "."
    if applic:
        k, v = applic[0]
        return f"Relatively strongest in {_STRONG_PHRASE[k]} ({v:.0f}%)."
    return "Not enough activity to assess strengths."


def _weak_sentence(k, sc):
    """Weakness phrasing — for a 0% Google Meet / Walk-In that comes from NOTHING
    being scheduled, say so rather than implying scheduled events went unattended."""
    if k == "gm" and sc.get("gm_s", 0) == 0:
        return "No Google Meets scheduled — schedule and drive attendance"
    if k == "wk" and sc.get("wk_s", 0) == 0:
        return "No Walk-Ins scheduled — schedule and drive Walk-In attendance"
    return _WEAK_SENTENCE[k]


def _area_improve(sc):
    """Short, value-driven comment on the counsellor's weakest APPLICABLE KPIs."""
    weak = sorted([(k, v) for k, v in sc["metrics"].items()
                   if v is not None and v < 60], key=lambda kv: kv[1])[:2]
    if weak:
        out = []
        for i, (k, v) in enumerate(weak):
            out.append(_weak_sentence(k, sc) + (f" ({v:.0f}%)." if i == 0 else "."))
        return " ".join(out)
    return "Maintain current performance; no major gaps."


def counsellor_scorecard(pending_leads, meetwalk_records):
    """Per-counsellor scorecard built from the SAME finalized datasets the other
    tabs use (single source of truth), so every number reconciles:

      * Follow-Up — from the Total Follow-Up Pending dataset (the exact set that
        drives Summary / Priority & Actions / each Counsellor Individual tab):
          Follow-Up Pending = that counsellor's Total Follow-Up Pending leads
                              (== their Counsellor Individual "Total Follow-Up
                              Pending"; sums to the Summary total),
          Follow-Up Done    = of those, the ones completed within the timeframe
                              (finalized _completion_status; sums to Summary Done),
          Follow-Up Done %  = Done / Pending.
      * Google Meet / Walk-In — from the finalized meet/walk records (future-
        excluded, deduped), so column totals match the "Google Meet & Walk-In"
        tab. 0 scheduled -> 0% attendance (counts fully in the score), not N/A.
      * Avg Conversion Chance % — over the same pending leads (matches the
        Counsellor Individual "Average Conversion Chance").

    Weighted performance redistributes any N/A KPI's weight across the remaining
    applicable ones. Returns dicts (unranked)."""
    lead_groups = defaultdict(list)
    for l in pending_leads:
        lead_groups[_norm_couns(l["counsellor"])].append(l)
    mw_groups = defaultdict(list)
    for r in (meetwalk_records or []):
        mw_groups[_norm_couns(r.get("counsellor"))].append(r)

    out = []
    for cb in set(lead_groups) | set(mw_groups):
        g = lead_groups.get(cb, [])
        recs = mw_groups.get(cb, [])
        pend = len(g)                       # Total Follow-Up Pending for counsellor
        done = sum(1 for l in g
                   if l.get("_completion_status", "Follow-Up Pending")
                   != "Follow-Up Pending")  # completed within timeframe (finalized)
        fu_pct = (100.0 * done / pend) if pend else None          # N/A when none pending
        # Google Meet / Walk-In from the finalized meet/walk records (so column
        # totals reconcile 1:1 with the "Google Meet & Walk-In" tab).
        gm_s = sum(1 for r in recs if r.get("gm"))
        gm_a = sum(1 for r in recs if r.get("gm") and r["gm"]["attend"] == "Attended")
        wk_s = sum(1 for r in recs if r.get("wk"))
        wk_a = sum(1 for r in recs if r.get("wk") and r["wk"]["attend"] == "Attended")
        # 0 scheduled -> 0% (counts fully in the score with its weight), NOT N/A.
        gm_pct = (100.0 * gm_a / gm_s) if gm_s else 0.0
        wk_pct = (100.0 * wk_a / wk_s) if wk_s else 0.0
        if pend:
            cc = sum(l["conversion_chance"] for l in g) / pend
        elif recs:
            cc = sum(r["conversion_chance"] for r in recs) / len(recs)
        else:
            cc = 0.0
        metrics = {"fu": fu_pct, "wk": wk_pct, "gm": gm_pct, "cc": cc}
        applic = {k: v for k, v in metrics.items() if v is not None}
        tw = sum(PERF_WEIGHTS[k] for k in applic)               # redistribute N/A weight
        perf = sum(v * PERF_WEIGHTS[k] / tw for k, v in applic.items()) if tw else 0.0
        out.append({
            "counsellor": cb, "n": pend, "fup_pend": pend, "fup_done": done,
            "fu_pct": fu_pct, "gm_s": gm_s, "gm_a": gm_a, "gm_pct": gm_pct,
            "wk_s": wk_s, "wk_a": wk_a, "wk_pct": wk_pct,
            "cc": cc, "perf": perf, "metrics": metrics,
        })
    return out


def build_counsellor_tab(leads, gen, period_label, period_range, meetwalk_records=None):
    t = Tab("Counsellor Performance")
    t.no_autocolor = True          # colour driven only by Counsellor Performance %
    add_header(t, period_label + " — counsellor scorecard (in %)", period_range, gen)
    t.title("Counsellor Performance — ranked by Performance %  "
            "(Follow-Up 40% · Walk-In 25% · Google Meet 15% · Conversion 20%; "
            "N/A metrics reweighted)")
    hdr = ["Rank", "Counsellor", "Follow-Up Pending", "Follow-Up Done",
           "Follow-Up Done %", "GMeet Scheduled", "GMeet Attended",
           "GMeet Attendance %", "Walk-In Scheduled", "Walk-In Attended",
           "Walk-In Attendance %", "Avg. Conversion Chance %",
           "Counsellor Performance %", "Performance Status",
           "Doing Good", "Area of Improvement"]
    t.header(hdr, filterable=True)
    perf_col = hdr.index("Counsellor Performance %")
    stat_col = hdr.index("Performance Status")

    cards = counsellor_scorecard(leads, meetwalk_records)
    # Rank by Performance % (highest = Rank 1); tie-breakers: Follow-Up Done %,
    # Walk-In Att %, Google Meet Att %, Avg Conversion Chance % (N/A -> lowest).
    def _tb(sc):
        m = sc["metrics"]
        val = lambda k: (m[k] if m[k] is not None else -1.0)
        return (-sc["perf"], -val("fu"), -val("wk"), -val("gm"), -val("cc"),
                sc["counsellor"])
    cards.sort(key=_tb)

    def fmtp(v):
        return "N/A" if v is None else f"{v:.1f}%"

    for rank, sc in enumerate(cards, 1):
        status, rgb = _perf_tier(sc["perf"])
        ri = len(t.rows)
        t.row([rank, sc["counsellor"], sc["fup_pend"], sc["fup_done"],
               fmtp(sc["fu_pct"]), sc["gm_s"], sc["gm_a"], fmtp(sc["gm_pct"]),
               sc["wk_s"], sc["wk_a"], fmtp(sc["wk_pct"]),
               f"{sc['cc']:.1f}%", f"{sc['perf']:.1f}%", status,
               _doing_good(sc), _area_improve(sc)])
        # colour the ENTIRE row by the Counsellor Performance % 4-tier scheme;
        # the Performance % + Status cells are additionally bold.
        for ci in range(len(hdr)):
            t.set_fmt(ri, ci, bg=rgb, fg=BAND_TEXT_RGB,
                      bold=(ci in (perf_col, stat_col)))
    return t


def _safe_tab_name(base, used):
    name = re.sub(r"[:\\/?*\[\]]", " ", base).strip()[:31] or "Counsellor"
    cand = name
    k = 2
    while cand in used:
        suffix = f" ({k})"
        cand = name[:31 - len(suffix)] + suffix
        k += 1
    used.add(cand)
    return cand


def build_counsellor_detail_tabs(leads, gen, period_label, period_range):
    """One tab per counsellor: a counsellor summary followed by that
    counsellor's individual leads (lead -> current status -> interaction
    history -> follow-up progress -> next follow-up -> meet/walk-in ->
    next best action). Returns an OrderedDict {tab_name: Tab}."""
    groups = defaultdict(list)
    for l in leads:
        key = re.sub(r"\s+", " ", l["counsellor"]).strip() or "(Unassigned)"
        groups[key].append(l)

    out = OrderedDict()
    used = {"Summary", "Priority & Actions", "Counsellor Performance",
            "Lead Details", "Platform Quality", "Conversion Model"}
    ordered = sorted(groups, key=lambda k: (-len(groups[k]), k))
    for cb in ordered[:MAX_COUNSELLOR_TABS]:
        # sort by the SAME Priority & Actions rank (so ranks match across tabs);
        # fall back to conversion chance if a rank was not assigned.
        g = sorted(groups[cb],
                   key=lambda x: x.get("_rank", 10**9 - int(x["conversion_chance"] * 10)))
        n = len(g)
        incl = [l for l in g if l["followup"]["open"]]
        # Total Follow-Up Pending for this counsellor = the same set counted in
        # Counsellor Performance; every one of these leads is listed below.
        fu_pending = sum(1 for l in g if l["followup"].get("total_pending"))
        fu_done = sum(l["followup"]["done"] for l in incl)
        fu_remaining = sum(l["followup"]["remaining"] for l in incl)
        overdue = sum(1 for l in incl if l["followup"]["overdue"])
        gm_s = sum(1 for l in g if l["meet_state"] != "none")
        gm_a = sum(1 for l in g if l["meet_state"] == "attended")
        wk_s = sum(1 for l in g if l["walk_state"] != "none")
        wk_a = sum(1 for l in g if l["walk_state"] == "attended")
        conv = sum(1 for l in g if l["converted"])
        avg_cc = sum(l["conversion_chance"] for l in g) / n if n else 0

        t = Tab(_safe_tab_name(cb, used))
        add_header(t, f"Counsellor — {cb}", period_range, gen)
        t.title(f"Summary — {cb}")
        t.header(["Metric", "Value", "%"])
        t.row(["Total Leads", n, ""], kpi=True)
        t.row(["Total Follow-Up Pending", fu_pending, pct(fu_pending, n)], kpi=True)
        t.row(["Overdue Follow-Ups (Next Date Passed)", overdue, pct(overdue, len(incl))])
        t.row(["Total Follow-Up Done", fu_done, ""])
        t.row(["Total Follow-Ups Remaining", fu_remaining, ""])
        t.row(["Google Meets Scheduled / Attended", f"{gm_s}/{gm_a}", pct(gm_a, gm_s)])
        t.row(["Walk-Ins Scheduled / Attended", f"{wk_s}/{wk_a}", pct(wk_a, wk_s)])
        t.row(["Converted (Admission Confirmed)", conv, pct(conv, n)])
        t.row(["Average Conversion Chance", f"{avg_cc:.1f}%", ""], kpi=True)
        t.blank()

        t.title(f"Total Follow-Up Pending Leads — {cb}  ({fu_pending})")
        t.header(["Rank", "Priority", "Conversion Chance %", "Lead Name", "Mobile",
                  "Current Status", "Lead Interaction History",
                  "Follow-Ups (Done/Rem)", "Last Follow-Up", "Next Follow-Up",
                  "Follow-Up Status", "Follow-Up Completion Status",
                  "Meet", "Walk-In", "Next Best Action"], filterable=True)
        for l in g:
            fu = l["followup"]
            t.row([l.get("_rank", ""), l["priority"], f"{l['conversion_chance']:.1f}%",
                   l["name"], l["mobile"], l["current_status"],
                   clean_history(l["history"]),
                   f"{fu['done']}/{fu['remaining']}", fu["last_follow"],
                   fu["next_follow"], l.get("_followup_status", ""),
                   l.get("_completion_status", ""),
                   l["meet_state"], l["walk_state"], l["next_action"]])
        out[t.name] = t
    return out


def build_model_tab(model, bands, gen):
    t = Tab("Conversion Model")
    add_header(t, "How Conversion Chance % is calculated", "All history", gen)
    t.title("Model basis")
    t.header(["Metric", "Value"])
    t.row(["Training leads (historical)", model.n_train])
    t.row(["Converted in training (Admission Confirmed)", model.n_pos])
    t.row(["Base conversion rate", f"{model.base_rate*100:.2f}%"])
    t.row(["Method", "Weight-of-evidence (naive-Bayes log-odds) + shrinkage"])
    t.blank()
    t.title("Priority bands (anchored to base rate)")
    t.header(["Priority", "Conversion-Chance Range"])
    for name, lo, hi in bands:
        t.row([name, (f"{lo*100:.0f}%+" if hi > 1 else f"{lo*100:.0f}% – {hi*100:.0f}%")])
    t.blank()
    t.title("What moves the score — each factor's learned effect")
    t.header(["Factor", "Value", "Leads (hist.)", "Converted", "Hist. Rate", "Effect on odds"])
    for feat in FEATURE_ORDER:
        table = model.woe.get(feat, {})
        for v, (woe, n, pos, rate) in sorted(table.items(), key=lambda kv: -kv[1][0]):
            if n == 0:
                continue
            eff = ("↑ strong" if woe > 0.7 else "↑" if woe > 0.15 else
                   "↓ strong" if woe < -0.7 else "↓" if woe < -0.15 else "· neutral")
            t.row([FEATURE_LABEL[feat], v, n, pos, f"{rate*100:.1f}%", eff])
    return t


# ── next-best-action heuristic (decision support) ────────────────────────────
def next_best_action(l):
    fu = l["followup"]
    if l["converted"]:
        return "Converted — onboard / collect docs"
    if l["lost"]:
        return "Closed (not interested / backed out)"
    if fu["overdue"]:
        return "OVERDUE follow-up — call today"
    if l["meet_state"] == "noshow":
        return "Meet no-show — reschedule Google Meet"
    if l["walk_state"] == "scheduled":
        return "Walk-In scheduled — confirm attendance"
    if l["meet_state"] == "scheduled":
        return "Meet scheduled — confirm attendance"
    if l["priority"] in ("Hot", "Warm") and fu["remaining"] > 0:
        return "High chance — prioritise next follow-up"
    if fu["remaining"] == 0:
        return "Follow-up limit reached — decide/close"
    if fu["next_follow"]:
        return f"Follow-up due {fu['next_follow']}"
    return "Schedule next follow-up"


# ═══════════════════════════════════════════════════════════════════════════
#  WORKBOOK WRITE  (Google Sheet + local xlsx)
# ═══════════════════════════════════════════════════════════════════════════
def _fmt_range(sid, r0, r1, c0, c1, bold=False, size=None, bg=None, fg=None, wrap=False):
    tf = {"bold": bold}
    if size:
        tf["fontSize"] = size
    if fg:
        tf["foregroundColor"] = {"red": fg[0], "green": fg[1], "blue": fg[2]}
    fmt = {"textFormat": tf}
    if bg:
        fmt["backgroundColor"] = {"red": bg[0], "green": bg[1], "blue": bg[2]}
    if wrap:
        fmt["wrapStrategy"] = "WRAP"
    fields = "userEnteredFormat(textFormat,backgroundColor" + \
             (",wrapStrategy" if wrap else "") + ")" if bg else \
             "userEnteredFormat(textFormat" + (",wrapStrategy" if wrap else "") + ")"
    return {"repeatCell": {
        "range": {"sheetId": sid, "startRowIndex": r0, "endRowIndex": r1,
                  "startColumnIndex": c0, "endColumnIndex": c1},
        "cell": {"userEnteredFormat": fmt}, "fields": fields}}


def write_workbook(sheets, spreadsheet_id, tabs):
    meta = sheets.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    existing = {sh["properties"]["title"]: sh["properties"]["sheetId"] for sh in meta["sheets"]}
    reqs = [{"addSheet": {"properties": {"title": "__tmp__"}}}] if "__tmp__" not in existing else []
    for name in tabs:
        if name not in existing:
            reqs.append({"addSheet": {"properties": {"title": name}}})
    if reqs:
        sheets.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body={"requests": reqs}).execute()
        meta = sheets.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        existing = {sh["properties"]["title"]: sh["properties"]["sheetId"] for sh in meta["sheets"]}
    del_reqs = [{"deleteSheet": {"sheetId": sid}} for nm, sid in existing.items()
                if nm not in tabs and nm != "__tmp__"]
    if del_reqs:
        sheets.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body={"requests": del_reqs}).execute()
    meta = sheets.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    existing = {sh["properties"]["title"]: sh["properties"]["sheetId"] for sh in meta["sheets"]}

    sheets.spreadsheets().values().batchClear(
        spreadsheetId=spreadsheet_id,
        body={"ranges": ["'%s'!A1:ZZ100000" % n for n in tabs]}).execute()
    data = []
    for name, tab in tabs.items():
        w = tab.width()
        data.append({"range": "'%s'!A1" % name,
                     "values": [row + [""] * (w - len(row)) for row in tab.rows]})
    sheets.spreadsheets().values().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"valueInputOption": "RAW", "data": data}).execute()

    order_reqs, fmt_reqs = [], []
    for idx, (name, tab) in enumerate(tabs.items()):
        sid = existing[name]; maxw = max(tab.width(), 1)
        fh0 = tab.filter_headers[0] if tab.filter_headers else None
        frozen = (fh0 + 1) if (fh0 is not None and fh0 <= 5) else 1
        order_reqs.append({"updateSheetProperties": {
            "properties": {"sheetId": sid, "index": idx}, "fields": "index"}})
        order_reqs.append({"updateSheetProperties": {
            "properties": {"sheetId": sid, "gridProperties": {"frozenRowCount": frozen}},
            "fields": "gridProperties.frozenRowCount"}})
        for ri in tab.banners:
            fmt_reqs.append(_fmt_range(sid, ri, ri + 1, 0, tab.fill_width(ri),
                                       bold=True, size=13, bg=CLR_TITLE_BG, fg=CLR_TITLE_FG))
        for ri in tab.titles:
            fmt_reqs.append(_fmt_range(sid, ri, ri + 1, 0, tab.fill_width(ri),
                                       bold=True, size=12, bg=CLR_SEC_BG, fg=CLR_SEC_FG))
        for ri in tab.headers:
            fmt_reqs.append(_fmt_range(sid, ri, ri + 1, 0, tab.fill_width(ri),
                                       bold=True, bg=CLR_HDR_BG, fg=CLR_HDR_FG))
        for hi in tab.headers:
            a, b = tab.data_span(hi); w = len(tab.rows[hi])
            for k, ri in enumerate(range(a, b)):
                rgb = None if tab.no_autocolor else _row_rgb(tab.rows[ri])
                bold = ri in tab.kpi
                if rgb:
                    is_band = any(str(v).strip() in BAND_RGB for v in tab.rows[ri])
                    fg = BAND_TEXT_RGB if is_band else None
                    fmt_reqs.append(_fmt_range(sid, ri, ri + 1, 0, w,
                                               bold=bold, bg=rgb, fg=fg))
                elif k % 2 == 1:
                    fmt_reqs.append(_fmt_range(sid, ri, ri + 1, 0, w, bold=bold, bg=CLR_ALT_BG))
                elif bold:
                    fmt_reqs.append(_fmt_range(sid, ri, ri + 1, 0, w, bold=True))
        for (ri, ci), fm in tab.cell_fmt.items():          # explicit per-cell overrides
            fmt_reqs.append(_fmt_range(sid, ri, ri + 1, ci, ci + 1,
                                       bold=fm.get("bold", False),
                                       bg=fm.get("bg"), fg=fm.get("fg")))
        if tab.filter_headers:
            hi = tab.filter_headers[0]; a, b = tab.data_span(hi)
            fmt_reqs.append({"setBasicFilter": {"filter": {"range": {
                "sheetId": sid, "startRowIndex": hi, "endRowIndex": max(b, hi + 1),
                "startColumnIndex": 0, "endColumnIndex": len(tab.rows[hi])}}}})
        fmt_reqs.append({"autoResizeDimensions": {"dimensions": {
            "sheetId": sid, "dimension": "COLUMNS", "startIndex": 0, "endIndex": maxw}}})
    if "__tmp__" in existing:
        fmt_reqs.append({"deleteSheet": {"sheetId": existing["__tmp__"]}})
    try:
        sheets.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id, body={"requests": order_reqs + fmt_reqs}).execute()
    except Exception as e:
        print("  [format] non-fatal formatting error:", e)
        try:
            sheets.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id, body={"requests": order_reqs}).execute()
        except Exception:
            pass


def write_local_xlsx(path, tabs):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def fill(h):
        return PatternFill("solid", fgColor=h)
    thin = Side(style="thin", color="D9DEE8")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for name, tab in tabs.items():
        ws = wb.create_sheet(re.sub(r"[:\\/?*\[\]]", " ", name)[:31])
        maxw = max(tab.width(), 1)
        for row in tab.rows:
            ws.append(row if row else [""])

        def paint(ri, hexbg=None, fg=None, bold=False, size=None, span=None):
            span = span if span is not None else tab.fill_width(ri)
            for ci in range(1, span + 1):
                c = ws.cell(row=ri + 1, column=ci)
                c.font = Font(bold=bold, color=(fg or "000000"), size=(size or 11))
                if hexbg:
                    c.fill = fill(hexbg)
        for ri in tab.banners:
            paint(ri, HEX["TITLE"], HEX["TITLE_FG"], bold=True, size=13, span=maxw)
            if maxw > 1:
                ws.merge_cells(start_row=ri + 1, start_column=1, end_row=ri + 1, end_column=maxw)
            ws.cell(ri + 1, 1).alignment = Alignment(horizontal="left", vertical="center")
        for ri in tab.titles:
            paint(ri, HEX["SEC"], HEX["SEC_FG"], bold=True, size=12)
        for ri in tab.headers:
            paint(ri, HEX["HDR"], HEX["HDR_FG"], bold=True)
            for ci in range(1, len(tab.rows[ri]) + 1):
                ws.cell(ri + 1, ci).border = grid
                ws.cell(ri + 1, ci).alignment = Alignment(vertical="center", wrap_text=True)
        for hi in tab.headers:
            a, b = tab.data_span(hi); w = len(tab.rows[hi])
            for k, ri in enumerate(range(a, b)):
                rowhex = None if tab.no_autocolor else _row_hex(tab.rows[ri])
                band = rowhex or (HEX["ALT"] if k % 2 == 1 else None)
                bold = ri in tab.kpi
                is_band = any(str(v).strip() in BAND_HEX for v in tab.rows[ri])
                for ci in range(1, w + 1):
                    cell = ws.cell(row=ri + 1, column=ci)
                    cell.border = grid
                    if band:
                        cell.fill = fill(band)
                    if is_band:                      # dark, readable text on priority fills
                        cell.font = Font(bold=bold, color=BAND_TEXT_HEX)
                    elif bold:
                        cell.font = Font(bold=True)
                    if "\n" in str(cell.value or ""):
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
        for (ri, ci), fm in tab.cell_fmt.items():          # explicit per-cell overrides
            cell = ws.cell(row=ri + 1, column=ci + 1)
            cell.border = grid
            if fm.get("bg"):
                cell.fill = fill(rgb_hex(fm["bg"]))
            fg = rgb_hex(fm["fg"]) if fm.get("fg") else "000000"
            cell.font = Font(bold=fm.get("bold", False), color=fg)
        fh0 = tab.filter_headers[0] if tab.filter_headers else None
        ws.freeze_panes = f"A{(fh0 + 2) if (fh0 is not None and fh0 <= 5) else 2}"
        if tab.filter_headers:
            hi = tab.filter_headers[0]; a, b = tab.data_span(hi)
            ws.auto_filter.ref = f"A{hi + 1}:{get_column_letter(len(tab.rows[hi]))}{max(b, hi + 1)}"
        for ci in range(1, maxw + 1):
            longest = max((len(str(row[ci - 1])) for row in tab.rows if ci - 1 < len(row)), default=8)
            ws.column_dimensions[get_column_letter(ci)].width = min(max(longest + 2, 10), 52)
    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════════
#  REPORT ASSEMBLY
# ═══════════════════════════════════════════════════════════════════════════
def build_report(period_label, period_range, leads, model, bands, gen,
                 period_pending=None, total_pending=None, overdue_count=None,
                 overdue_leads=None, display_leads=None, pending_dataset=None,
                 done_count=None, remaining_count=None,
                 meetwalk_records=None, gm_metrics=None, wk_metrics=None):
    # display_leads = subset PLUS every Total Follow-Up Pending lead. Used for
    # the counsellor tabs (performance overview + pending). Falls back to `leads`.
    disp = display_leads if display_leads is not None else leads
    # pending_dataset = EXACTLY the Total Follow-Up Pending leads (same set that
    # produced the Summary count). Priority & Actions shows only these, so its
    # detail reconciles 1:1 with Summary -> Total Follow-Up Pending.
    pend = pending_dataset if pending_dataset is not None else disp
    tabs = OrderedDict()
    tabs["Summary"] = build_summary_tab(period_label, period_range, leads,
                                        model, bands, gen, period_pending,
                                        total_pending, overdue_count,
                                        done_count, remaining_count, pend,
                                        gm_metrics, wk_metrics)
    tabs["Priority & Actions"] = build_priority_tab(pend, gen, period_label, period_range)
    # detailed Google Meet & Walk-In tab (unique scheduled leads), right after
    # Priority & Actions and structured similarly.
    if meetwalk_records is not None:
        tabs["Google Meet & Walk-In"] = build_meetwalk_tab(
            meetwalk_records, gen, period_label, period_range)
    tabs["Counsellor Performance"] = build_counsellor_tab(
        pend, gen, period_label, period_range, meetwalk_records)
    # one tab per counsellor — built from the SAME pending dataset so each
    # counsellor shows only their share of the Total Follow-Up Pending leads.
    for name, tab in build_counsellor_detail_tabs(
            pend, gen, period_label, period_range).items():
        tabs[name] = tab
    tabs["Conversion Model"] = build_model_tab(model, bands, gen)
    return tabs


def score_and_classify(leads, model, bands):
    for l in leads:
        p = model.score(l["features"])
        l["conversion_chance"] = round(p * 100, 1)
        l["priority"] = band_for(p, bands)
    for l in leads:                              # next action needs priority set
        l["next_action"] = next_best_action(l)


def apply_period_followups(leads, today, cutoff, ov_start=None, ov_end=None,
                           cur_start=None, cur_end=None):
    """Re-derive each lead's follow-up metrics for a specific report period
    (cutoff = period end datetime, or None for the live/all-time view), then
    refresh the next-best-action which depends on the follow-up state.

    ov_start / ov_end define the Overdue window (previous completed period);
    cur_start / cur_end define the current-period pending window. Both are
    evaluated on the lead's effective Next Follow-Up Date across both sheets."""
    for l in leads:
        l["followup"] = follow_up_metrics(
            l["_next_raw"], l["_versions"], l["converted"], l["lost"],
            today, cutoff, ov_start, ov_end, l.get("_master_next", ""),
            cur_start, cur_end)
        l["next_action"] = next_best_action(l)


# ── Follow-Up completion within the report timeframe (RecordTimeStamp) ────────
TERMINAL_STATUSES = ("admission confirmed", "not interested", "irrelevant")


def followup_completion(lead, cur_start, cur_end):
    """For a lead in the Total Follow-Up Pending set, decide whether its
    follow-up was actually DONE within the applicable report timeframe.

    Returns (completion_status, followup_status):
      completion_status : 'Follow-Up Done' | 'Follow-Up Pending'
                          | 'Follow-Up Done (Invalid Info)'
      followup_status   : 'Overdue' | 'Current'   (from the finalized dataset)

    Logic (from the Active tab's RecordTimeStamp + Admission Status):
      * No Active update inside the report timeframe  -> Follow-Up Pending.
      * Update exists + terminal status               -> Follow-Up Done.
      * Update exists + must-reschedule status:
          - Next Follow-Up Date moved forward         -> Follow-Up Done.
          - not moved forward                         -> Follow-Up Done (Invalid Info).
    """
    fu = lead["followup"]
    followup_status = "Overdue" if fu.get("overdue") else "Current"

    ts = lead.get("_record_ts")
    updated = bool(ts and cur_start is not None
                   and cur_start <= ts.date() <= cur_end)
    # a lead actioned away this timeframe (Next Follow-Up Date moved out of the
    # window during the run) IS a completed follow-up even if its RecordTimeStamp
    # signal is unavailable (e.g. master-sourced date).
    if fu.get("actioned_away"):
        updated = True
    if not updated:
        return "Follow-Up Pending", followup_status

    admit = s(lead.get("admission_status")).lower()
    if any(k in admit for k in TERMINAL_STATUSES):
        return "Follow-Up Done", followup_status

    # must-reschedule status: Next Follow-Up Date must be moved forward past the
    # previous scheduled date (the one that put the lead on the follow-up list).
    cur_next = (parse_followup_date(lead.get("_next_raw"))
                or parse_followup_date(lead.get("_master_next")))
    prev_next, latest_arch = None, None
    for v in lead.get("_versions", []):
        a = parse_dt(v.get("archived_at"))
        if a and (latest_arch is None or a > latest_arch):
            latest_arch = a
            prev_next = parse_followup_date(v.get("next_followup"))
    if cur_next and prev_next and cur_next > prev_next:
        return "Follow-Up Done", followup_status
    if cur_next and prev_next is None and cur_next > cur_end:
        return "Follow-Up Done", followup_status   # no prior version, moved to future
    return "Follow-Up Done (Invalid Info)", followup_status


# ═══════════════════════════════════════════════════════════════════════════
#  EMAIL  (same configuration / sending logic as
#          pyConsolidatedLeadPerformanceReport.py — Gmail SMTP from
#          config_files/email_config.py)
# ═══════════════════════════════════════════════════════════════════════════
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _valid_recipients(recipients):
    """Clean + validate recipients before sending: split accidental joins
    (comma / semicolon / whitespace), trim, drop anything that is not a single
    valid e-mail address, and de-duplicate case-insensitively."""
    seen, out = set(), []
    for raw in recipients or []:
        for part in re.split(r"[,\s;]+", str(raw).strip()):
            addr = part.strip()
            if not addr:
                continue
            if not _EMAIL_RE.match(addr):
                print(f"  [email] WARNING skipping malformed recipient {addr!r} "
                      f"— check EMAIL_RECIPIENTS for a missing comma")
                continue
            key = addr.lower()
            if key not in seen:
                seen.add(key)
                out.append(addr)
    return out


def send_email(subject, html_body, recipients=None):
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    recipients = _valid_recipients(
        recipients if recipients is not None else EMAIL_RECIPIENTS)
    if not recipients:
        print("  [email] no valid recipients — nothing sent")
        return
    sys.path.insert(0, PROJECT_ROOT)
    try:
        import email_config as ec
        sender, app_pass = ec.GMAIL_SENDER, ec.GMAIL_APP_PASS
    except Exception as e:
        print("  [email] skipped - could not load config_files/email_config.py:", e)
        return
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = ", ".join(recipients)
    msg.attach(MIMEText(html_body, "html"))
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as srv:
            srv.login(sender, app_pass)
            srv.sendmail(sender, recipients, msg.as_string())
        print("  [email] sent to", ", ".join(recipients))
    except Exception as e:
        print("  [email] FAILED:", e)


def build_lfa_email_body(report_type, period_range, url, link_name, gen_stamp,
                         fu_pending, fu_done, fu_remaining,
                         gm_sched, gm_att, gm_showoff,
                         wk_sched, wk_att, wk_showoff):
    """Professional, self-contained HTML email in the SAME style as the
    consolidated report: header bar, KPI cards, a named call-to-action link, and
    the IntelliBI signature. Cards show the Follow-Up / Google Meet / Walk-In
    summary; every value is the exact figure used in the report's Summary tab.
    `url`/`link_name` may be empty — the report link is then omitted."""
    def _cards(items):
        return "".join(
            "<td style='padding:6px'>"
            "<div style='background:#f4f8fd;border:1px solid #e2e8f0;border-radius:8px;"
            "padding:14px 10px;text-align:center'>"
            f"<div style='font-size:24px;font-weight:700;color:#{color}'>{v}</div>"
            f"<div style='font-size:12px;color:#5b6b86;margin-top:2px'>{lbl}</div>"
            "</div></td>"
            for lbl, v, color in items)

    def _section(title, items):
        return (f"<div style='font-size:13px;font-weight:700;color:#1B355E;"
                f"margin:6px 0 4px'>{title}</div>"
                "<table role='presentation' width='100%' style='border-collapse:"
                "separate;margin:0 -6px 12px'><tr>" + _cards(items) + "</tr></table>")

    NAVY = "1B355E"
    # Percentages use the SAME basis the report's Summary tab shows:
    #   Follow-Up Done %      = Done / Total Follow-Up Pending  (Done+Remaining=Pending)
    #   Google Meet Attended% = Attended / (Attended + Show Off)  (past-due scheduled)
    #   Walk-In Attended %    = Attended / (Attended + Show Off)  (past-due scheduled)
    fu_done_pct = pct(fu_done, fu_pending)
    gm_att_pct = pct(gm_att, gm_att + gm_showoff)
    wk_att_pct = pct(wk_att, wk_att + wk_showoff)
    fu = _section("Follow-Up Summary", [
        ("Total Follow-Up Pending", fu_pending, NAVY),
        ("Total Follow-Up Done", fu_done, NAVY),
        ("Total Follow-Ups Remaining", fu_remaining, NAVY),
        ("Total Follow-Up Done %", fu_done_pct, NAVY)])
    gm = _section("Google Meet Summary", [
        ("Total Google Meet Scheduled", gm_sched, NAVY),
        ("Total Google Meet Attended", gm_att, NAVY),
        ("Total Google Meet Show Off", gm_showoff, NAVY),
        ("Total Google Meet Attended %", gm_att_pct, NAVY)])
    wk = _section("Walk-In Summary", [
        ("Total Walk-In Scheduled", wk_sched, NAVY),
        ("Total Walk-In Attended", wk_att, NAVY),
        ("Total Walk-In Show Off", wk_showoff, NAVY),
        ("Total Walk-In Attended %", wk_att_pct, NAVY)])

    if url:
        cta = (f"<p style='text-align:center;margin:6px 0 24px'>"
               f"<a href='{url}' style='background:#2B547E;color:#ffffff;"
               f"text-decoration:none;padding:13px 30px;border-radius:6px;"
               f"font-weight:600;font-size:15px;display:inline-block'>"
               f"Open {link_name} &nbsp;&rsaquo;</a></p>"
               f"<p style='margin:0;color:#5b6b86;font-size:13px'>Or open it here: "
               f"<a href='{url}' style='color:#2B547E;font-weight:600;"
               f"text-decoration:none'>{link_name}</a></p>")
    else:
        cta = ("<p style='margin:0;color:#5b6b86;font-size:13px'>The detailed "
               "report is shared separately with authorised recipients.</p>")

    return f"""<html><body style="margin:0;padding:24px;background:#eef2f8;
  font-family:'Segoe UI',Roboto,Arial,sans-serif;color:#1a2a48">
  <div style="max-width:640px;margin:0 auto;background:#ffffff;border-radius:10px;
    overflow:hidden;border:1px solid #e2e8f0">
    <div style="background:#1B355E;padding:22px 28px;color:#ffffff">
      <div style="font-size:20px;font-weight:700">IntelliBI &nbsp;&middot;&nbsp; {report_type} Lead Follow-Up Analysis Report</div>
      <div style="font-size:13px;opacity:.85;margin-top:4px">Reporting Period: {period_range}</div>
    </div>
    <div style="padding:24px 28px">
      <p style="margin:0 0 14px">Hello Team,</p>
      <p style="margin:0 0 18px;line-height:1.5">
        Please find the <b>{report_type}</b> lead follow-up analysis summary for
        <b>{period_range}</b>. Here is the follow-up, Google Meet and Walk-In snapshot:</p>
      {fu}
      {gm}
      {wk}
      {cta}
      <p style="margin:26px 0 0;line-height:1.5">
        Thanks &amp; Regards,<br><b>IntelliBI Automation Team</b></p>
    </div>
    <div style="background:#f4f8fd;padding:12px 28px;font-size:12px;color:#8494ad;
      border-top:1px solid #e8edf5">
      Automated report &middot; Generated {gen_stamp}</div>
  </div>
</body></html>"""


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════
def run():
    today = now_ist().date()
    gen = now_ist().strftime("%d-%b-%Y %I:%M %p") + " IST"

    sheets = drive = None
    if not LOCAL_DIR:
        sheets = get_read_service()
    if not DRY_RUN and not LOCAL_DIR:
        drive = get_drive_service()

    # ── read every source ────────────────────────────────────────────────────
    active   = read_source(sheets, RESULT_SHEET_ID, RESULT_ACTIVE_TABS, "result_active")
    inactive = read_source(sheets, RESULT_SHEET_ID, RESULT_INACTIVE_TABS, "result_inactive")
    master   = read_source(sheets, MASTER_SHEET_ID, MASTER_TABS, "master")
    meet     = read_source(sheets, MEET_SHEET_ID, MEET_TABS, "meet", optional=True)
    walkin   = read_source(sheets, WALKIN_SHEET_ID, WALKIN_TABS, "walkin", optional=True)
    # Walk-In attendance = the "Walk-In New" tab of the same Student Inquiry
    # Tracker spreadsheet (Mobile Number + Timestamp).
    walkin_new = read_source(sheets, WALKIN_SHEET_ID, WALKIN_ATTEND_TABS,
                             "walkin_new", optional=True)

    print(f"Sources: active={len(active or [])} inactive={len(inactive or [])} "
          f"master={len(master or [])} meet={len(meet or [])} walkin={len(walkin or [])} "
          f"walkin_new={len(walkin_new or [])}")

    master_idx = index_master(master)
    meet_map = load_meet_attendance(meet)
    walkin_set = load_walkin_attended(walkin)
    # Google Meet & Walk-In attendance lookups for the new reporting section/tab.
    meet_attended = meet_attended_set(meet_map)
    walk_ts = load_walkin_timestamps(walkin_new)

    # ── train the conversion model on ALL available history ──────────────────
    # Live master first (authoritative), then augment with past exports in the
    # intellibi_lead_cycle folder for any leads not already present.
    train_idx = dict(master_idx)
    added = 0
    for mob, row in load_history_master_index().items():
        if mob not in train_idx:
            train_idx[mob] = row
            added += 1
    if added:
        print(f"Training augmented with {added} historical leads from {HISTORY_DIR}")
    model = ConversionModel().train(build_training_samples(train_idx, meet_map, walkin_set))
    bands = priority_bands(model.base_rate)
    print(f"Model: base rate {model.base_rate*100:.2f}%  "
          f"({model.n_pos}/{model.n_train} converted)")

    # ── assemble & score active leads ────────────────────────────────────────
    leads = assemble_leads(active, inactive, master_idx, meet_map, walkin_set, today)
    score_and_classify(leads, model, bands)
    print(f"Active leads scored: {len(leads)}")

    # ── canonicalise counsellor names (merge case/spacing variants → one) ─────
    # Build one alias/case map from every counsellor spelling seen (master +
    # active leads), then rewrite both stores in place so ALL downstream
    # grouping — scorecard, per-counsellor detail tabs, ranking, meet/walk-in,
    # reconciliation — sees a single canonical name per counsellor.
    _couns_names = [m.get("counsellor", "") for m in master_idx.values()]
    _couns_names += [l.get("counsellor", "") for l in leads]
    _couns_canon = build_counsellor_canon_map(_couns_names)
    for _m in master_idx.values():
        _c = _m.get("counsellor", "")
        if str(_c).strip():
            _m["counsellor"] = canon_couns_display(_c, _couns_canon)
    apply_counsellor_canon(leads, _couns_canon)

    # Already-enrolled students (by normalized phone) — excluded from every
    # Follow-Up Pending calculation/detail below. Loaded once (same for all
    # periods). Empty set = no exclusion (e.g. sheet unreadable).
    enrolled_phones = load_enrolled_phones(sheets)
    if enrolled_phones:
        print(f"Enrolled-student exclusion active: {len(enrolled_phones)} phone(s).")

    # ── periods ──────────────────────────────────────────────────────────────
    jobs = []

    # ── AUTO mode: decide reports from the execution date (ignores the manual
    #    Daily/Weekly/Monthly/Manual flags for this run) ─────────────────────
    if GENERATE_AUTO:
        # Daily — every run, for the current day.
        st, en = day_bounds(today)
        jobs.append(("Daily", today.strftime("%d-%b-%Y"), st, en,
                     f"Lead Follow-Up Analysis - Daily - {today.strftime('%d-%b-%Y')}"))
        # Weekly — only on Monday, for the previous COMPLETE Mon–Sun week.
        if today.weekday() == 0:
            st, en, mon, sun = week_bounds(today - timedelta(days=7))
            jobs.append(("Weekly",
                         f"{mon.strftime('%d-%b-%Y')} to {sun.strftime('%d-%b-%Y')}",
                         st, en,
                         f"Lead Follow-Up Analysis - Weekly - {mon.strftime('%d-%b-%Y')} to {sun.strftime('%d-%b-%Y')}"))
        # Monthly — only on the last calendar day of the month, for that whole month.
        if today.day == calendar.monthrange(today.year, today.month)[1]:
            st, en, ms, me = month_bounds(today.year, today.month)
            jobs.append(("Monthly",
                         f"{ms.strftime('%d-%b-%Y')} to {me.strftime('%d-%b-%Y')}",
                         st, en, f"Lead Follow-Up Analysis - Monthly - {ms.strftime('%b-%Y')}"))

    # ── MANUAL mode: existing behaviour, unchanged (only runs when AUTO is off) ─
    if GENERATE_DAILY and not GENERATE_AUTO:
        d = datetime.strptime(DAILY_DATE, "%Y-%m-%d").date() if DAILY_DATE else today
        st, en = day_bounds(d)
        jobs.append(("Daily", d.strftime("%d-%b-%Y"), st, en,
                     f"Lead Follow-Up Analysis - Daily - {d.strftime('%d-%b-%Y')}"))
    if GENERATE_WEEKLY and not GENERATE_AUTO:
        ref = datetime.strptime(WEEKLY_REFERENCE_DATE, "%Y-%m-%d").date() if WEEKLY_REFERENCE_DATE else today
        st, en, mon, sun = week_bounds(ref)
        jobs.append(("Weekly", f"{mon.strftime('%d-%b-%Y')} to {sun.strftime('%d-%b-%Y')}",
                     st, en,
                     f"Lead Follow-Up Analysis - Weekly - {mon.strftime('%d-%b-%Y')} to {sun.strftime('%d-%b-%Y')}"))
    if GENERATE_MONTHLY and not GENERATE_AUTO:
        yr = MONTHLY_YEAR or today.year; mo = MONTHLY_MONTH or today.month
        st, en, ms, me = month_bounds(yr, mo)
        jobs.append(("Monthly", f"{ms.strftime('%d-%b-%Y')} to {me.strftime('%d-%b-%Y')}",
                     st, en, f"Lead Follow-Up Analysis - Monthly - {ms.strftime('%b-%Y')}"))
    if GENERATE_MANUAL and not GENERATE_AUTO:
        if not (MANUAL_START_DATE and MANUAL_END_DATE):
            print("  [manual] GENERATE_MANUAL is on but MANUAL_START_DATE / "
                  "MANUAL_END_DATE are not set — skipping the Manual report.")
        else:
            ms = datetime.strptime(MANUAL_START_DATE, "%Y-%m-%d").date()
            me = datetime.strptime(MANUAL_END_DATE, "%Y-%m-%d").date()
            st = datetime.combine(ms, time.min); en = datetime.combine(me, time.max)
            jobs.append(("Manual",
                         f"{ms.strftime('%d-%b-%Y')} to {me.strftime('%d-%b-%Y')}",
                         st, en,
                         f"Lead Follow-Up Analysis - Manual - {ms.strftime('%d-%b-%Y')} "
                         f"to {me.strftime('%d-%b-%Y')}"))

    folder_cache = {}
    for label, rng, st, en, fname in jobs:
        # re-derive follow-ups for THIS period's cutoff + windows.
        cutoff = en if en is not None else datetime.combine(today, time.max)
        ov_start, ov_end = overdue_window(label, st, en, today)
        # current-period pending window (Monthly = month-to-date, no future).
        cur_start = st.date() if st else None
        cur_end = en.date() if en else None
        if label == "Monthly" and cur_end:
            cur_end = min(cur_end, today)
        # apply over ALL active leads so pending/overdue (which look at follow-up
        # date windows, independent of the activity filter) are complete.
        apply_period_followups(leads, today, cutoff, ov_start, ov_end,
                               cur_start, cur_end)

        subset = [l for l in leads if in_period(l, st, en)]

        # ── ONE common, de-duplicated Total Follow-Up Pending dataset ─────────
        # = active/relevant pending leads + master-only pending leads, so every
        # counted lead carries full detail. Same set drives the count and the
        # Priority & Actions / Counsellor Performance / Counsellor tabs.
        active_mobiles = {l["mobile"] for l in leads if l["mobile"]}
        master_pending = build_master_pending_leads(
            master_idx, active_mobiles, meet_map, walkin_set, model, bands,
            today, cutoff, cur_start, cur_end, ov_start, ov_end)
        active_pending = [l for l in leads if l["followup"].get("total_pending")]

        def _key(l):
            return l["mobile"] or ("email:" + norm_email(l.get("email", "")))

        pending_seen, pending_dataset = set(), []
        for l in active_pending + master_pending:
            # Exclude already-enrolled students at the BASE pending level, so every
            # dependent count/detail (Summary, Priority & Actions, Counsellor
            # Performance, per-counsellor tabs, ranking) stays consistent.
            if l.get("mobile") and l["mobile"] in enrolled_phones:
                continue
            k = _key(l)
            if not k or k in pending_seen:
                continue
            pending_seen.add(k); pending_dataset.append(l)

        total_pending = len(pending_dataset)
        period_pending = sum(1 for l in pending_dataset
                             if l["followup"].get("period_pending"))
        overdue_leads = [l for l in pending_dataset if l["followup"].get("overdue")]
        overdue_count = len(overdue_leads)

        # Assign ONE Priority & Actions rank (by Conversion Chance %, highest
        # first) on the shared pending dataset so Priority & Actions and each
        # Counsellor Individual tab show the SAME rank for the same lead.
        for i, l in enumerate(sorted(pending_dataset,
                                     key=lambda x: -x["conversion_chance"]), 1):
            l["_rank"] = i

        # Follow-up completion within the report timeframe (RecordTimeStamp).
        # Done + Remaining = Total Follow-Up Pending, by construction.
        done_count = 0
        for l in pending_dataset:
            comp, fstat = followup_completion(l, cur_start, cur_end)
            l["_completion_status"] = comp
            l["_followup_status"] = fstat
            if comp != "Follow-Up Pending":
                done_count += 1
        remaining_count = total_pending - done_count

        # display set for the counsellor-performance tab = period subset + every
        # pending lead (so its Follow-Up Pending column covers all pending leads).
        disp_seen, display_leads = set(), []
        for l in subset + pending_dataset:
            # keep the Counsellor Performance follow-up counts consistent with the
            # excluded pending dataset (drop enrolled students here too).
            if l.get("mobile") and l["mobile"] in enrolled_phones:
                continue
            k = _key(l)
            if k in disp_seen:
                continue
            disp_seen.add(k); display_leads.append(l)

        # ── validation: count must reconcile with detail (requirements #6/#7) ──
        if total_pending != period_pending + overdue_count:
            print(f"  [validate] WARNING: total-pending {total_pending} != "
                  f"period {period_pending} + overdue {overdue_count}")
        # Priority & Actions shows exactly `pending_dataset`; its unique mobiles
        # must equal the Summary Total Follow-Up Pending count.
        pa_unique = len({_key(l) for l in pending_dataset})
        if pa_unique != total_pending:
            print(f"  [validate] WARNING: Priority&Actions unique {pa_unique} "
                  f"!= Summary total-pending {total_pending}")
        def _couns_key(l):
            return re.sub(r"\s+", " ", l["counsellor"]).strip() or "(Unassigned)"

        # Level 3: counsellor distribution of the SAME dataset sums to the total.
        by_couns = defaultdict(int)
        for l in pending_dataset:
            by_couns[_couns_key(l)] += 1
        if sum(by_couns.values()) != total_pending:
            print(f"  [validate] WARNING: counsellor pending sum "
                  f"{sum(by_couns.values())} != total {total_pending}")
        # Per-counsellor: Counsellor Performance (over display set) must match
        # Counsellor Individual (over pending dataset) for every counsellor.
        perf_pend = defaultdict(int)
        for l in display_leads:
            if l["followup"].get("total_pending"):
                perf_pend[_couns_key(l)] += 1
        if dict(perf_pend) != dict(by_couns):
            print(f"  [validate] WARNING: per-counsellor mismatch "
                  f"perf={dict(perf_pend)} indiv={dict(by_couns)}")
        # Done + Remaining must equal Total Follow-Up Pending.
        if done_count + remaining_count != total_pending:
            print(f"  [validate] WARNING: done {done_count} + remaining "
                  f"{remaining_count} != total-pending {total_pending}")

        # ── Google Meet & Walk-In (this period) ───────────────────────────────
        # Scheduled meets/walk-ins (from master) in the period window, deduped
        # per unique lead, with attendance from the Meet form / Walk-In New tab.
        lead_by_mob = {l["mobile"]: l for l in leads if l["mobile"]}
        meetwalk_records = build_meet_walk_records(
            master_idx, lead_by_mob, meet_attended, walk_ts, model, bands,
            today, cur_start, cur_end, ov_start, ov_end)
        gm_metrics = meet_walk_metrics(meetwalk_records, "gm")
        wk_metrics = meet_walk_metrics(meetwalk_records, "wk")
        # reconciliation: Total = Overdue + Current; detailed-tab unique leads.
        for lbl_m, mtr, k in (("GMeet", gm_metrics, "gm"), ("Walk-In", wk_metrics, "wk")):
            if mtr["total"] != mtr["overdue"] + mtr["current"]:
                print(f"  [validate] WARNING: {lbl_m} total {mtr['total']} != "
                      f"overdue {mtr['overdue']} + current {mtr['current']}")
            uniq = sum(1 for r in meetwalk_records if r.get(k))
            if uniq != mtr["total"]:
                print(f"  [validate] WARNING: {lbl_m} tab-unique {uniq} != "
                      f"total {mtr['total']}")

        tabs = build_report(label, rng, subset, model, bands, gen,
                            period_pending, total_pending, overdue_count,
                            overdue_leads, display_leads, pending_dataset,
                            done_count, remaining_count,
                            meetwalk_records, gm_metrics, wk_metrics)
        print(f"\n{label} | {rng} | subset: {len(subset)} | "
              f"period-pending: {period_pending} | overdue: {overdue_count} | "
              f"total-pending: {total_pending} (master-only "
              f"{len(master_pending)}) [{ov_start}..{ov_end}] | "
              f"GMeet {gm_metrics['total']} Walk {wk_metrics['total']} | "
              f"tabs: {len(tabs)}")

        xlsx_path = os.path.join(OUTPUT_DIR, fname + ".xlsx")
        try:
            write_local_xlsx(xlsx_path, tabs)
            print(f"  local xlsx: {xlsx_path}")
        except Exception as e:
            print("  local xlsx FAILED:", e)

        if DRY_RUN or LOCAL_DIR:
            continue
        sub_name = OUTPUT_SUBFOLDERS.get(label)
        if sub_name not in folder_cache:
            folder_cache[sub_name] = resolve_output_folder(
                drive, OUTPUT_PARENT_FOLDER_ID, sub_name)
        target_folder = folder_cache[sub_name]
        url, created, _fid = upload_report_to_drive(drive, target_folder, fname, xlsx_path)
        print(f"  -> {sub_name}  {'created' if created else 'replaced'}: {url}")

        # ── Email the summary (same config/format as the consolidated report) ──
        # Values below are the EXACT figures shown on this report's Summary tab.
        if SEND_EMAIL:
            subject = f"{label} Lead Follow-Up Analysis Report - {rng}"
            link_name = f"{label} Lead Follow-Up Analysis Report"
            masked_recips = [r for r in EMAIL_RECIPIENTS
                             if s(r).lower() in {m.lower() for m in MASK_RECIPIENTS}]
            normal_recips = [r for r in EMAIL_RECIPIENTS if r not in masked_recips]
            _fu = (total_pending, done_count, remaining_count)
            _gm = (gm_metrics["total"], gm_metrics["attended"], gm_metrics["showoff"])
            _wk = (wk_metrics["total"], wk_metrics["attended"], wk_metrics["showoff"])

            # Restricted recipient(s): build a MASKED copy of THIS report (Mobile
            # Number + Email Address masked, same rules as the consolidated
            # report), upload it as a SEPARATE Drive file, and share it with them
            # as Editor. The original full report is never modified.
            url_masked = None
            if masked_recips:
                masked_xlsx_path = os.path.join(OUTPUT_DIR, fname + " (Masked).xlsx")
                try:
                    write_local_xlsx(masked_xlsx_path, mask_tabs(tabs))
                    if os.path.exists(masked_xlsx_path):
                        url_masked, _mc, masked_fid = upload_report_to_drive(
                            drive, target_folder, fname + " (Masked)", masked_xlsx_path)
                        print(f"  masked copy: {url_masked}")
                        share_file_with(drive, masked_fid, masked_recips, role="writer")
                except Exception as e:
                    print("  [drive] masked report build/upload/share FAILED:", e)

            # Authorised recipients: the full report link (unchanged).
            if normal_recips:
                send_email(subject,
                           build_lfa_email_body(label, rng, url, link_name, gen,
                                                *_fu, *_gm, *_wk),
                           normal_recips)
            # Restricted recipient(s): the MASKED report link (Editor access
            # granted above) — NEVER the full report. If the masked copy could not
            # be produced, fall back to the summary WITHOUT any link so no
            # unmasked lead detail ever leaves the building.
            if masked_recips:
                send_email(subject,
                           build_lfa_email_body(label, rng, url_masked or "", link_name,
                                                gen, *_fu, *_gm, *_wk),
                           masked_recips)

    print("\nDone.")
    return 0


if __name__ == "__main__":
    sys.exit(run())