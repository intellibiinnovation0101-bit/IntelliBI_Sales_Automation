#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IntelliBI - Consolidated Lead Performance Report
================================================

Generates Daily / Weekly / Monthly Lead Performance reports from the
**Consolidated Master Lead** sheet (the single source of truth produced by
consolidate_leads.py). This script only READS the master sheet; it never
touches the consolidation logic or the individual source sheets.

Each report is written as a Google Sheet (one file per period) into a dedicated
Drive folder, with these tabs:
    1. Summary          - executive dashboard (exec summary, source & counsellor
                          performance, course interest, follow-ups)
    2. Lead Details     - lead-level detail for every lead active in the period
    3. CB - <Counsellor> - one tab per counsellor: their summary + their leads

Daily  -> a fresh file per date.
Weekly -> one file per Mon-Sun week (updated in place if it already exists).
Monthly-> one file per calendar month (updated in place if it already exists).

Optionally emails the report links to management.

Auth / conventions match the rest of the repo: service account
config_files/service_account.json, Sheets v4 + Drive v3, Gmail SMTP from
config_files/email_config.py.

Run:  python report/pyConsolidatedLeadPerformanceReport.py
Offline test (no Google I/O):
      REPORT_LOCAL_MASTER_CSV=/path/Consolidated_Master_Lead.csv REPORT_DRY_RUN=1 \
      python report/pyConsolidatedLeadPerformanceReport.py
"""

import os
import re
import sys
import calendar
from datetime import datetime, date, timedelta, time, timezone
from collections import OrderedDict, defaultdict

import pandas as pd

# The team reads IST; compute "Generated" in IST from UTC so the timestamp is
# correct regardless of the machine's / scheduler's local timezone.
IST = timezone(timedelta(hours=5, minutes=30))


def now_ist():
    return datetime.now(timezone.utc).astimezone(IST).replace(tzinfo=None)

# ============================================================
# REPORT GENERATION CONFIGURATION
# ============================================================
GENERATE_DAILY_REPORT   = True
GENERATE_WEEKLY_REPORT   = False
GENERATE_MONTHLY_REPORT  = False

# Optional manual report periods (leave as None to use the defaults below).
DAILY_REPORT_DATE            = None      # e.g. "2026-07-30"
WEEKLY_REPORT_REFERENCE_DATE = None      # e.g. "2026-07-30" (any day in the wanted week)
MONTHLY_REPORT_MONTH         = None      # e.g. 7   (1-12)
MONTHLY_REPORT_YEAR          = None      # e.g. 2026 (defaults to current year)

# Email the report links to management?  True / False
SEND_EMAIL       = True
EMAIL_RECIPIENTS = [
    "harishintellibi@gmail.com",
    "salesintellibi01@gmail.com",
    "info@intellibiinnovationstechnologies.in",
    "salesintellibi03@gmail.com",
    "163manish.sharma@gmail.com"
]

# Recipients who must receive a MASKED copy of the report — every lead's Mobile
# Number and Email ID are masked in THEIR copy only. Everyone else gets the full,
# unmasked report exactly as before. Compared case-insensitively. Masking is done
# on a separate copy of the data; the source/original report is never modified.
MASK_RECIPIENTS = {"163manish.sharma@gmail.com"}

# ---- Source ----
MASTER_SHEET_ID = "1zZQjXnMJD96Ca0MNyfSt4-XS0z5w3rT7WPdb9qsP1Gs"   # Consolidated Master Lead
MASTER_TAB_NAME = None            # None -> first tab

# Already-enrolled students — "IntelliBI — Student Admission Responses". A lead
# whose phone/mobile appears here is an enrolled student following up post-
# admission (onboarding / payment / batch / documents / course process …) and
# must NOT be counted as a Repeat lead or in ANY dependent metric. Their rows are
# filtered out of this in-memory report only; the source sheet is never modified.
ENROLLED_SHEET_ID = "1oaXxg3JdtxFp8lFWijIMZKaMZvS0SiglI1K2JTrN2fs"
ENROLLED_TAB_NAME = None           # None -> first tab

# ---- Output Drive folders ----
DAILY_FOLDER_ID   = "1kuGgoyseH49tiEnwmKBgz8xceF5u7uJP"
WEEKLY_FOLDER_ID  = "1iUzEaoOS2ViCC7qH4W8Kj-DcXh3RQM_I"
MONTHLY_FOLDER_ID = "1DICOV0iW5W2oIs7tKvsFfVUKlsfz4TxW"

# ---- Auth ----
# --- IntelliBI Sales Automation portability bootstrap (auto-inserted) ---
import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "common"))
import _bootstrap  # noqa: E402  (sys.path + env defaults + config.yaml)
from paths import CREDENTIALS_DIR, CONFIG_DIR, LOGS_DIR  # noqa: E402
# --- end bootstrap ---

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SERVICE_ACCOUNT_FILE = os.environ.get(
    "GOOGLE_SERVICE_ACCOUNT_FILE",
    os.path.join(CREDENTIALS_DIR, "service_account.json"))
# Scopes are split so the IMPERSONATED (domain-wide-delegation) token requests
# ONLY the drive scope. DWD requires every requested scope to be authorised for
# the client id; your grant authorises `drive` (what pyAttendaceFeedbackReport.py
# uses), so requesting spreadsheets+drive together was being rejected. Reading the
# master uses the service account's OWN identity (no DWD), so the spreadsheets
# scope there needs no grant.
SCOPES = ["https://www.googleapis.com/auth/spreadsheets",
          "https://www.googleapis.com/auth/drive"]        # kept for reference
READ_SCOPES  = ["https://www.googleapis.com/auth/spreadsheets"]   # SA-direct read
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive"]          # impersonated upload

# A service account has NO Drive storage of its own, so it cannot CREATE files in
# a My Drive folder (Google returns "storageQuotaExceeded"). To create the report
# files in your folders, the service account impersonates a Workspace user (who
# owns the folders / has storage). This requires a ONE-TIME domain-wide-delegation
# authorisation for the service account's client ID (see README). Set to None only
# if you instead use a Shared Drive (which has no per-user quota).
IMPERSONATE_USER = os.environ.get(
    "REPORT_IMPERSONATE_USER", "info@intellibiinnovationstechnologies.in")

# ---- Testing / offline switches (env-driven; leave unset in production) ----
LOCAL_MASTER_CSV = os.environ.get("REPORT_LOCAL_MASTER_CSV")   # read master from CSV instead of Sheets
DRY_RUN          = os.environ.get("REPORT_DRY_RUN") == "1"     # build + write local xlsx, no Google I/O
OUTPUT_DIR       = os.environ.get("REPORT_OUTPUT_DIR", os.path.join(os.path.dirname(os.path.abspath(__file__)), "output"))

MAX_COUNSELLOR_TABS = 40

# ---- Master column names ----
C_FIRST   = "First Enquiry Date"
C_LATEST  = "Latest Enquiry Date"
C_INIT    = "LeadInitalTimestamp"
C_NAME    = "Full Name"
C_MOBILE  = "Mobile Number"
C_EMAIL   = "Email Address"
C_COURSE  = "Which technology are you interested in learning?"
C_COURSE_ADV = "Course Advised"
C_STATUS  = "Lead Status"
C_ADM     = "Admission Status"
C_BACKOUT = "Backout Reason"
C_COUNSEL = "Counselling By"
C_GMEET   = "IsGoogleMeetSchedule"
C_WALKSCH = "IsWalkInSchedule"
C_VALID   = "IsPhoneNumberValid"
C_RELEV   = "IsLeadRelevant"
C_REF     = "IsReferral"
C_REFNAME = "Referrer's Name"
C_HIST    = "Lead Interaction History"
C_PLAT    = "Platforms Used"
C_NINT    = "Number of Interactions"

SOURCES = [("Walk-In", "IsWalk-In"), ("Website", "IsWebsite"),
           ("WhatsApp", "IsWhatsapp"), ("Call", "IsCall")]
SOURCE_LABELS = [s for s, _ in SOURCES]

DATE_FMTS = ("%d-%b-%Y %I:%M %p", "%d-%b-%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d")

try:
    from dateutil import parser as _dtparser
except Exception:
    _dtparser = None


# ============================================================
# Small helpers
# ============================================================
def s(v):
    return "" if v is None else str(v).strip()


def yes(v):
    return s(v).lower() == "yes"


# IntelliBI is the internal portal that re-processes Call / WhatsApp / Website leads.
# It must be EXCLUDED from every calculation (it would double-count the underlying
# lead) but is still shown, greyed + struck through, for reference. Matching is
# trimmed + case-insensitive.
INTELLIBI_SRC = "intellibi"


def is_intellibi_src(label):
    return s(label).casefold() == INTELLIBI_SRC


def pct(n, d):
    return "0.0%" if not d else f"{100.0 * n / d:.1f}%"


def norm_phone(value):
    """Normalize a phone/mobile for matching: strip every non-digit (same digit
    extraction the masking uses) and keep the last 10 digits, so
    '+91 9876543210', '98765 43210' and '9876543210' all reduce to the same key.
    Returns '' when there are no digits (blanks never match anything)."""
    digits = re.sub(r"\D", "", s(value))
    return digits[-10:] if len(digits) >= 10 else digits


def mask_mobile(value):
    """Mask a mobile number: keep the first 2 and last 2 digits, replace the
    digits in between with '*'.  e.g. 9876543210 -> 98******10.  Values that are
    blank or have <=4 digits are returned unchanged (nothing meaningful to hide)."""
    v = s(value)
    if not v:
        return v
    digits = re.sub(r"\D", "", v)
    if len(digits) <= 4:
        return v
    return digits[:2] + ("*" * (len(digits) - 4)) + digits[-2:]


def mask_email(value):
    """Mask an email's local part: keep the first 2 and last 2 characters, star
    the rest; the domain stays visible.  e.g. rahulkumar@gmail.com ->
    ra******ar@gmail.com,  abcdef@company.com -> ab**ef@company.com."""
    v = s(value)
    if not v or "@" not in v:
        return v
    local, domain = v.split("@", 1)
    n = len(local)
    if n <= 2:
        masked = local                                   # too short to mask
    elif n <= 4:
        masked = local[0] + ("*" * (n - 2)) + local[-1]  # keep 1 + 1
    else:
        masked = local[:2] + ("*" * (n - 4)) + local[-2:]
    return masked + "@" + domain


def mask_dataframe(df):
    """Return a COPY of the master DataFrame with every lead's Mobile Number and
    Email Address masked. The ORIGINAL df is never modified — other (authorised)
    recipients still receive the complete data."""
    d = df.copy()
    if C_MOBILE in d.columns:
        d[C_MOBILE] = d[C_MOBILE].map(mask_mobile)
    if C_EMAIL in d.columns:
        d[C_EMAIL] = d[C_EMAIL].map(mask_email)
    return d


def parse_dt(value):
    v = s(value)
    if not v:
        return None
    for f in DATE_FMTS:
        try:
            return datetime.strptime(v, f)
        except ValueError:
            pass
    if _dtparser:
        try:
            return _dtparser.parse(v, dayfirst=False, fuzzy=True)
        except Exception:
            return None
    return None


def parse_history(text):
    """Parse a Lead Interaction History cell into a list of
    (source, datetime, counsellor). Returns [] if unparseable."""
    t = s(text)
    if not t:
        return []
    srcs, dates, cbys = [], [], []
    for line in t.splitlines():
        low = line.lower()
        if low.startswith("lead source:"):
            srcs = [x.strip() for x in line.split(":", 1)[1].split(",")]
        elif low.startswith("lead date:"):
            dates = [x.strip() for x in line.split(":", 1)[1].split(",")]
        elif low.startswith("counselling by:"):
            cbys = [x.strip() for x in line.split(":", 1)[1].split(",")]
    out = []
    for i, dstr in enumerate(dates):
        dt = parse_dt(dstr)
        if not dt:
            continue
        src = srcs[i] if i < len(srcs) else ""
        cby = cbys[i] if i < len(cbys) else ""
        out.append((src, dt, cby))
    return out


def is_referral(row):
    return yes(row.get(C_REF)) or bool(s(row.get(C_REFNAME)))


def is_admission_confirmed(row):
    return "confirm" in s(row.get(C_ADM)).lower()


def is_lost(row):
    adm = s(row.get(C_ADM)).lower()
    st = s(row.get(C_STATUS)).lower()
    if adm in ("not interested", "lost", "backed out", "backout", "dropped"):
        return True
    if st == "not interested":
        return True
    if s(row.get(C_BACKOUT)):
        return True
    return False


# ============================================================
# Period definitions
# ============================================================
def day_bounds(d):
    return (datetime.combine(d, time.min), datetime.combine(d, time.max))


def week_bounds(ref):
    monday = ref - timedelta(days=ref.weekday())
    sunday = monday + timedelta(days=6)
    return (datetime.combine(monday, time.min), datetime.combine(sunday, time.max), monday, sunday)


def month_bounds(year, month):
    last = calendar.monthrange(year, month)[1]
    start = date(year, month, 1)
    end = date(year, month, last)
    return (datetime.combine(start, time.min), datetime.combine(end, time.max), start, end)


# ============================================================
# Tab model
# ============================================================
class Tab:
    def __init__(self, name):
        self.name = name
        self.rows = []
        self.banners = set()     # full-width report title banner
        self.info = set()        # header info rows (Reporting Period / Generated)
        self.titles = set()      # section titles (span their table's width)
        self.headers = set()     # table header rows
        self.filter_headers = [] # header row indices that get a filter
        self.kpi = set()         # rows whose text is bold (headline KPIs)
        self.chart = None        # optional chart spec (dict) drawn on this tab

    def banner(self, text):
        self.banners.add(len(self.rows)); self.rows.append([text])

    def info_row(self, label, value):
        self.info.add(len(self.rows)); self.rows.append([label, value])

    def title(self, text):
        self.titles.add(len(self.rows)); self.rows.append([text])

    def header(self, cols, filterable=False):
        i = len(self.rows)
        self.headers.add(i)
        if filterable:
            self.filter_headers.append(i)
        self.rows.append([str(c) for c in cols])

    def row(self, cols, kpi=False):
        if kpi:
            self.kpi.add(len(self.rows))
        self.rows.append(["" if c is None else c for c in cols])

    def blank(self):
        self.rows.append([])

    def width(self):
        return max((len(r) for r in self.rows), default=1)

    def _stops(self):
        return self.banners | self.info | self.titles | self.headers

    def data_span(self, header_index):
        """(start, end_exclusive) row range of contiguous data rows following a
        header, stopping at the next blank/title/header/banner/info row."""
        stops = self._stops()
        start = header_index + 1
        end = start
        while end < len(self.rows):
            if end in stops:
                break
            if not any(str(c).strip() for c in self.rows[end]):  # blank row
                break
            end += 1
        return start, end

    def fill_width(self, ri):
        """How many columns a row's fill/colour should span: a table's header and
        data rows span the table's own column count (so e.g. the 3-column
        Executive Summary is never coloured across the 12-column sheet); the
        report banner spans the full width; info rows span their label+value."""
        if ri in self.headers:
            return len(self.rows[ri])
        for hi in self.headers:
            a, b = self.data_span(hi)
            if a <= ri < b:
                return len(self.rows[hi])
        if ri in self.banners:
            return max(self.width(), 1)
        if ri in self.titles:
            nh = sorted(h for h in self.headers if h > ri)
            return len(self.rows[nh[0]]) if nh else max(self.width(), 1)
        # info / other: span its own non-empty cells
        ln = len([c for c in self.rows[ri] if str(c).strip() != ""])
        return max(ln, 1)


# ============================================================
# Metric computation
# ============================================================
def prepare_active(df, start, end):
    """Return list of active-lead dicts (leads with >=1 interaction in [start,end]),
    each augmented with _ninper, _srcs_inper, _is_new."""
    active = []
    for _, r in df.iterrows():
        row = {k: s(v) for k, v in r.items()}
        inter = parse_history(row.get(C_HIST))
        if not inter:
            fdt = parse_dt(row.get(C_FIRST) or row.get(C_INIT))
            if fdt:
                first_src = next((lbl for lbl, col in SOURCES if yes(row.get(col))), "")
                inter = [(first_src, fdt, row.get(C_COUNSEL, ""))]
        inper = [i for i in inter if start <= i[1] <= end]
        if not inper:
            continue
        first_dt = min(i[1] for i in inter) if inter else None
        row["_ninper"] = len(inper)
        row["_srcs_inper"] = [i[0] for i in inper]
        # (source, datetime) of every in-period interaction — used ONLY by the new
        # Lead Source Share by hour/day report; nothing existing reads it.
        row["_inper_pairs"] = [(i[0], i[1]) for i in inper]
        row["_is_new"] = bool(first_dt and start <= first_dt <= end)
        row["_first_dt"] = first_dt          # first-ever enquiry datetime (for day-wise trend)
        # Originating channel = the source of the lead's EARLIEST interaction
        # (first non-blank). Each lead has exactly ONE, so grouping leads by it
        # gives a clean partition that reconciles with the Executive Summary.
        row["_primary_src"] = next(
            (s(src) for src, _dt, _cb in sorted(inter, key=lambda i: i[1]) if s(src)),
            "")
        active.append(row)
    return active


def exec_summary(active):
    total = len(active)
    inter = sum(a["_ninper"] for a in active)
    new = sum(1 for a in active if a["_is_new"])
    fup = total - new
    valid = sum(1 for a in active if yes(a.get(C_VALID)))
    invalid = sum(1 for a in active if s(a.get(C_VALID)).lower() == "no")
    rel = sum(1 for a in active if yes(a.get(C_RELEV)))
    irr = sum(1 for a in active if s(a.get(C_RELEV)).lower() == "no")
    ref = sum(1 for a in active if is_referral(a))
    gmeet = sum(1 for a in active if yes(a.get(C_GMEET)))
    walk = sum(1 for a in active if yes(a.get(C_WALKSCH)))
    adm = sum(1 for a in active if is_admission_confirmed(a))
    lost = sum(1 for a in active if is_lost(a))
    return OrderedDict([
        ("Total Leads (Fresh + Repeat)", (total, "")),
        ("Fresh (New) Leads", (new, pct(new, total))),
        ("Repeat Leads", (fup, pct(fup, total))),
        ("Referral Leads", (ref, pct(ref, total))),
        ("Total Lead Interactions", (inter, "")),
        ("Valid Phone Number Leads", (valid, pct(valid, total))),
        ("Invalid Phone Number Leads", (invalid, pct(invalid, total))),
        ("Relevant Leads", (rel, pct(rel, total))),
        ("Irrelevant Leads", (irr, pct(irr, total))),
        ("Google Meets Scheduled", (gmeet, pct(gmeet, total))),
        ("Walk-ins Scheduled", (walk, pct(walk, total))),
    ])


def source_perf(active):
    # Count each lead under EVERY acquisition channel it actually USED WITHIN THE
    # PERIOD (from _srcs_inper — the source of each in-period interaction), NOT by
    # the channel of its earliest-ever interaction. A daily "Lead Source
    # Performance" has to answer "how many leads came through each channel today":
    # a repeat lead whose first enquiry was weeks ago by Call but who WALKS IN today
    # must be counted under Walk-In. The previous single-attribution-by-origin model
    # missed exactly that (it counted such a walk-in under Call, and it counted a
    # lead whose earliest-ever touch was a walk-in as a walk-in even on a day it did
    # NOT walk in) — which under-counted the true daily Walk-In leads.
    #
    # Because a lead can use several channels in one day, it is counted under each
    # channel it used, so these source rows measure CHANNEL ACTIVITY and can sum to
    # more than the distinct active-lead total.
    #
    # "% Contribution" denominator = the total of ONLY the four real acquisition
    # channels (Walk-In + Website + WhatsApp + Call) — NOT Total Leads
    # (Fresh + Repeat) — so each source's share is out of the source total and the
    # four sources sum to ~100%. (Any extra/internal label such as IntelliBI
    # follow-up is NOT part of this denominator.)
    source_total = sum(
        sum(1 for a in active if lbl in a.get("_srcs_inper", []))
        for lbl, _col in SOURCES)
    # Known acquisition channels first (always shown, in configured order), then any
    # other in-period source label that actually occurs (e.g. the internal IntelliBI
    # follow-up), so no lead's in-period activity is ever hidden.
    known = [label for label, _col in SOURCES]
    present = []
    for a in active:
        for lbl in a.get("_srcs_inper", []):
            if lbl and lbl not in present:
                present.append(lbl)
    extra = sorted(l for l in present if l not in known)
    rows = []
    for label in known + extra:
        g = [a for a in active if label in a.get("_srcs_inper", [])]
        n = len(g)                                       # leads that used this channel today
        fresh = sum(1 for a in g if a["_is_new"])        # of those, first enquiry in period
        repeat = n - fresh                               # of those, first enquiry earlier
        # IntelliBI is the internal portal. Its Fresh/Repeat split is DATA-DRIVEN
        # from each lead's own history (never hard-coded): a lead counts as Fresh
        # here ONLY if the portal touch is its FIRST-EVER enquiry — i.e. the lead's
        # earliest interaction of all (_primary_src) is IntelliBI, meaning it had no
        # prior enquiry through any other source. If it already enquired earlier via
        # any channel, it is a Repeat. So IntelliBI Fresh > 0 iff a genuinely new
        # lead is actually received through the portal; in the usual case (every
        # IntelliBI lead came in via a real channel first) it computes to 0.
        if is_intellibi_src(label):
            fresh = sum(1 for a in g
                        if a["_is_new"] and is_intellibi_src(a.get("_primary_src")))
            repeat = n - fresh
        inter = sum(a.get("_srcs_inper", []).count(label) for a in g)  # in-period touches via this channel
        valid = sum(1 for a in g if yes(a.get(C_VALID)))
        rel = sum(1 for a in g if yes(a.get(C_RELEV)))
        irr = sum(1 for a in g if s(a.get(C_RELEV)).lower() == "no")
        ref = sum(1 for a in g if is_referral(a))
        gm = sum(1 for a in g if yes(a.get(C_GMEET)))
        wk = sum(1 for a in g if yes(a.get(C_WALKSCH)))
        # Order: Source | Total | Unique(Fresh) | Repeat | Interactions | Valid |
        #        Relevant | Irrelevant | Referral | GMeet | Walk-in | % Contribution
        rows.append([label, n, fresh, repeat, inter, valid, rel, irr, ref, gm, wk,
                     pct(n, source_total)])
    return rows


def group_by_counsellor(active):
    """Group active leads by counsellor, case/space-insensitively, so name
    variants (e.g. 'Arshkhan Pathan' vs 'ArshKhan Pathan') don't fragment.
    Returns (groups_by_key, display_name_by_key)."""
    groups = defaultdict(list)
    forms = defaultdict(lambda: defaultdict(int))
    for a in active:
        raw = s(a.get(C_COUNSEL)) or "(Unassigned)"
        key = re.sub(r"\s+", " ", raw).strip().casefold()
        groups[key].append(a)
        forms[key][raw] += 1
    disp = {k: max(v.items(), key=lambda kv: kv[1])[0] for k, v in forms.items()}
    return groups, disp


def counsellor_perf(active):
    total = len(active)
    groups, disp = group_by_counsellor(active)
    rows = []
    for key in sorted(groups, key=lambda k: (-len(groups[k]), disp[k])):
        g = groups[key]
        n = len(g)                                       # Total = Fresh + Repeat
        fresh = sum(1 for a in g if a["_is_new"])
        repeat = n - fresh
        # Order: Counselling By | Total | Unique(Fresh) | Repeat | Interactions |
        #        Valid | Relevant | Irrelevant | Referral | GMeet | Walk-in | % Contribution
        rows.append([
            disp[key], n, fresh, repeat,
            sum(a["_ninper"] for a in g),
            sum(1 for a in g if yes(a.get(C_VALID))),
            sum(1 for a in g if yes(a.get(C_RELEV))),
            sum(1 for a in g if s(a.get(C_RELEV)).lower() == "no"),
            sum(1 for a in g if is_referral(a)),
            sum(1 for a in g if yes(a.get(C_GMEET))),
            sum(1 for a in g if yes(a.get(C_WALKSCH))),
            pct(n, total),
        ])
    return groups, rows


def course_breakdown(active):
    cnt = defaultdict(int)
    for a in active:
        c = s(a.get(C_COURSE)) or s(a.get(C_COURSE_ADV))
        if c:
            cnt[c] += 1
    return sorted(cnt.items(), key=lambda kv: -kv[1])


# ============================================================
# Tab building
# ============================================================
LEAD_COLS = [C_FIRST, C_LATEST, C_NAME, C_MOBILE, C_EMAIL, C_PLAT, "_ninper",
             "New/Follow-Up", C_VALID, C_RELEV, C_REF, C_REFNAME, C_COURSE,
             C_STATUS, C_ADM, C_BACKOUT, C_COUNSEL, C_GMEET, C_WALKSCH]
LEAD_HEADERS = ["First Enquiry", "Latest Enquiry", "Full Name", "Mobile Number",
                "Email", "Platforms Used", "Interactions (in period)", "Lead Type",
                "Phone Valid", "Relevant", "Is Referral", "Referrer's Name",
                "Course Interested", "Lead Status", "Admission Status",
                "Backout Reason", "Counselling By", "Google Meet Sch.", "Walk-in Sch.",
                "Lead Journey (Enquiry → Latest)"]


def format_journey(hist_text):
    """Render the master's 'Lead Interaction History' as a readable, chronological
    numbered journey (one interaction per line): initial enquiry → latest."""
    inter = parse_history(hist_text)
    if not inter:
        return ""
    steps = []
    for i, (src, dt, cby) in enumerate(sorted(inter, key=lambda x: x[1]), 1):
        when = dt.strftime("%d-%b-%Y %I:%M %p")
        who = f"  ·  {cby}" if cby and cby.strip() not in ("", "-") else ""
        steps.append(f"{i}. {when}  ·  {src or '—'}{who}")
    return "\n".join(steps)


def platforms_in_sequence(row):
    """'Platforms Used' ordered by the lead's ACTUAL enquiry datetime (earliest
    first), de-duplicated — derived from the chronological Lead Interaction
    History. Falls back to the stored value if the history can't be parsed; any
    platform present in the stored value but not in the history is appended at
    the end so a platform is never lost."""
    seq = []
    for src, dt, cby in sorted(parse_history(row.get(C_HIST)), key=lambda x: x[1]):
        name = s(src)
        if name and name != "—" and name not in seq:
            seq.append(name)
    for p in [p.strip() for p in s(row.get(C_PLAT)).split(",") if p.strip()]:
        if p not in seq:
            seq.append(p)
    return ", ".join(seq) if seq else s(row.get(C_PLAT))


def lead_detail_rows(tab, leads):
    tab.header(LEAD_HEADERS, filterable=True)
    # Sort every lead-detail tab by the First Enquiry datetime, ascending
    # (oldest first). Uses the parsed datetime, not text; blanks sort last.
    for a in sorted(leads, key=lambda x: (parse_dt(x.get(C_FIRST)) or datetime.max)):
        tab.row([
            a.get(C_FIRST), a.get(C_LATEST), a.get(C_NAME), a.get(C_MOBILE),
            a.get(C_EMAIL), platforms_in_sequence(a), a["_ninper"],
            "Fresh" if a["_is_new"] else "Repeat",
            a.get(C_VALID), a.get(C_RELEV), a.get(C_REF), a.get(C_REFNAME),
            a.get(C_COURSE), a.get(C_STATUS), a.get(C_ADM), a.get(C_BACKOUT),
            a.get(C_COUNSEL), a.get(C_GMEET), a.get(C_WALKSCH),
            format_journey(a.get(C_HIST)),
        ])


# Headline exec-summary metrics rendered in bold.
EXEC_KPI = {"Total Leads (Fresh + Repeat)", "Fresh (New) Leads", "Repeat Leads",
            "Referral Leads"}

SRC_HEADER = ["Source", "Total Leads", "Fresh (New) Leads", "Repeat Leads",
              "Interactions", "Valid", "Relevant", "Irrelevant", "Referral",
              "GMeet Sch.", "Walk-in Sch.", "% Contribution"]
CB_HEADER = ["Counselling By", "Total Leads", "Fresh (New) Leads", "Repeat Leads",
             "Interactions", "Valid", "Relevant", "Irrelevant", "Referral",
             "GMeet Sch.", "Walk-in Sch.", "% Contribution"]


def add_report_header(tab, title_text, period_range, gen_stamp, multiline=False):
    """One merged header row combining title + reporting period + generated time
    (applies to every report sheet). multiline=True stacks the three parts on
    three separate lines (wrapped) so a long title never truncates on a narrow
    tab; the single-line form is unchanged for every other tab."""
    if multiline:
        tab.banner(f"{title_text}\nReporting Period:  {period_range}\n"
                   f"Generated:  {gen_stamp}")
        tab.multiline_banner = True
    else:
        tab.banner(f"{title_text}   |   Reporting Period:   {period_range}"
                   f"   |   Generated:   {gen_stamp}")
    tab.blank()


def add_exec_summary(tab, active):
    tab.title("Executive Summary")
    tab.header(["Metric", "Value", "% of Total Leads"])
    for metric, (val, p) in exec_summary(active).items():
        tab.row([metric, val, p], kpi=(metric in EXEC_KPI))
    tab.blank()


def build_summary_tab(period_label, period_range, active, gen_stamp,
                      start=None, end=None):
    t = Tab("Summary")
    add_report_header(
        t, f"IntelliBI  -  Lead Performance Report  ({period_label})",
        period_range, gen_stamp)

    # Use the (otherwise blank) second line for the fresh non-referral metrics —
    # bold, size 13, with the SAME colour-coding as the email. Referral leads are
    # excluded; Fresh-Relevant counts the Relevant leads among the non-referral
    # fresh ones. Rendered as a colour-run rich-text line (col A, overflow band).
    _SEP = "     |     "
    fresh_nonref = sum(1 for a in active
                       if a.get("_is_new") and not is_referral(a))
    fresh_rel_nonref = sum(1 for a in active
                           if a.get("_is_new") and not is_referral(a)
                           and yes(a.get(C_RELEV)))
    rel_pct = (fresh_rel_nonref / fresh_nonref * 100.0) if fresh_nonref else 0.0
    segs = None

    if period_label == "Daily":
        # Daily: colour Fresh by its count > 15; the Fresh-Relevant count and %
        # by that % > 80.
        fresh_rgb, fresh_hex = ((TXT_GREEN, TXT_GREEN_HEX) if fresh_nonref > 15
                                else (TXT_RED, TXT_RED_HEX))
        rel_rgb, rel_hex = ((TXT_GREEN, TXT_GREEN_HEX) if rel_pct > 80
                            else (TXT_RED, TXT_RED_HEX))
        segs = [
            ("Fresh (Non-Referral):  ", CLR_SUB_FG, HEX["SUB_FG"]),
            (str(fresh_nonref), fresh_rgb, fresh_hex),
            (_SEP + "Fresh-Relevant (Non-Referral):  ", CLR_SUB_FG, HEX["SUB_FG"]),
            (str(fresh_rel_nonref), rel_rgb, rel_hex),
            (_SEP + "Fresh-Relevant (Non-Referral):  ", CLR_SUB_FG, HEX["SUB_FG"]),
            (f"{rel_pct:.1f}%", rel_rgb, rel_hex),
        ]
    elif period_label in ("Weekly", "Monthly") and start is not None and end is not None:
        # Weekly/Monthly: add per-day averages and colour all four the same way
        # the email does — the two Fresh figures share the Avg-Daily-Fresh > 15
        # verdict; the two Relevant figures share the Avg-Relevant-% > 80 verdict.
        last = min(end.date(), now_ist().date())
        ndays = max((last - start.date()).days + 1, 1)
        avg_daily_fresh = fresh_nonref / ndays
        avg_daily_fresh_rel = fresh_rel_nonref / ndays
        fresh_rgb, fresh_hex = ((TXT_GREEN, TXT_GREEN_HEX) if avg_daily_fresh > 15
                                else (TXT_RED, TXT_RED_HEX))
        rel_rgb, rel_hex = ((TXT_GREEN, TXT_GREEN_HEX) if rel_pct > 80
                            else (TXT_RED, TXT_RED_HEX))
        segs = [
            ("Non-Referral →   Fresh:  ", CLR_SUB_FG, HEX["SUB_FG"]),
            (str(fresh_nonref), fresh_rgb, fresh_hex),
            (_SEP + "Fresh-Relevant:  ", CLR_SUB_FG, HEX["SUB_FG"]),
            (str(fresh_rel_nonref), rel_rgb, rel_hex),
            (_SEP + "Avg Daily Fresh:  ", CLR_SUB_FG, HEX["SUB_FG"]),
            (f"{avg_daily_fresh:.2f}", fresh_rgb, fresh_hex),
            (_SEP + "Avg Daily Fresh-Relevant:  ", CLR_SUB_FG, HEX["SUB_FG"]),
            (f"{avg_daily_fresh_rel:.2f}", fresh_rgb, fresh_hex),
            (_SEP + "Avg Daily Fresh-Relevant %:  ", CLR_SUB_FG, HEX["SUB_FG"]),
            (f"{rel_pct:.1f}%", rel_rgb, rel_hex),
        ]

    if segs:
        text = "".join(seg[0] for seg in segs)
        t.rows[1] = [text]                        # replace the blank second line
        t.avg_line = {"row": 1, "segments": segs, "text": text, "size": 13}
        t.blank()                                 # separator before the table

    add_exec_summary(t, active)

    t.title("Lead Source Performance")
    t.header(SRC_HEADER)
    for r in source_perf(active):
        t.row(r)
    t.blank()

    t.title("Counsellor Performance")
    t.header(CB_HEADER)
    _, crows = counsellor_perf(active)
    for r in crows:
        t.row(r)
    t.blank()

    t.title("Course Interest")
    t.header(["Course Interested In", "Leads"])
    cb = course_breakdown(active)
    if cb:
        for course, n in cb:
            t.row([course, n])
    else:
        t.row(["(none recorded)", 0])
    t.blank()

    t.title("Scheduled Follow-ups (Meetings)")
    t.header(["Metric", "Count"])
    t.row(["Walk-ins Scheduled", sum(1 for a in active if yes(a.get(C_WALKSCH)))])
    t.row(["Google Meets Scheduled", sum(1 for a in active if yes(a.get(C_GMEET)))])
    return t


def sanitize_tab_name(name):
    name = re.sub(r"[:\\/?*\[\]]", " ", name).strip()
    return (name[:99]) if len(name) > 99 else (name or "Sheet")


# ============================================================
# Trend tab — day-wise (Weekly/Monthly) and hourly (Daily)
# ============================================================
DAYWISE_TAB_NAME = "Day-wise Trend"
HOURLY_TAB_NAME  = "Hourly Trend"
# Display names only. The two trend metrics are, per period bucket:
#   "Fresh Leads" = Fresh (Non-Referral)          = new leads excluding referrals
#   "Relevant"    = Fresh-Relevant (Non-Referral) = of those, the Relevant ones
DW_M1_LABEL = "Fresh Leads"          # was "Total Fresh"  (Fresh, excl. referrals)
DW_M2_LABEL = "Relevant"             # was "Fresh Relevant" (Relevant of the above)
# Requested chart line colours: green = Fresh Leads, brown = Relevant.
TREND_GREEN = (0.263, 0.627, 0.278)   # #43A047
TREND_BROWN = (0.627, 0.322, 0.176)   # #A0522D
TREND_GREEN_HEX = "43A047"
TREND_BROWN_HEX = "A0522D"
# Row-highlight fills for the day-wise conditional colouring.
ROW_GREEN = (0.78, 0.90, 0.74); ROW_GREEN_HEX = "C6EFCE"
ROW_RED   = (0.97, 0.80, 0.78); ROW_RED_HEX   = "FFC7CE"
# Bold text colours for the header average metrics (font, not fill).
TXT_GREEN = (0.13, 0.50, 0.15); TXT_GREEN_HEX = "217A26"
TXT_RED   = (0.78, 0.11, 0.11); TXT_RED_HEX   = "C71C1C"
NAVY_HEX  = "1B355E"                  # default (non-colour-coded) email KPI value
# How many columns the day-wise header colour band spans. The full single-line
# header (and the combined average line) lives in column A and OVERFLOWS across
# this many empty, same-coloured cells, so a long header shows in full on one
# line without truncation or merging the narrow table columns.
TREND_HEADER_SPAN = 22

# ── Hourly Lead Type Analysis tab ─────────────────────────────────────────────
# New tab (after the trend tab) cross-tabbing enquiry HOUR × Lead Type. The Lead
# Type value comes straight from the consolidated master's "Lead Type" column, so
# the categories are driven by the consolidation config sheet — never hard-coded
# here. LEAD_TYPE_ORDER only fixes the DISPLAY order of the known categories; any
# category the config adds later still appears (as an extra column before
# 'Unidentified'), and blank/unknown values fall under 'Unidentified'.
LEADTYPE_TAB_NAME      = "Hourly Lead Type Analysis"
C_LEADTYPE             = "Lead Type"
LEAD_TYPE_UNIDENTIFIED = "Unidentified"
# Fixed BUSINESS display/sort order (not alphabetical). 'Unidentified' is always
# forced last by _ordered_lead_types(). "Student - Pursuing" is the current config
# category name; "Fresher - Pursuing" is listed alongside it so the order still
# holds if that category is ever renamed in the config sheet.
LEAD_TYPE_ORDER = [
    "Working Professional - IT", "Working Professional - Non IT",
    "Fresher - Passed Out", "Student - Pursuing", "Fresher - Pursuing",
    "Career Break", "Unreachable",
]
# One distinct, legible colour per Lead Type for the stacked-column chart.
LEAD_TYPE_COLORS_HEX = {
    "Working Professional - IT":      "2B6CB0",   # blue
    "Working Professional - Non IT":  "38A169",   # green
    "Fresher - Passed Out":           "DD6B20",   # orange
    "Student - Pursuing":             "805AD5",   # purple
    "Fresher - Pursuing":             "805AD5",   # purple (same category, alt name)
    "Career Break":                   "D53F8C",   # magenta
    "Unreachable":                    "8B0000",   # dark red
    LEAD_TYPE_UNIDENTIFIED:           "A0AEC0",   # grey (unchanged)
}
_DEFAULT_LT_HEX = "718096"                        # any future/unexpected category
# Blank-row gap between the stacked hourly chart (which floats ~24 rows tall) and
# the Lead Type Share table/pie added below it, so the two never overlap.
LEADTYPE_CHART_GAP = 24


def _hex_rgb(h):
    """'2B6CB0' -> (0.169, 0.424, 0.690) for the Google Sheets chart colour API."""
    return (int(h[0:2], 16) / 255.0, int(h[2:4], 16) / 255.0, int(h[4:6], 16) / 255.0)


def _lt_hex(lead_type):
    return LEAD_TYPE_COLORS_HEX.get(lead_type, _DEFAULT_LT_HEX)


def _lt_rgb(lead_type):
    return _hex_rgb(_lt_hex(lead_type))


def _relpct_str(total_fresh, fresh_relevant):
    """Relevant % = (Fresh Relevant / Total Fresh) × 100, as a 1-dp string."""
    return f"{(fresh_relevant / total_fresh * 100.0):.1f}%" if total_fresh else "0.0%"


def daywise_trend(active, start, end):
    """Day-wise counts from the period start up to the report date (whichever of
    the period end / today is earlier). Each lead is attributed to the DAY IT WAS
    RECEIVED (its first-ever enquiry date), so only leads whose first enquiry
    falls in the period ('fresh'/new leads) contribute. Referral leads are
    excluded up front, so per day:
        Fresh Leads = Fresh (Non-Referral)          = new leads, excl. referrals
        Relevant    = Fresh-Relevant (Non-Referral) = of those, the Relevant ones
    Returns [(date, fresh_nonref, fresh_rel_nonref)] for every calendar day in
    the range (missing days show 0)."""
    last = min(end.date(), now_ist().date())
    fresh_by, rel_by = defaultdict(int), defaultdict(int)
    for a in active:
        if not a.get("_is_new"):
            continue                              # only leads received in the period
        if is_referral(a):
            continue                              # exclude referral leads entirely
        fd = a.get("_first_dt")
        if not fd:
            continue
        d = fd.date()
        if d < start.date() or d > last:
            continue
        fresh_by[d] += 1                          # Fresh (Non-Referral)
        if yes(a.get(C_RELEV)):
            rel_by[d] += 1                        # Fresh-Relevant (Non-Referral)
    rows = []
    d = start.date()
    while d <= last:
        rows.append((d, fresh_by.get(d, 0), rel_by.get(d, 0)))
        d += timedelta(days=1)
    return rows


def _hour_label(h):
    """0..23 -> '12 AM','1 AM',…,'12 PM','1 PM',…"""
    ap = "AM" if h < 12 else "PM"
    hh = h % 12 or 12
    return f"{hh} {ap}"


def hourly_trend(active, day):
    """Hourly counts for a single day, grouped by the ENQUIRY hour of each lead
    received that day (first-enquiry hour). Same two metrics as daywise_trend:
    Fresh Leads = Fresh (Non-Referral) and Relevant = Fresh-Relevant (Non-Referral),
    with referral leads excluded up front. Returns [(hour_label, fresh_nonref,
    fresh_rel_nonref)] for the span of hours from the first to the last hour that
    had a lead (gaps filled with 0). [] if none."""
    fresh_by, rel_by = defaultdict(int), defaultdict(int)
    hours = set()
    for a in active:
        if not a.get("_is_new"):
            continue
        if is_referral(a):
            continue                              # exclude referral leads entirely
        fd = a.get("_first_dt")
        if not fd or fd.date() != day:
            continue
        h = fd.hour
        hours.add(h)
        fresh_by[h] += 1                          # Fresh (Non-Referral)
        if yes(a.get(C_RELEV)):
            rel_by[h] += 1                        # Fresh-Relevant (Non-Referral)
    if not hours:
        return []
    rows = []
    for h in range(min(hours), max(hours) + 1):
        rows.append((_hour_label(h), fresh_by.get(h, 0), rel_by.get(h, 0)))
    return rows


def _build_trend_tab(tab_name, banner_title, chart_title, period_range, gen_stamp,
                     section_title, headers, table_rows, series_cols, x_title,
                     legend_g="BOTTOM_LEGEND", legend_x="b", avg_lines=None):
    """Shared builder: header banner + optional average-metric lines + a data
    table (whose last column is Relevant %) + a green/brown line chart BELOW the
    table. Line 1 is the full report header on ONE single line (title | reporting
    period | generated), overflowing across a wide colour band so it never
    truncates. Each avg_line is (label, value, value_rgb, value_hex) rendered
    right below the main header with the value bold and colour-coded."""
    t = Tab(tab_name)
    t.color_data_rows = False        # keep the trend table plain (the % column
    #                                  must not trigger the traffic-light shading)
    # Line 1: full report header on ONE line. The text lives in column A and
    # overflows across TREND_HEADER_SPAN same-coloured empty cells, so the
    # complete header always shows on a single line without truncation.
    full_header = (f"{banner_title}   |   Reporting Period:   {period_range}"
                   f"   |   Generated:   {gen_stamp}")
    t.banner(full_header)
    t.wide_banner = True
    t.header_span = TREND_HEADER_SPAN
    t.blank()
    # Average-metric lines right under the main header (label merged A:C, value
    # in col D bold + green/red). Tracked in t.avg_rows for the writers.
    if avg_lines:
        t.avg_rows = []
        for label, value, rgb, hexc in avg_lines:
            ri = len(t.rows)
            row = [label, "", "", value] + [""] * (max(len(headers), 4) - 4)
            t.rows.append(row)
            t.avg_rows.append({"row": ri, "value_col": 3, "rgb": rgb, "hex": hexc})
        t.blank()
    hdr_idx = len(t.rows)
    t.header(headers)
    first_data = len(t.rows)
    for row in table_rows:
        t.row(row)
    if table_rows:
        last_data = len(t.rows) - 1
        t.chart = {
            "title": chart_title,
            "header_row": hdr_idx, "first_row": first_data, "last_row": last_data,
            "date_col": 0, "series_cols": series_cols,
            "series_titles": [DW_M1_LABEL, DW_M2_LABEL],
            "series_colors": [TREND_GREEN, TREND_BROWN],
            "series_colors_hex": [TREND_GREEN_HEX, TREND_BROWN_HEX],
            "x_title": x_title, "y_title": "Number of Leads",
            "anchor_row": last_data + 2, "anchor_col": 0,   # BELOW the table
            "legend_g": legend_g, "legend_x": legend_x,
        }
    else:
        t.row(["(no data)"] + [0] * (len(headers) - 2) + ["0.0%"])
    return t


# ── Lead Source Share (by hour/day) ──────────────────────────────
# A NEW report appended below the existing Fresh-vs-Relevant table+graph. Counts
# each in-period interaction under its own channel (the finalized _srcs_inper
# classification) and its own day/hour, for the four acquisition channels only.
# One distinct, legible colour per source for its multi-line trend chart. Four
# well-separated hues: Walk-In green, Website blue, WhatsApp purple, Call orange.
# (Walk-In & Call keep their colours; WhatsApp moved off teal so it no longer
# reads like Website's blue — the four are now easy to tell apart.)
SOURCE_COLORS_HEX = {"Walk-In": "2E7D32", "Website": "1565C0",
                     "WhatsApp": "8E24AA", "Call": "EF6C00"}

# Header for the hidden helper column that carries the graph's short X-axis label.
AXIS_LABEL_HDR = "AxisLabel"
# Light fill for weekend (Sat/Sun) rows in the day-wise tables.
WEEKEND_BG_HEX = "FFF2CC"


def _is_weekend(d):
    return d.weekday() >= 5          # Saturday (5) or Sunday (6)


def _axis_date_label(d):
    """Short graph X-axis label from a date: 'dd-Mon-yy|<weekday initial>', e.g.
    01-Aug-2026 (a Saturday) -> '01-Aug-26|S'. Weekday initial is derived
    dynamically from the date (S/M/T/W/T/F/S). Weekend (Sat/Sun) labels are
    WRAPPED in brackets — e.g. '[01-Aug-26|S]' — to highlight them on the axis
    (a spreadsheet chart cannot bold individual axis labels)."""
    lbl = f"{d.strftime('%d-%b-%y')}|{d.strftime('%a')[0]}"
    return f"[{lbl}]" if _is_weekend(d) else lbl


def source_daywise(active, start, end):
    """Per-DAY lead-source interaction counts for the four channels, from the
    period start to the report date (period end or today, whichever is earlier).
    Returns [(date_display, [Walk-In, Website, WhatsApp, Call], total)] for every
    day (missing days show 0), matching the day-wise trend's date span."""
    last = min(end.date(), now_ist().date())
    by = defaultdict(lambda: defaultdict(int))
    for a in active:
        for src, dt in a.get("_inper_pairs", []):
            if src in SOURCE_LABELS and start.date() <= dt.date() <= last:
                by[dt.date()][src] += 1
    rows, d = [], start.date()
    while d <= last:
        counts = by.get(d, {})
        vals = [counts.get(sl, 0) for sl in SOURCE_LABELS]
        rows.append((d, vals, sum(vals)))          # date object (formatted by caller)
        d += timedelta(days=1)
    return rows


def source_hourly(active, day):
    """Per-HOUR lead-source interaction counts for the four channels on a single
    day. Returns [(hour_label, [Walk-In, Website, WhatsApp, Call], total)] across
    the span of hours that had activity (gaps filled with 0); [] if none."""
    by = defaultdict(lambda: defaultdict(int))
    hours = set()
    for a in active:
        for src, dt in a.get("_inper_pairs", []):
            if src in SOURCE_LABELS and dt.date() == day:
                by[dt.hour][src] += 1
                hours.add(dt.hour)
    if not hours:
        return []
    rows = []
    for h in range(min(hours), max(hours) + 1):
        counts = by.get(h, {})
        vals = [counts.get(sl, 0) for sl in SOURCE_LABELS]
        rows.append((_hour_label(h), vals, sum(vals)))
    return rows


def append_source_share(t, buckets, x_label, period_label, weekday=False):
    """Append the Lead Source Share table (+ multi-line trend chart) to a trend
    tab, placed well BELOW the tab's first (Fresh-vs-Relevant) floating chart so
    the two never overlap.

    weekday=True  (Weekly / Monthly day-wise): each bucket's x is a DATE object;
      the table shows 'Date' then a 'Day Of Week' column, and the chart X-axis
      reads a short 'dd-Mon-yy|<weekday initial>' label from a HIDDEN helper
      column (so the visible table keeps the full date/weekday).
    weekday=False (Daily hourly): x is the hour label (unchanged behaviour).

    buckets = [(x, [4 source counts], total)]. Sets t.chart2."""
    title = f"Lead Source Share - {period_label}"
    # clear the vertical space the first floating chart occupies
    base = t.chart["anchor_row"] if getattr(t, "chart", None) else len(t.rows)
    while len(t.rows) < base + LEADTYPE_CHART_GAP:
        t.blank()
    t.title(title)
    if weekday:
        # The HIDDEN axis-label column sits at index 6 (same column the Fresh
        # table hides), BEFORE Total, so hiding that one column never hides the
        # visible Total (which lands at index 7 and shows right after Call).
        headers = (["Date", "Day Of Week"] + list(SOURCE_LABELS)
                   + [AXIS_LABEL_HDR, "Total"])
        src_start = 2
        axis_col = len(SOURCE_LABELS) + 2          # Date, Day Of Week, 4 srcs, [axis], Total
    else:
        headers = [x_label] + list(SOURCE_LABELS) + ["Total"]
        src_start = 1
        axis_col = None
    hdr_idx = len(t.rows)
    t.header(headers)
    first_data = len(t.rows)
    if not buckets:
        t.row(["(no leads in this period)"] + [0] * (len(headers) - 1))
        return
    weekend_rows = []
    for x, vals, total in buckets:
        ri = len(t.rows)
        if weekday:
            t.row([x.strftime("%d-%b-%Y"), x.strftime("%A")] + list(vals)
                  + [_axis_date_label(x), total])
            if _is_weekend(x):
                weekend_rows.append(ri)
        else:
            t.row([x] + list(vals) + [total])
    last_data = len(t.rows) - 1                 # last day/hour row (chart stops here)

    # Average Leads per source (and Total), averaged dynamically over the shown
    # day/hour buckets for this Daily/Weekly/Monthly report. Bold + light band,
    # and EXCLUDED from the chart's line range (chart uses first_data..last_data).
    n = len(buckets)
    col_sums = [sum(b[1][j] for b in buckets) for j in range(len(SOURCE_LABELS))]
    avg_vals = [f"{(cs / n):.1f}" for cs in col_sums]
    avg_total = f"{(sum(b[2] for b in buckets) / n):.1f}"
    if weekday:
        t.row(["Average Leads", ""] + avg_vals + ["", avg_total])
    else:
        t.row(["Average Leads"] + avg_vals + [avg_total])
    avg_idx = len(t.rows) - 1
    t.kpi.add(avg_idx)
    if getattr(t, "row_fills", None) is None:
        t.row_fills = {}
    t.row_fills[avg_idx] = (CLR_HDR_BG, HEX["HDR"])
    # Highlight weekend (Sat/Sun) rows: bold + a light shade.
    for ri in weekend_rows:
        t.kpi.add(ri)
        t.row_fills[ri] = (_hex_rgb(WEEKEND_BG_HEX), WEEKEND_BG_HEX)

    series_cols = list(range(src_start, src_start + len(SOURCE_LABELS)))
    t.chart2 = {
        "title": title,
        "header_row": hdr_idx, "first_row": first_data, "last_row": last_data,
        "date_col": axis_col if weekday else 0, "series_cols": series_cols,
        "series_titles": list(SOURCE_LABELS),
        "series_colors":     [_hex_rgb(SOURCE_COLORS_HEX[s]) for s in SOURCE_LABELS],
        "series_colors_hex": [SOURCE_COLORS_HEX[s] for s in SOURCE_LABELS],
        "x_title": x_label, "y_title": "Number of Leads",
        "anchor_row": avg_idx + 2, "anchor_col": 0,
        "legend_g": "BOTTOM_LEGEND", "legend_x": "b",
    }
    if weekday:
        if getattr(t, "hidden_cols", None) is None:
            t.hidden_cols = set()
        t.hidden_cols.add(axis_col)


def build_daywise_tab(period_label, period_range, active, start, end, gen_stamp):
    """Weekly/Monthly day-wise tab:
        Day | Day of Week | Date | Fresh Leads | Relevant | Relevant %
    Line 1 is the full report header on ONE single line (title | reporting period
    | generated) — it overflows across a wide colour band so it never truncates.
    Line 2 combines both average metrics on ONE single line, each value bold and
    colour-coded. The table starts right after these two header lines. Chart
    X-axis = Day number; legend top-right; each row is highlighted GREEN when
    Fresh Leads > 15 AND Relevant % > 80, otherwise RED."""
    data = daywise_trend(active, start, end)          # [(date, m1, m2)]
    # Last column is a HIDDEN helper carrying the short graph X-axis label
    # (dd-Mon-yy|<weekday initial>); the visible Day / Day of Week / Date columns
    # and all values are unchanged.
    rows = [[d.day, d.strftime("%A"), d.strftime("%d-%b-%Y"), m1, m2,
             _relpct_str(m1, m2), _axis_date_label(d)]
            for d, m1, m2 in data]
    # Header average metrics: Daily Average Leads (= total Fresh Leads / #days,
    # green if >=15 else red) and Average Relevant % (= total Relevant / total
    # Fresh Leads × 100, green if >=80 else red).
    total_fresh = sum(m1 for _d, m1, _m2 in data)
    total_rel = sum(m2 for _d, _m1, m2 in data)
    ndays = len(data)
    avg_leads = (total_fresh / ndays) if ndays else 0.0
    avg_rel = (total_rel / total_fresh * 100.0) if total_fresh else 0.0

    def _tc(good):
        return (TXT_GREEN, TXT_GREEN_HEX) if good else (TXT_RED, TXT_RED_HEX)
    lead_rgb, lead_hex = _tc(avg_leads >= 15)
    rel_rgb, rel_hex = _tc(avg_rel >= 80)

    title = f"Day-wise Fresh Vs Relevant - {period_label}"
    t = Tab(DAYWISE_TAB_NAME)
    t.color_data_rows = False        # the % column must not trigger row shading

    # --- Line 1: full report header on ONE line. The text lives in column A and
    #     overflows across TREND_HEADER_SPAN same-coloured empty cells, so the
    #     complete header always shows on a single line without truncation. ---
    full_header = (f"{title}   |   Reporting Period:   {period_range}"
                   f"   |   Generated:   {gen_stamp}")
    t.banner(full_header)
    t.wide_banner = True
    t.header_span = TREND_HEADER_SPAN

    # --- Line 2: both averages on ONE line, each value bold + colour-coded.
    #     Rendered as rich text (Google textFormatRuns / openpyxl CellRichText)
    #     in column A over a light band spanning the same width. ---
    avg_segments = [
        ("Daily Average Leads:  ", CLR_SUB_FG, HEX["SUB_FG"]),
        (str(int(round(avg_leads))), lead_rgb, lead_hex),
        ("     |     Average Relevant %:  ", CLR_SUB_FG, HEX["SUB_FG"]),
        (f"{avg_rel:.1f}%", rel_rgb, rel_hex),
    ]
    avg_text = "".join(seg[0] for seg in avg_segments)
    ri_avg = len(t.rows)
    t.rows.append([avg_text])
    t.avg_line = {"row": ri_avg, "segments": avg_segments, "text": avg_text}
    t.blank()

    # --- table (starts right after the two header lines) ---
    headers = ["Day", "Day of Week", "Date", DW_M1_LABEL, DW_M2_LABEL, "Relevant %",
               AXIS_LABEL_HDR]
    hdr_idx = len(t.rows)
    t.header(headers)
    first_data = len(t.rows)
    for row in rows:
        t.row(row)
    if rows:
        last_data = len(t.rows) - 1
        t.chart = {
            "title": title,
            "header_row": hdr_idx, "first_row": first_data, "last_row": last_data,
            # X-axis reads the hidden short-label helper column (col 6), so the
            # graph shows dd-Mon-yy|<weekday initial> instead of the day number.
            "date_col": 6, "series_cols": [3, 4],
            "series_titles": [DW_M1_LABEL, DW_M2_LABEL],
            "series_colors": [TREND_GREEN, TREND_BROWN],
            "series_colors_hex": [TREND_GREEN_HEX, TREND_BROWN_HEX],
            "x_title": "Date", "y_title": "Number of Leads",
            "anchor_row": last_data + 2, "anchor_col": 0,   # BELOW the table
            "legend_g": "RIGHT_LEGEND", "legend_x": "tr",
        }
        t.hidden_cols = {6}
        # Row-level colour coding (green if both conditions hold, else red).
        fills = {}
        for i, (dday, m1, m2) in enumerate(data):
            relpct = (m2 / m1 * 100.0) if m1 else 0.0
            good = (m1 > 15) and (relpct > 80.0)
            fills[first_data + i] = (
                (ROW_GREEN, ROW_GREEN_HEX) if good else (ROW_RED, ROW_RED_HEX))
            if _is_weekend(dday):            # bold the weekend rows (keep green/red)
                t.kpi.add(first_data + i)
        t.row_fills = fills
    else:
        t.row(["(no data)"] + [0] * (len(headers) - 2) + ["0.0%"])

    # NEW: Lead Source Share by day, appended below the Fresh-vs-Relevant graph.
    append_source_share(t, source_daywise(active, start, end), "Date", period_label,
                        weekday=True)
    return t


def build_hourly_tab(period_label, period_range, active, day, gen_stamp):
    """Daily hourly tab: Hour | Fresh Leads | Relevant | Relevant %. Chart X-axis
    = hour. Each row is highlighted GREEN when the hour has at least 1 Fresh Lead
    AND Relevant % > 80, otherwise RED."""
    data = hourly_trend(active, day)                  # [(hour_label, m1, m2)]
    rows = [[hl, m1, m2, _relpct_str(m1, m2)] for hl, m1, m2 in data]
    title = f"Hourly Fresh vs Relevant - {period_label}"
    t = _build_trend_tab(
        HOURLY_TAB_NAME, title, title, period_range, gen_stamp, title,
        ["Hour", DW_M1_LABEL, DW_M2_LABEL, "Relevant %"],
        rows, [1, 2], "Hour")
    # Row-level colour coding (green if both conditions hold, else red).
    if data and t.chart:
        fills = {}
        for i, (_hl, m1, m2) in enumerate(data):
            relpct = (m2 / m1 * 100.0) if m1 else 0.0
            good = (m1 >= 1) and (relpct > 80.0)
            fills[t.chart["first_row"] + i] = (
                (ROW_GREEN, ROW_GREEN_HEX) if good else (ROW_RED, ROW_RED_HEX))
        t.row_fills = fills

    # NEW: Lead Source Share by hour, appended below the Fresh-vs-Relevant graph.
    append_source_share(t, source_hourly(active, day), "Hour", period_label)
    return t


def _leadtype_hourly(active, day=None):
    """Cross-tab of enquiry HOUR -> {Lead Type: count}, restricted to
    Fresh-Relevant (Non-Referral) leads only:
        * Fresh     — first enquiry falls in the period (_is_new)
        * Relevant  — IsLeadRelevant = Yes
        * Non-Referral — not a referral lead
    Repeat, Irrelevant and Referral leads are excluded. Uses each lead's
    first-enquiry hour and its 'Lead Type' from the master. day=None aggregates the
    whole period by hour-of-day (Weekly/Monthly); a specific date restricts to that
    day (Daily). Returns (by_hour, hours_set, present_types)."""
    by = defaultdict(lambda: defaultdict(int))
    hours, present = set(), set()
    for a in active:
        if not a.get("_is_new"):
            continue                          # Fresh only (drop Repeat leads)
        if is_referral(a):
            continue                          # Non-Referral only
        if not yes(a.get(C_RELEV)):
            continue                          # Relevant only (drop Irrelevant)
        fd = a.get("_first_dt")
        if not fd:
            continue
        if day is not None and fd.date() != day:
            continue
        lt = s(a.get(C_LEADTYPE)) or LEAD_TYPE_UNIDENTIFIED
        by[fd.hour][lt] += 1
        hours.add(fd.hour)
        present.add(lt)
    return by, hours, present


def _ordered_lead_types(present):
    """Column order: known categories first (fixed display order), then any extra
    categories the config introduced (alphabetical), then 'Unidentified' last."""
    known = [lt for lt in LEAD_TYPE_ORDER if lt in present]
    extra = sorted(lt for lt in present
                   if lt not in LEAD_TYPE_ORDER and lt != LEAD_TYPE_UNIDENTIFIED)
    tail = [LEAD_TYPE_UNIDENTIFIED] if LEAD_TYPE_UNIDENTIFIED in present else []
    return known + extra + tail


def build_leadtype_hourly_tab(period_label, period_range, active, gen_stamp, day=None):
    """Hourly Lead Type Analysis tab: Hour × Lead Type enquiry counts (+ Total row
    and Total column) with a stacked-column chart below. Applies to Daily (that
    day), Weekly and Monthly (period aggregated by hour-of-day)."""
    by, hours, present = _leadtype_hourly(active, day)
    lead_types = _ordered_lead_types(present)
    title = f"Hourly Lead Type Analysis - {period_label}"

    t = Tab(LEADTYPE_TAB_NAME)
    t.color_data_rows = False        # counts must not trigger traffic-light shading
    # Line 1: full report header on one line (wide overflow band, never truncates).
    full_header = (f"{title}   |   Reporting Period:   {period_range}"
                   f"   |   Generated:   {gen_stamp}")
    t.banner(full_header)
    t.wide_banner = True
    t.header_span = TREND_HEADER_SPAN
    t.blank()

    if not lead_types:
        lead_types = [LEAD_TYPE_UNIDENTIFIED]
    headers = ["Hour"] + lead_types + ["Total"]
    hdr_idx = len(t.rows)
    t.header(headers)
    first_data = len(t.rows)

    if hours:
        col_tot = defaultdict(int)
        grand = 0
        lo, hi = min(hours), max(hours)
        for h in range(lo, hi + 1):                 # continuous hour span (gaps -> 0)
            counts = by.get(h, {})
            rowvals, rtot = [_hour_label(h)], 0
            for lt in lead_types:
                c = int(counts.get(lt, 0))
                rowvals.append(c)
                col_tot[lt] += c
                rtot += c
            rowvals.append(rtot)
            grand += rtot
            t.row(rowvals)
        last_data = len(t.rows) - 1                 # last HOUR row (chart stops here)
        # Total row (bold, light band) — excluded from the chart categories.
        t.row(["Total"] + [col_tot[lt] for lt in lead_types] + [grand])
        total_idx = len(t.rows) - 1
        t.kpi.add(total_idx)
        t.row_fills = {total_idx: (CLR_HDR_BG, HEX["HDR"])}
        # Stacked-column chart: one series per Lead Type (Total column excluded).
        series_cols = list(range(1, 1 + len(lead_types)))
        t.chart = {
            "title": title, "chart_kind": "COLUMN_STACKED",
            "header_row": hdr_idx, "first_row": first_data, "last_row": last_data,
            "date_col": 0, "series_cols": series_cols,
            "series_titles": lead_types,
            "series_colors":     [_lt_rgb(lt) for lt in lead_types],
            "series_colors_hex": [_lt_hex(lt) for lt in lead_types],
            "x_title": "Hour of Day", "y_title": "Number of Enquiries",
            "anchor_row": total_idx + 2, "anchor_col": 0,
            "legend_g": "BOTTOM_LEGEND", "legend_x": "b",
        }

        # ---- Lead Type Share Analysis (count + share %) : small table + bar ----
        # Same Fresh-Relevant (Non-Referral) totals as above, one row per Lead
        # Type. Placed well below the stacked chart (which floats ~24 rows tall)
        # so the two charts never overlap.
        #
        # The chart is a horizontal bar (one bar per Lead Type). Google Sheets
        # drops on-bar value/percent data labels when it imports an xlsx chart,
        # so we bake BOTH numbers straight into each Lead Type's axis label,
        # e.g. "Unidentified  -  291 (61.8%)" = Lead Count 291, Lead Share 61.8%.
        # That is the one place the count + % render reliably in Google Sheets.
        while len(t.rows) < (total_idx + 2) + LEADTYPE_CHART_GAP:
            t.blank()
        t.title("Lead Type Share Analysis")
        share_hdr = len(t.rows)
        t.header(["Lead Type  -  Count (Share %)", "Lead Count"])
        share_first = len(t.rows)
        for lt in lead_types:                       # fixed business order
            cnt = int(col_tot[lt])
            pct = (cnt / grand * 100.0) if grand else 0.0
            t.row([f"{lt}  -  {cnt} ({pct:.1f}%)", cnt])
        share_last = len(t.rows) - 1
        t.row([f"Total  -  {int(grand)} (100.0%)", int(grand)])
        share_total = len(t.rows) - 1
        t.kpi.add(share_total)
        t.row_fills[share_total] = (CLR_HDR_BG, HEX["HDR"])
        # Horizontal bar: one bar per Lead Type, count shown on the bar, share %
        # in the axis label. Bars render Unidentified (top) -> WP-IT (bottom).
        t.bar_share = {
            "title": f"Lead Type Share - {period_label}",
            "cat_col": 0, "val_col": 1,
            "header_row": share_hdr, "first_row": share_first, "last_row": share_last,
            "colors_hex": [_lt_hex(lt) for lt in lead_types],
            "anchor_row": share_total + 2, "anchor_col": 0,
        }
    else:
        t.row(["(no enquiries in this period)"] + [0] * len(lead_types) + [0])
    return t


def _gchart_source(sid, r0, r1, c0, c1):
    return {"sourceRange": {"sources": [{"sheetId": sid,
            "startRowIndex": r0, "endRowIndex": r1,
            "startColumnIndex": c0, "endColumnIndex": c1}]}}


def _rgb(t):
    return {"red": t[0], "green": t[1], "blue": t[2]}


def build_gchart_request(sid, spec):
    """A Google Sheets addChart request from a Tab.chart spec. Default is a LINE
    chart (with markers + value labels); chart_kind='COLUMN_STACKED' produces a
    stacked column chart (one series per column, no per-point markers/labels)."""
    hr, lr = spec["header_row"], spec["last_row"]
    dcol = spec["date_col"]
    domain = {"domain": _gchart_source(sid, hr, lr + 1, dcol, dcol + 1)}
    is_col = spec.get("chart_kind") == "COLUMN_STACKED"
    # dataLabel type DATA shows ONLY the numeric value (e.g. "26"); pointStyle
    # puts a round marker on every point; colorStyle sets the requested colours.
    _dlabel = {"type": "DATA", "placement": "ABOVE", "textFormat": {"fontSize": 9}}
    colors = spec.get("series_colors") or []
    # Pair each series column with its colour so they stay together if reordered.
    pairs = [(c, colors[i] if i < len(colors) else None)
             for i, c in enumerate(spec["series_cols"])]
    # Google Sheets renders a STACKED-column legend in REVERSE series order. To make
    # the bottom legend read in the intended forward order (Working Professional -
    # IT first … Unidentified last), add the series in reverse here — the double
    # reversal yields a forward legend. (Line charts are unaffected.)
    if is_col:
        pairs = list(reversed(pairs))
    series = []
    for c, col in pairs:
        sser = {"series": _gchart_source(sid, hr, lr + 1, c, c + 1),
                "targetAxis": "LEFT_AXIS"}
        if not is_col:                              # lines get markers + value labels
            sser["dataLabel"] = _dlabel
            sser["pointStyle"] = {"size": 7, "shape": "CIRCLE"}
        if col is not None:
            sser["colorStyle"] = {"rgbColor": _rgb(col)}
        series.append(sser)
    basic = {
        "chartType": "COLUMN" if is_col else "LINE",
        "legendPosition": spec.get("legend_g", "BOTTOM_LEGEND"), "headerCount": 1,
        "axis": [{"position": "BOTTOM_AXIS", "title": spec.get("x_title", "Date")},
                 {"position": "LEFT_AXIS", "title": spec.get("y_title", "Number of Leads")}],
        "domains": [domain], "series": series}
    if is_col:
        basic["stackedType"] = "STACKED"
    return {"addChart": {"chart": {
        "spec": {"title": spec["title"], "basicChart": basic},
        "position": {"overlayPosition": {
            "anchorCell": {"sheetId": sid, "rowIndex": spec["anchor_row"],
                           "columnIndex": spec["anchor_col"]},
            "offsetXPixels": 5, "offsetYPixels": 5,
            "widthPixels": 940, "heightPixels": 470}}}}}


def build_report(period_label, period_range, df, start, end):
    """Return OrderedDict[tab_name -> Tab]."""
    active = prepare_active(df, start, end)
    fresh = [a for a in active if a["_is_new"]]           # first enquiry in period
    repeat = [a for a in active if not a["_is_new"]]      # enquired again (first was earlier)
    gen_stamp = now_ist().strftime("%d-%b-%Y %I:%M %p") + " IST"   # one stamp per run
    tabs = OrderedDict()
    tabs["Summary"] = build_summary_tab(period_label, period_range, active, gen_stamp,
                                        start=start, end=end)

    # Graphical trend tab, immediately after Summary:
    #   Weekly / Monthly -> day-wise;  Daily -> hourly (by enquiry hour).
    # Then, immediately after it, the Hourly Lead Type Analysis tab (all periods).
    if period_label in ("Weekly", "Monthly"):
        tabs[DAYWISE_TAB_NAME] = build_daywise_tab(
            period_label, period_range, active, start, end, gen_stamp)
        tabs[LEADTYPE_TAB_NAME] = build_leadtype_hourly_tab(
            period_label, period_range, active, gen_stamp, day=None)
    elif period_label == "Daily":
        tabs[HOURLY_TAB_NAME] = build_hourly_tab(
            period_label, period_range, active, start.date(), gen_stamp)
        tabs[LEADTYPE_TAB_NAME] = build_leadtype_hourly_tab(
            period_label, period_range, active, gen_stamp, day=start.date())

    # Two separate lead-detail tabs: Fresh (new in the period) and Repeat
    # (enquired again in the period via any of the four platforms).
    fresh_tab = Tab("Fresh Lead Details")
    add_report_header(fresh_tab, f"Fresh Lead Details  ({period_label})",
                      period_range, gen_stamp)
    fresh_tab.title(f"Fresh (New) Leads in Period  ({len(fresh)})")
    lead_detail_rows(fresh_tab, fresh)
    tabs["Fresh Lead Details"] = fresh_tab

    repeat_tab = Tab("Repeat Lead Details")
    add_report_header(repeat_tab, f"Repeat Lead Details  ({period_label})",
                      period_range, gen_stamp)
    repeat_tab.title(f"Repeat Leads in Period — enquired again via any platform  "
                     f"({len(repeat)})")
    lead_detail_rows(repeat_tab, repeat)
    tabs["Repeat Lead Details"] = repeat_tab

    groups, disp = group_by_counsellor(active)
    for key in sorted(groups, key=lambda k: (-len(groups[k]), disp[k]))[:MAX_COUNSELLOR_TABS]:
        leads = groups[key]
        cb = disp[key]
        tname = sanitize_tab_name("CB - " + cb)
        # avoid collision with reserved names
        if tname in tabs:
            tname = sanitize_tab_name("CB - " + cb + " ")
        tb = Tab(tname)
        add_report_header(tb, f"Counsellor: {cb}", period_range, gen_stamp)
        add_exec_summary(tb, leads)
        # Lead Source Performance for THIS counsellor's leads — same structure and
        # logic as the main Summary tab (source_perf + SRC_HEADER), scoped to the
        # counsellor's own active leads.
        tb.title("Lead Source Performance")
        tb.header(SRC_HEADER)
        for r in source_perf(leads):
            tb.row(r)
        tb.blank()
        # Fresh + Repeat lead details for this counsellor.
        cb_fresh = [a for a in leads if a["_is_new"]]
        cb_repeat = [a for a in leads if not a["_is_new"]]
        tb.title(f"Fresh Lead Details  ({len(cb_fresh)})")
        if cb_fresh:
            lead_detail_rows(tb, cb_fresh)
        else:
            tb.row(["(no fresh leads in this period)"])
        tb.blank()
        tb.title(f"Repeat Lead Details  ({len(cb_repeat)})")
        if cb_repeat:
            lead_detail_rows(tb, cb_repeat)
        else:
            tb.row(["(no repeat leads in this period)"])
        tabs[tname] = tb
    return tabs, active


# ============================================================
# Presentation palette + conditional-colour helpers
# ============================================================
# RGB tuples (0-1) for Google Sheets; HEX mirrors for the xlsx backup.
CLR_TITLE_BG = (0.11, 0.21, 0.37); CLR_TITLE_FG = (1, 1, 1)
CLR_SUB_BG   = (0.90, 0.94, 0.99); CLR_SUB_FG   = (0.11, 0.21, 0.37)
CLR_SEC_BG   = (0.17, 0.33, 0.53); CLR_SEC_FG   = (1, 1, 1)
CLR_HDR_BG   = (0.82, 0.87, 0.95); CLR_HDR_FG   = (0.10, 0.16, 0.28)
CLR_ALT_BG   = (0.955, 0.970, 0.990)
CLR_GREEN    = (0.78, 0.90, 0.74)
CLR_ORANGE   = (1.00, 0.90, 0.66)
CLR_RED      = (0.97, 0.80, 0.78)
CLR_FRESH    = (0.83, 0.92, 0.98)     # Fresh lead-type chip
CLR_REPEAT   = (1.00, 0.93, 0.80)     # Repeat lead-type chip

HEX = {"TITLE":"1B355E","TITLE_FG":"FFFFFF","SUB":"E6EEFB","SUB_FG":"1B355E",
       "SEC":"2B547E","SEC_FG":"FFFFFF","HDR":"D2DEF2","HDR_FG":"1A2A48",
       "ALT":"F4F8FD","GREEN":"C6EFCE","ORANGE":"FFEB9C","RED":"FFC7CE",
       "FRESH":"D3EAFB","REPEAT":"FFEBCC"}

# % thresholds for the traffic-light colouring of contribution / percentage cells
PCT_HIGH = 25.0     # >= green (high)
PCT_MED  = 10.0     # >= orange (medium); below = red (low)

# Metrics where a LOWER percentage is the GOOD result (data-quality KPIs). Their
# %-cell colouring must run in REVERSE of the normal traffic light: low = green,
# moderate = amber, high = red. Matched on the metric label (the row's first cell).
INVERSE_PCT_METRICS = {"Invalid Phone Number Leads", "Irrelevant Leads"}


def _pct_rgb_inverse(v):
    """Reverse traffic-light for 'lower-is-better' metrics: below PCT_MED is green
    (good), PCT_MED..PCT_HIGH is amber, at/above PCT_HIGH is red (undesirable)."""
    if v is None:
        return None
    return CLR_GREEN if v < PCT_MED else (CLR_ORANGE if v < PCT_HIGH else CLR_RED)


def _pct_hex_inverse(v):
    if v is None:
        return None
    return HEX["GREEN"] if v < PCT_MED else (HEX["ORANGE"] if v < PCT_HIGH else HEX["RED"])


def _pct_value(cell):
    m = re.match(r"^\s*([\d.]+)\s*%\s*$", str(cell))
    try:
        return float(m.group(1)) if m else None
    except ValueError:
        return None


def _pct_rgb(v):
    if v is None:
        return None
    return CLR_GREEN if v >= PCT_HIGH else (CLR_ORANGE if v >= PCT_MED else CLR_RED)


def _cell_conditional_rgb(value):
    """Return an RGB fill for a data cell based on its value, or None."""
    v = _pct_value(value)
    if v is not None:
        return _pct_rgb(v)
    sv = str(value).strip()
    if sv == "Fresh":
        return CLR_FRESH
    if sv == "Repeat":
        return CLR_REPEAT
    return None


def _pct_hex(value):
    v = _pct_value(value)
    if v is not None:
        return HEX["GREEN"] if v >= PCT_HIGH else (HEX["ORANGE"] if v >= PCT_MED else HEX["RED"])
    sv = str(value).strip()
    if sv == "Fresh":
        return HEX["FRESH"]
    if sv == "Repeat":
        return HEX["REPEAT"]
    return None


def _row_conditional_rgb(rowvals):
    """Whole-row fill colour: the %-traffic-light category takes precedence
    (green/orange/red), then a Fresh/Repeat lead-type tint. None -> no colour.
    For 'lower-is-better' metrics (Invalid Phone / Irrelevant) the %-colouring is
    inverted so a low percentage reads green and a high one reads red."""
    inverse = bool(rowvals) and str(rowvals[0]).strip() in INVERSE_PCT_METRICS
    for v in rowvals:
        pv = _pct_value(v)
        if pv is not None:
            return _pct_rgb_inverse(pv) if inverse else _pct_rgb(pv)
    for v in rowvals:
        sv = str(v).strip()
        if sv == "Fresh":
            return CLR_FRESH
        if sv == "Repeat":
            return CLR_REPEAT
    return None


def _row_conditional_hex(rowvals):
    inverse = bool(rowvals) and str(rowvals[0]).strip() in INVERSE_PCT_METRICS
    for v in rowvals:
        pv = _pct_value(v)
        if pv is not None:
            return _pct_hex_inverse(pv) if inverse else _pct_hex(v)
    for v in rowvals:
        sv = str(v).strip()
        if sv == "Fresh":
            return HEX["FRESH"]
        if sv == "Repeat":
            return HEX["REPEAT"]
    return None


# ============================================================
# Google Sheets / Drive I/O
# ============================================================
def _creds(scopes, impersonate=False):
    """Service-account credentials for the given scopes. impersonate=True applies
    domain-wide delegation (with_subject) — used ONLY for Drive writes, and ONLY
    with the drive scope, so the DWD grant (drive) covers every requested scope."""
    from google.oauth2 import service_account
    if not os.path.exists(SERVICE_ACCOUNT_FILE):
        sys.exit(f"ERROR: service account not found at {SERVICE_ACCOUNT_FILE}")
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=scopes)
    if impersonate and IMPERSONATE_USER:
        creds = creds.with_subject(IMPERSONATE_USER)
    return creds


def get_read_service():
    """Sheets client as the service account itself (no impersonation, spreadsheets
    scope) — used to READ the master sheet."""
    from googleapiclient.discovery import build
    return build("sheets", "v4",
                 credentials=_creds(READ_SCOPES, impersonate=False),
                 cache_discovery=False)


def get_drive_service():
    """Drive client impersonating the Workspace user with the DRIVE scope only —
    used to UPLOAD the report files (matching pyAttendaceFeedbackReport.py)."""
    from googleapiclient.discovery import build
    return build("drive", "v3",
                 credentials=_creds(DRIVE_SCOPES, impersonate=True),
                 cache_discovery=False)


def read_master_df(sheets):
    """Read the consolidated master sheet into a DataFrame."""
    if LOCAL_MASTER_CSV:
        return pd.read_csv(LOCAL_MASTER_CSV, dtype=str, keep_default_na=False)
    meta = sheets.spreadsheets().get(spreadsheetId=MASTER_SHEET_ID).execute()
    title = MASTER_TAB_NAME or meta["sheets"][0]["properties"]["title"]
    resp = sheets.spreadsheets().values().get(
        spreadsheetId=MASTER_SHEET_ID, range=title,
        valueRenderOption="FORMATTED_VALUE").execute()
    values = resp.get("values", [])
    if not values:
        return pd.DataFrame()
    header = [h if h else f"col_{i}" for i, h in enumerate(values[0])]
    n = len(header)
    rows = [(r + [""] * n)[:n] for r in values[1:]]
    return pd.DataFrame(rows, columns=header).astype(str)


def load_enrolled_mobiles(sheets=None):
    """Return the set of NORMALIZED phone numbers of already-enrolled students
    from 'IntelliBI — Student Admission Responses'. Matched by every Phone/Mobile
    column present. Never raises — on any problem (sheet not shared, no phone
    column, offline) it returns an empty set so the report still runs with no
    exclusion applied."""
    try:
        if sheets is None:
            sheets = get_read_service()
        meta = sheets.spreadsheets().get(spreadsheetId=ENROLLED_SHEET_ID).execute()
        title = ENROLLED_TAB_NAME or meta["sheets"][0]["properties"]["title"]
        resp = sheets.spreadsheets().values().get(
            spreadsheetId=ENROLLED_SHEET_ID, range=title,
            valueRenderOption="FORMATTED_VALUE").execute()
        values = resp.get("values", [])
        if not values:
            return set()
        header = [s(h) for h in values[0]]
        phone_idx = [i for i, h in enumerate(header)
                     if any(k in h.lower()
                            for k in ("mobile", "phone", "whatsapp", "contact number"))]
        if not phone_idx:
            print("  [enrolled] no Phone/Mobile column found in Student Admission "
                  "Responses — no exclusion applied.")
            return set()
        out = set()
        for r in values[1:]:
            for i in phone_idx:
                if i < len(r):
                    p = norm_phone(r[i])
                    if p:
                        out.add(p)
        return out
    except Exception as exc:                       # noqa: BLE001 — degrade gracefully
        print(f"  [enrolled] could not load Student Admission Responses "
              f"({exc}); no exclusion applied.")
        return set()


def upload_report_to_drive(drive, folder_id, name, xlsx_path):
    """Upload the styled .xlsx into the Drive folder as a Google Sheet, replacing
    any same-named file first. Uses ONLY the drive scope (impersonated), exactly
    like pyAttendaceFeedbackReport.py — so it does NOT need the `spreadsheets`
    scope in the domain-wide-delegation grant. Returns (url, created)."""
    from googleapiclient.http import MediaFileUpload

    safe = name.replace("'", "\\'")
    # Remove any previous file(s) with this name in the folder (Sheet or xlsx).
    try:
        res = drive.files().list(
            q="name = '%s' and '%s' in parents and trashed = false" % (safe, folder_id),
            fields="files(id)", pageSize=20,
            supportsAllDrives=True, includeItemsFromAllDrives=True).execute()
        old = res.get("files", [])
        for fobj in old:
            drive.files().delete(fileId=fobj["id"], supportsAllDrives=True).execute()
    except Exception as e:
        print("  [drive] could not remove previous file(s):", e)
        old = []

    media = MediaFileUpload(
        xlsx_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False)
    body = {"name": name, "parents": [folder_id],
            "mimeType": "application/vnd.google-apps.spreadsheet"}  # convert to Sheet
    try:
        f = drive.files().create(body=body, media_body=media,
                                 fields="id,webViewLink",
                                 supportsAllDrives=True).execute()
    except Exception as e:
        if "storageQuota" in str(e):
            sys.exit(
                "\nERROR: 'storageQuotaExceeded' creating the report file.\n"
                "The Drive upload must be owned by a Workspace user. Ensure\n"
                "IMPERSONATE_USER is set to a user who owns the output folders and\n"
                "that domain-wide delegation authorises the `drive` scope for the\n"
                "service account client ID. See README_LeadPerformanceReport.md.\n")
        raise
    url = f.get("webViewLink") or \
        f"https://docs.google.com/spreadsheets/d/{f['id']}/edit"
    return url, (len(old) == 0), f["id"]


def share_file_with(drive, file_id, emails, role="writer"):
    """Grant `role` (default Editor/writer) on a Drive file to one or more email
    addresses. Best-effort per address; failures are logged, never fatal. No
    Google notification email is sent (our own report email carries the link)."""
    for em in (emails if isinstance(emails, (list, tuple, set)) else [emails]):
        em = s(em)
        if not em:
            continue
        try:
            drive.permissions().create(
                fileId=file_id,
                body={"type": "user", "role": role, "emailAddress": em},
                sendNotificationEmail=False,
                supportsAllDrives=True).execute()
            print(f"  [drive] shared masked report with {em} ({role})")
        except Exception as e:
            print(f"  [drive] could NOT share masked report with {em}:", e)


def write_workbook(sheets, spreadsheet_id, tabs):
    """Create/replace exactly the given tabs (name->Tab) and apply light formatting."""
    meta = sheets.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    existing = {sh["properties"]["title"]: sh["properties"]["sheetId"] for sh in meta["sheets"]}

    # add a temporary tab so we can freely delete the rest
    reqs = [{"addSheet": {"properties": {"title": "__tmp__"}}}] if "__tmp__" not in existing else []
    for name in tabs:
        if name not in existing:
            reqs.append({"addSheet": {"properties": {"title": name}}})
    if reqs:
        sheets.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body={"requests": reqs}).execute()
        meta = sheets.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        existing = {sh["properties"]["title"]: sh["properties"]["sheetId"] for sh in meta["sheets"]}

    # delete any tab not in our target set (keep __tmp__ until the end)
    del_reqs = [{"deleteSheet": {"sheetId": sid}} for nm, sid in existing.items()
                if nm not in tabs and nm != "__tmp__"]
    if del_reqs:
        sheets.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body={"requests": del_reqs}).execute()

    # write values + set order
    order_reqs, fmt_reqs, clear_ranges, chart_reqs, merge_reqs = [], [], [], [], []
    rich_reqs = []                    # rich-text (per-value colour) cell updates
    meta = sheets.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    existing = {sh["properties"]["title"]: sh["properties"]["sheetId"] for sh in meta["sheets"]}
    for idx, (name, tab) in enumerate(tabs.items()):
        clear_ranges.append("'%s'!A1:ZZ10000" % name)
    sheets.spreadsheets().values().batchClear(
        spreadsheetId=spreadsheet_id, body={"ranges": clear_ranges}).execute()

    data = []
    for name, tab in tabs.items():
        w = tab.width()
        vals = [row + [""] * (w - len(row)) for row in tab.rows]
        data.append({"range": "'%s'!A1" % name, "values": vals})
    sheets.spreadsheets().values().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"valueInputOption": "RAW", "data": data}).execute()

    for idx, (name, tab) in enumerate(tabs.items()):
        sid = existing[name]
        maxw = max(tab.width(), 1)
        # Freeze only the top banner by default. For a tab whose detail table sits
        # near the top (Fresh/Repeat Details), freeze through its header so the
        # column labels stay visible. CB tabs (detail table is far down, below the
        # summary) freeze ONLY the banner — never the whole summary.
        fh0 = tab.filter_headers[0] if tab.filter_headers else None
        frozen = (fh0 + 1) if (fh0 is not None and fh0 <= 5) else 1
        order_reqs.append({"updateSheetProperties": {
            "properties": {"sheetId": sid, "index": idx},
            "fields": "index"}})
        order_reqs.append({"updateSheetProperties": {
            "properties": {"sheetId": sid,
                           "gridProperties": {"frozenRowCount": frozen}},
            "fields": "gridProperties.frozenRowCount"}})

        # report title banner (full width), header info rows, section titles,
        # and table headers — each spanning only its own width.
        _ml = getattr(tab, "multiline_banner", False)
        _wide = getattr(tab, "wide_banner", False)
        _span = getattr(tab, "header_span", maxw)
        for ri in tab.banners:
            if _wide:
                # full single-line header over a wide colour band (text in col A
                # overflows across the empty band cells — never truncated).
                fmt_reqs.append(_fmt_range(sid, ri, ri + 1, 0, _span,
                                           bold=True, size=13,
                                           bg=CLR_TITLE_BG, fg=CLR_TITLE_FG,
                                           valign="MIDDLE"))
            else:
                fmt_reqs.append(_fmt_range(sid, ri, ri + 1, 0, tab.fill_width(ri),
                                           bold=True, size=13,
                                           bg=CLR_TITLE_BG, fg=CLR_TITLE_FG,
                                           wrap=_ml, valign=("MIDDLE" if _ml else None)))
        for ri in tab.info:
            fmt_reqs.append(_fmt_range(sid, ri, ri + 1, 0, tab.fill_width(ri),
                                       bold=True, bg=CLR_SUB_BG, fg=CLR_SUB_FG))
        for ri in tab.titles:
            fmt_reqs.append(_fmt_range(sid, ri, ri + 1, 0, tab.fill_width(ri),
                                       bold=True, size=12,
                                       bg=CLR_SEC_BG, fg=CLR_SEC_FG))
        for ri in tab.headers:
            fmt_reqs.append(_fmt_range(sid, ri, ri + 1, 0, tab.fill_width(ri),
                                       bold=True, bg=CLR_HDR_BG, fg=CLR_HDR_FG))

        # data rows: whole-row conditional colour (across the TABLE's own width
        # only), else a subtle alternating band. KPI rows are bold.
        for hi in tab.headers:
            a, b = tab.data_span(hi)
            w = len(tab.rows[hi])
            _color_rows = getattr(tab, "color_data_rows", True)
            _fills = getattr(tab, "row_fills", None) or {}
            for k, ri in enumerate(range(a, b)):
                if ri in _fills:                       # explicit row colour wins
                    rgb = _fills[ri][0]
                else:
                    rgb = _row_conditional_rgb(tab.rows[ri]) if _color_rows else None
                bold = ri in tab.kpi
                if rgb:
                    fmt_reqs.append(_fmt_range(sid, ri, ri + 1, 0, w, bold=bold, bg=rgb))
                elif k % 2 == 1:
                    fmt_reqs.append(_fmt_range(sid, ri, ri + 1, 0, w, bold=bold,
                                               bg=CLR_ALT_BG))
                elif bold:
                    fmt_reqs.append(_fmt_range(sid, ri, ri + 1, 0, w, bold=True))

        # filter on the primary detail table (spanning that table's width)
        if tab.filter_headers:
            hi = tab.filter_headers[0]
            a, b = tab.data_span(hi)
            fmt_reqs.append({"setBasicFilter": {"filter": {"range": {
                "sheetId": sid, "startRowIndex": hi, "endRowIndex": max(b, hi + 1),
                "startColumnIndex": 0, "endColumnIndex": len(tab.rows[hi])}}}})

        fmt_reqs.append({"autoResizeDimensions": {"dimensions": {
            "sheetId": sid, "dimension": "COLUMNS", "startIndex": 0,
            "endIndex": maxw}}})
        # For a wide-banner tab, the long header/average text lives in col A, so
        # auto-resize would blow that column up. Pin it back to a narrow width
        # (the text still overflows across the band to its right).
        if _wide:
            fmt_reqs.append({"updateDimensionProperties": {
                "range": {"sheetId": sid, "dimension": "COLUMNS",
                          "startIndex": 0, "endIndex": 1},
                "properties": {"pixelSize": 72}, "fields": "pixelSize"}})

        # combined average line: light band across the header span + rich text in
        # col A with each metric value bold and colour-coded.
        al = getattr(tab, "avg_line", None)
        if al:
            fmt_reqs.append(_fmt_range(sid, al["row"], al["row"] + 1, 0, _span,
                                       bg=CLR_SUB_BG, valign="MIDDLE"))
            rich_reqs.append(_avg_line_request(sid, al["row"], al))

        # optional chart(s) for this tab (drawn in its own batch, below)
        if getattr(tab, "chart", None):
            chart_reqs.append(build_gchart_request(sid, tab.chart))
        if getattr(tab, "chart2", None):
            chart_reqs.append(build_gchart_request(sid, tab.chart2))

        # header average-metric lines: label merged A:C (bold, right), value in
        # col D (bold, colour-coded); the whole line gets the light header band.
        for av in getattr(tab, "avg_rows", []):
            ri = av["row"]
            merge_reqs.append({"mergeCells": {"range": {
                "sheetId": sid, "startRowIndex": ri, "endRowIndex": ri + 1,
                "startColumnIndex": 0, "endColumnIndex": 3}, "mergeType": "MERGE_ALL"}})
            fmt_reqs.append(_fmt_range(sid, ri, ri + 1, 0, maxw,
                                       bg=CLR_SUB_BG, valign="MIDDLE"))
            fmt_reqs.append(_fmt_range(sid, ri, ri + 1, 0, 3, bold=True,
                                       fg=CLR_SUB_FG, bg=CLR_SUB_BG,
                                       halign="RIGHT", valign="MIDDLE"))
            fmt_reqs.append(_fmt_range(sid, ri, ri + 1, 3, 4, bold=True,
                                       fg=av["rgb"], bg=CLR_SUB_BG,
                                       halign="LEFT", valign="MIDDLE"))
    # remove temporary tab
    if "__tmp__" in existing:
        fmt_reqs.append({"deleteSheet": {"sheetId": existing["__tmp__"]}})
    # Formatting must never break the (already-written) data: isolate it.
    try:
        sheets.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": order_reqs + merge_reqs + fmt_reqs + rich_reqs}).execute()
    except Exception as e:
        print("  [format] non-fatal formatting error:", e)
        try:
            sheets.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id, body={"requests": order_reqs}).execute()
        except Exception:
            pass
    # Charts in their own isolated batch — a chart error never affects the data.
    if chart_reqs:
        try:
            sheets.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id, body={"requests": chart_reqs}).execute()
        except Exception as e:
            print("  [chart] non-fatal chart error:", e)


def _fmt_range(sid, r0, r1, c0, c1, bold=False, size=None, bg=None, fg=None,
               wrap=False, valign=None, halign=None):
    tf = {"bold": bold}
    if size:
        tf["fontSize"] = size
    if fg:
        tf["foregroundColor"] = {"red": fg[0], "green": fg[1], "blue": fg[2]}
    fmt = {"textFormat": tf}
    if bg:
        fmt["backgroundColor"] = {"red": bg[0], "green": bg[1], "blue": bg[2]}
    if wrap:
        fmt["wrapStrategy"] = "WRAP"
    if valign:
        fmt["verticalAlignment"] = valign
    if halign:
        fmt["horizontalAlignment"] = halign
    parts = ["textFormat"]
    if bg:
        parts.append("backgroundColor")
    if wrap:
        parts.append("wrapStrategy")
    if valign:
        parts.append("verticalAlignment")
    if halign:
        parts.append("horizontalAlignment")
    fields = "userEnteredFormat(%s)" % ",".join(parts)
    return {"repeatCell": {
        "range": {"sheetId": sid, "startRowIndex": r0, "endRowIndex": r1,
                  "startColumnIndex": c0, "endColumnIndex": c1},
        "cell": {"userEnteredFormat": fmt},
        "fields": fields}}


def _avg_line_request(sid, ri, al):
    """updateCells request writing the combined average line (col A) with per-value
    colour runs: the whole line is bold, the label segments use the sub-header
    colour, and each metric value keeps its green/red colour-coding."""
    size = al.get("size")
    runs, idx = [], 0
    for text, rgb, _hex in al["segments"]:
        rf = {"bold": True,
              "foregroundColor": {"red": rgb[0], "green": rgb[1], "blue": rgb[2]}}
        if size:
            rf["fontSize"] = size
        runs.append({"startIndex": idx, "format": rf})
        idx += len(text)
    base_tf = {"bold": True, "foregroundColor": {
        "red": CLR_SUB_FG[0], "green": CLR_SUB_FG[1], "blue": CLR_SUB_FG[2]}}
    if size:
        base_tf["fontSize"] = size
    return {"updateCells": {
        "rows": [{"values": [{
            "userEnteredValue": {"stringValue": al["text"]},
            "userEnteredFormat": {
                "textFormat": base_tf,
                "backgroundColor": {
                    "red": CLR_SUB_BG[0], "green": CLR_SUB_BG[1], "blue": CLR_SUB_BG[2]},
                "verticalAlignment": "MIDDLE"},
            "textFormatRuns": runs}]}],
        "fields": "userEnteredValue,userEnteredFormat,textFormatRuns",
        "start": {"sheetId": sid, "rowIndex": ri, "columnIndex": 0}}}


# ---- offline xlsx backup (same professional look as the Google Sheet) ----
def write_local_xlsx(path, tabs):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def fill(hexcode):
        return PatternFill("solid", fgColor=hexcode)

    thin = Side(style="thin", color="D9DEE8")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, tab in tabs.items():
        ws = wb.create_sheet(sanitize_tab_name(name)[:31])
        maxw = max(tab.width(), 1)
        for row in tab.rows:
            ws.append(row if row else [""])

        def paint(ri, hexbg=None, fg=None, bold=False, size=None, span=None):
            span = span if span is not None else tab.fill_width(ri)
            for ci in range(1, span + 1):
                c = ws.cell(row=ri + 1, column=ci)
                c.font = Font(bold=bold, color=(fg or "000000"), size=(size or 11))
                if hexbg:
                    c.fill = fill(hexbg)

        # banner (single merged header row) / section titles / headers
        _ml = getattr(tab, "multiline_banner", False)
        _wide = getattr(tab, "wide_banner", False)
        _span = getattr(tab, "header_span", maxw)
        for ri in tab.banners:
            if _wide:
                # full single-line header over a wide colour band; NO merge so the
                # text overflows across the empty band cells and never truncates.
                paint(ri, HEX["TITLE"], HEX["TITLE_FG"], bold=True, size=13, span=_span)
                ws.cell(ri + 1, 1).alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=False)
                continue
            paint(ri, HEX["TITLE"], HEX["TITLE_FG"], bold=True, size=13, span=maxw)
            if maxw > 1:
                ws.merge_cells(start_row=ri + 1, start_column=1,
                               end_row=ri + 1, end_column=maxw)
            ws.cell(ri + 1, 1).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=_ml)
            if _ml:
                # three stacked lines -> give the row enough height to show them.
                n_lines = str(tab.rows[ri][0]).count("\n") + 1
                ws.row_dimensions[ri + 1].height = 18 * n_lines + 6
        for ri in tab.info:
            paint(ri, HEX["SUB"], HEX["SUB_FG"], bold=True)
        for ri in tab.titles:
            paint(ri, HEX["SEC"], HEX["SEC_FG"], bold=True, size=12)
        for ri in tab.headers:
            paint(ri, HEX["HDR"], HEX["HDR_FG"], bold=True)
            for ci in range(1, len(tab.rows[ri]) + 1):
                c = ws.cell(row=ri + 1, column=ci)
                c.border = grid
                c.alignment = Alignment(vertical="center", wrap_text=True)

        # header average-metric lines: light band, label merged A:C (bold, right),
        # value in col D (bold, colour-coded green/red).
        for av in getattr(tab, "avg_rows", []):
            ri = av["row"]
            for ci in range(1, maxw + 1):
                ws.cell(ri + 1, ci).fill = fill(HEX["SUB"])
            if maxw >= 3:
                ws.merge_cells(start_row=ri + 1, start_column=1,
                               end_row=ri + 1, end_column=3)
            lab = ws.cell(ri + 1, 1)
            lab.font = Font(bold=True, color=HEX["SUB_FG"])
            lab.alignment = Alignment(horizontal="right", vertical="center")
            val = ws.cell(ri + 1, av["value_col"] + 1)
            val.font = Font(bold=True, color=av["hex"])
            val.alignment = Alignment(horizontal="left", vertical="center")

        # combined average line (single line, each value bold + colour-coded):
        # a light band across the header span with rich text in col A (overflows).
        al = getattr(tab, "avg_line", None)
        if al:
            from openpyxl.cell.rich_text import CellRichText, TextBlock
            from openpyxl.cell.text import InlineFont
            from openpyxl.styles.colors import Color
            ri = al["row"]
            span = getattr(tab, "header_span", maxw)
            size = al.get("size")
            for ci in range(1, span + 1):
                ws.cell(ri + 1, ci).fill = fill(HEX["SUB"])

            def _inline(hexc):
                kw = dict(b=True, color=Color(rgb="FF" + hexc))
                if size:
                    kw["sz"] = size
                return InlineFont(**kw)
            blocks = [TextBlock(_inline(hexc), text)
                      for text, _rgb, hexc in al["segments"]]
            c = ws.cell(ri + 1, 1)
            c.value = CellRichText(blocks)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

        # data rows: whole-row conditional colour across the TABLE width only.
        _color_rows = getattr(tab, "color_data_rows", True)
        _fills = getattr(tab, "row_fills", None) or {}
        for hi in tab.headers:
            a, b = tab.data_span(hi)
            w = len(tab.rows[hi])
            for k, ri in enumerate(range(a, b)):
                if ri in _fills:                       # explicit row colour wins
                    rowhex = _fills[ri][1]
                else:
                    rowhex = _row_conditional_hex(tab.rows[ri]) if _color_rows else None
                band = rowhex or (HEX["ALT"] if k % 2 == 1 else None)
                bold = ri in tab.kpi
                for ci in range(1, w + 1):
                    cell = ws.cell(row=ri + 1, column=ci)
                    cell.border = grid
                    if band:
                        cell.fill = fill(band)
                    if bold:
                        cell.font = Font(bold=True)
                    if "\n" in str(cell.value or ""):   # multi-line (Lead Journey)
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

        # freeze: through the detail header only if it's near the top; else the
        # banner row (CB tabs never freeze the whole summary).
        fh0 = tab.filter_headers[0] if tab.filter_headers else None
        if fh0 is not None and fh0 <= 5:
            ws.freeze_panes = f"A{fh0 + 2}"
        else:
            ws.freeze_panes = "A2"
        if tab.filter_headers:
            hi = tab.filter_headers[0]
            a, b = tab.data_span(hi)
            last_col = get_column_letter(len(tab.rows[hi]))
            ws.auto_filter.ref = f"A{hi + 1}:{last_col}{max(b, hi + 1)}"

        # auto-ish column widths (skip the wide header banner and the average-metric
        # lines so their long overflow text never widens the narrow Day column).
        _skip_w = {av["row"] for av in getattr(tab, "avg_rows", [])}
        if getattr(tab, "wide_banner", False):
            _skip_w |= set(tab.banners)
        if getattr(tab, "avg_line", None):
            _skip_w.add(tab.avg_line["row"])
        for ci in range(1, maxw + 1):
            longest = 0
            for rj, row in enumerate(tab.rows):
                if rj in _skip_w:
                    continue
                if ci - 1 < len(row):
                    longest = max(longest, len(str(row[ci - 1])))
            ws.column_dimensions[get_column_letter(ci)].width = min(max(longest + 2, 10), 48)

        # hide helper columns (e.g. the short graph X-axis label) so the visible
        # table is unchanged while the chart can still read them for categories.
        for _hc in sorted(getattr(tab, "hidden_cols", None) or []):
            ws.column_dimensions[get_column_letter(_hc + 1)].hidden = True

        # optional line chart (same data table the Google Sheet chart uses)
        if getattr(tab, "chart", None):
            spec = tab.chart
            if spec.get("chart_kind") == "COLUMN_STACKED":
                try:
                    from openpyxl.chart import BarChart, Reference
                    from openpyxl.chart.series import SeriesLabel
                    from openpyxl.chart.shapes import GraphicalProperties
                    ch = BarChart()
                    ch.type = "col"
                    ch.grouping = "stacked"
                    ch.overlap = 100
                    ch.title = spec["title"]
                    ch.x_axis.title = spec.get("x_title", "Hour of Day")
                    ch.y_axis.title = spec.get("y_title", "Number of Enquiries")
                    ch.x_axis.delete = False
                    ch.y_axis.delete = False
                    ch.height, ch.width, ch.style = 12, 26, 10
                    cats = Reference(ws, min_col=spec["date_col"] + 1, max_col=spec["date_col"] + 1,
                                     min_row=spec["first_row"] + 1, max_row=spec["last_row"] + 1)
                    titles = spec.get("series_titles") or []
                    colors = spec.get("series_colors_hex") or []
                    _cols = spec["series_cols"]
                    # Add the series in FORWARD column order (Working Professional
                    # - IT first ... Unidentified last). The first series sits at the
                    # BOTTOM of the stack, so this puts WP-IT at the bottom and
                    # Unidentified at the TOP. Google Sheets locks the legend order
                    # to the stack (top-of-stack = first legend entry), so the
                    # bottom legend then reads Unidentified first ... WP-IT last.
                    for j in range(len(_cols)):
                        col1 = _cols[j] + 1
                        ref = Reference(ws, min_col=col1, max_col=col1,
                                        min_row=spec["header_row"] + 1, max_row=spec["last_row"] + 1)
                        ch.add_data(ref, titles_from_data=True)
                    ch.set_categories(cats)
                    _rt = list(titles)
                    _rc = list(colors)
                    for i, ser in enumerate(ch.series):
                        if i < len(_rt):
                            ser.tx = SeriesLabel(v=_rt[i])
                        if i < len(_rc) and _rc[i]:
                            ser.graphicalProperties = GraphicalProperties(solidFill=_rc[i])
                    if ch.legend is not None:
                        ch.legend.position = spec.get("legend_x", "b")
                        ch.legend.overlay = False
                    anchor = get_column_letter(spec["anchor_col"] + 1) + str(spec["anchor_row"] + 1)
                    ws.add_chart(ch, anchor)
                except Exception as e:
                    print("  [chart] non-fatal local-xlsx stacked-column error:", e)
                # done with this tab's chart
                spec = None
            if spec is not None:
                try:
                    from openpyxl.chart import LineChart, Reference
                    from openpyxl.chart.label import DataLabelList
                    ch = LineChart()
                    ch.title = spec["title"]
                    ch.x_axis.title = spec.get("x_title", "Date")
                    ch.y_axis.title = spec.get("y_title", "Number of Leads")
                    ch.x_axis.delete = False
                    ch.y_axis.delete = False
                    ch.height, ch.width, ch.style = 12, 26, 12
                    # show ONLY the numeric value at each point (note: the correct
                    # openpyxl attribute is showSerName — showSeriesName is ignored,
                    # which is why the series name used to leak into the label).
                    ch.dataLabels = DataLabelList()
                    ch.dataLabels.showVal = True
                    ch.dataLabels.showSerName = False
                    ch.dataLabels.showCatName = False
                    ch.dataLabels.showLegendKey = False
                    ch.dataLabels.numFmt = "0"
                    smin, smax = min(spec["series_cols"]) + 1, max(spec["series_cols"]) + 1
                    data = Reference(ws, min_col=smin, max_col=smax,
                                     min_row=spec["header_row"] + 1, max_row=spec["last_row"] + 1)
                    ch.add_data(data, titles_from_data=True)
                    cats = Reference(ws, min_col=spec["date_col"] + 1, max_col=spec["date_col"] + 1,
                                     min_row=spec["first_row"] + 1, max_row=spec["last_row"] + 1)
                    ch.set_categories(cats)
                    # round markers + explicit series names + line/marker colours
                    # (green = Total Fresh, brown = Fresh Relevant).
                    from openpyxl.chart.marker import Marker
                    from openpyxl.chart.series import SeriesLabel
                    from openpyxl.chart.shapes import GraphicalProperties
                    from openpyxl.drawing.line import LineProperties
                    titles = spec.get("series_titles") or []
                    colors = spec.get("series_colors_hex") or []
                    for i, ser in enumerate(ch.series):
                        ser.smooth = False
                        if i < len(titles):
                            ser.tx = SeriesLabel(v=titles[i])
                        hexc = colors[i] if i < len(colors) else None
                        ser.marker = Marker(symbol="circle", size=7)
                        if hexc:
                            ser.graphicalProperties = GraphicalProperties()
                            ser.graphicalProperties.line = LineProperties(solidFill=hexc, w=28000)
                            ser.marker.graphicalProperties = GraphicalProperties(solidFill=hexc)
                    if ch.legend is not None:
                        ch.legend.position = spec.get("legend_x", "b")   # 'tr' = top-right
                        ch.legend.overlay = False
                    anchor = get_column_letter(spec["anchor_col"] + 1) + str(spec["anchor_row"] + 1)
                    ws.add_chart(ch, anchor)
                except Exception as e:
                    print("  [chart] non-fatal local-xlsx chart error:", e)

        # optional horizontal BAR chart — Lead Type Share. Count + share % are
        # carried in the axis labels (Google drops on-bar data labels on import).
        # Bars render Unidentified (top) -> WP-IT (bottom).
        bspec = getattr(tab, "bar_share", None)
        if bspec:
            try:
                from openpyxl.chart import BarChart, Reference
                from openpyxl.chart.label import DataLabelList
                from openpyxl.chart.shapes import GraphicalProperties
                from openpyxl.chart.series import DataPoint
                bch = BarChart()
                bch.type = "bar"                     # horizontal bars
                bch.grouping = "clustered"
                bch.title = bspec["title"]
                bch.height, bch.width, bch.style = 10, 18, 11
                bch.varyColors = True                # one colour per Lead Type bar
                data = Reference(ws, min_col=bspec["val_col"] + 1, max_col=bspec["val_col"] + 1,
                                 min_row=bspec["header_row"] + 1, max_row=bspec["last_row"] + 1)
                bch.add_data(data, titles_from_data=True)
                cats = Reference(ws, min_col=bspec["cat_col"] + 1, max_col=bspec["cat_col"] + 1,
                                 min_row=bspec["first_row"] + 1, max_row=bspec["last_row"] + 1)
                bch.set_categories(cats)
                # one colour per bar, in the fixed Lead Type order
                colors = bspec.get("colors_hex") or []
                if bch.series:
                    ser = bch.series[0]
                    pts = []
                    for i, hexc in enumerate(colors):
                        dp = DataPoint(idx=i)
                        if hexc:
                            dp.graphicalProperties = GraphicalProperties(solidFill=hexc)
                        pts.append(dp)
                    ser.data_points = pts
                # Count is shown in the axis label (Google ignores on-bar labels).
                # Keep a value label too so the count still shows if the workbook
                # is opened directly in Excel / LibreOffice.
                bch.dataLabels = DataLabelList()
                bch.dataLabels.showVal = True
                bch.dataLabels.showSerName = False
                bch.dataLabels.showCatName = False
                bch.dataLabels.showLegendKey = False
                bch.dataLabels.showPercent = False
                # no legend — each Lead Type is already named on the axis
                bch.legend = None
                anchor = get_column_letter(bspec["anchor_col"] + 1) + str(bspec["anchor_row"] + 1)
                ws.add_chart(bch, anchor)
            except Exception as e:
                print("  [chart] non-fatal local-xlsx bar-share error:", e)

        # optional SECOND line chart — Lead Source Share (multi-line, one line per
        # source). Rendered exactly like tab.chart's line chart, from its own table.
        c2 = getattr(tab, "chart2", None)
        if c2:
            try:
                from openpyxl.chart import LineChart, Reference
                from openpyxl.chart.label import DataLabelList
                from openpyxl.chart.marker import Marker
                from openpyxl.chart.series import SeriesLabel
                from openpyxl.chart.shapes import GraphicalProperties
                from openpyxl.drawing.line import LineProperties
                ch = LineChart()
                ch.title = c2["title"]
                ch.x_axis.title = c2.get("x_title", "Date")
                ch.y_axis.title = c2.get("y_title", "Number of Leads")
                ch.x_axis.delete = False
                ch.y_axis.delete = False
                ch.height, ch.width, ch.style = 12, 26, 12
                ch.dataLabels = DataLabelList()
                ch.dataLabels.showVal = True
                ch.dataLabels.showSerName = False
                ch.dataLabels.showCatName = False
                ch.dataLabels.showLegendKey = False
                ch.dataLabels.numFmt = "0"
                smin, smax = min(c2["series_cols"]) + 1, max(c2["series_cols"]) + 1
                data = Reference(ws, min_col=smin, max_col=smax,
                                 min_row=c2["header_row"] + 1, max_row=c2["last_row"] + 1)
                ch.add_data(data, titles_from_data=True)
                cats = Reference(ws, min_col=c2["date_col"] + 1, max_col=c2["date_col"] + 1,
                                 min_row=c2["first_row"] + 1, max_row=c2["last_row"] + 1)
                ch.set_categories(cats)
                titles = c2.get("series_titles") or []
                colors = c2.get("series_colors_hex") or []
                for i, ser in enumerate(ch.series):
                    ser.smooth = False
                    if i < len(titles):
                        ser.tx = SeriesLabel(v=titles[i])
                    hexc = colors[i] if i < len(colors) else None
                    ser.marker = Marker(symbol="circle", size=7)
                    if hexc:
                        ser.graphicalProperties = GraphicalProperties()
                        ser.graphicalProperties.line = LineProperties(solidFill=hexc, w=28000)
                        ser.marker.graphicalProperties = GraphicalProperties(solidFill=hexc)
                if ch.legend is not None:
                    ch.legend.position = c2.get("legend_x", "b")
                    ch.legend.overlay = False
                anchor = get_column_letter(c2["anchor_col"] + 1) + str(c2["anchor_row"] + 1)
                ws.add_chart(ch, anchor)
            except Exception as e:
                print("  [chart] non-fatal local-xlsx chart2 error:", e)
    wb.save(path)
    _patch_charts_plot_hidden(path)


def _patch_charts_plot_hidden(path):
    """openpyxl always writes <c:plotVisOnly val="1"/>, which makes Excel/Sheets
    plot ONLY visible cells — so a chart whose category range is a HIDDEN helper
    column (our short graph X-axis label) renders BLANK axis labels. Rewrite every
    chart to plotVisOnly=0 so those hidden category cells are still plotted. Charts
    without hidden data are unaffected. Never raises."""
    try:
        import zipfile
        tmp = path + ".tmp"
        with zipfile.ZipFile(path, "r") as zin, \
             zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if (item.filename.startswith("xl/charts/chart")
                        and item.filename.endswith(".xml")):
                    text = data.decode("utf-8")
                    if 'plotVisOnly val="1"' in text:
                        text = text.replace('plotVisOnly val="1"',
                                            'plotVisOnly val="0"')
                    data = text.encode("utf-8")
                zout.writestr(item, data)
        os.replace(tmp, path)
    except Exception as e:                          # cosmetic — never fail the write
        print("  [chart] plotVisOnly patch skipped:", e)


# ============================================================
# Email
# ============================================================
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _valid_recipients(recipients):
    """Clean + validate recipients before sending. Splits accidental joins
    (comma / semicolon / whitespace), trims, drops anything that is not a single
    valid e-mail address, and de-duplicates case-insensitively.

    This makes a missing comma in EMAIL_RECIPIENTS (which Python silently fuses
    into one bad address, e.g. 'info@x.in' + 'manish@y.com' ->
    'info@x.inmanish@y.com' with two '@') surface as a loud warning and get
    skipped, instead of being mailed to a dead address that never arrives."""
    seen, out = set(), []
    for raw in recipients or []:
        for part in re.split(r"[,\s;]+", str(raw).strip()):
            addr = part.strip()
            if not addr:
                continue
            if not _EMAIL_RE.match(addr):
                print(f"  [email] WARNING skipping malformed recipient {addr!r} "
                      f"— check EMAIL_RECIPIENTS for a missing comma")
                continue
            key = addr.lower()
            if key not in seen:
                seen.add(key)
                out.append(addr)
    return out


def send_email(subject, html_body, recipients=None):
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    recipients = _valid_recipients(
        recipients if recipients is not None else EMAIL_RECIPIENTS)
    if not recipients:
        print("  [email] no valid recipients — nothing sent")
        return
    sys.path.insert(0, PROJECT_ROOT)
    try:
        import email_config as ec
        sender, app_pass = ec.GMAIL_SENDER, ec.GMAIL_APP_PASS
    except Exception as e:
        print("  [email] skipped - could not load config_files/email_config.py:", e)
        return
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = ", ".join(recipients)
    msg.attach(MIMEText(html_body, "html"))
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as srv:
            srv.login(sender, app_pass)
            srv.sendmail(sender, recipients, msg.as_string())
        print("  [email] sent to", ", ".join(recipients))
    except Exception as e:
        print("  [email] FAILED:", e)


def build_email_body(report_type, period_range, url, link_name, active, gen_stamp,
                     start=None, end=None):
    """Professional, self-contained HTML email: intro, period, KPI snapshot,
    a named call-to-action link, and the IntelliBI signature."""
    es = exec_summary(active)

    def val(metric):
        v = es.get(metric)
        return v[0] if v else 0

    kpis = [("Total Leads", val("Total Leads (Fresh + Repeat)"), NAVY_HEX),
            ("Fresh (New)", val("Fresh (New) Leads"), NAVY_HEX),
            ("Repeat", val("Repeat Leads"), NAVY_HEX),
            ("Referral", val("Referral Leads"), NAVY_HEX)]
    # Fresh (Non-Referral) = fresh/new leads excluding referrals; Fresh-Relevant
    # (Non-Referral) = of those, the ones marked Relevant. Referral leads are
    # never counted in either figure.
    fresh_nonref = sum(1 for a in active
                       if a.get("_is_new") and not is_referral(a))
    fresh_rel_nonref = sum(1 for a in active
                           if a.get("_is_new") and not is_referral(a)
                           and yes(a.get(C_RELEV)))

    # Weekly/Monthly also show per-day averages of those two figures, and colour-
    # code all four: the two Fresh figures share the Avg-Daily-Fresh > 15 verdict
    # (green/red); the two Relevant figures share the Avg-Relevant-% > 80 verdict.
    if report_type in ("Weekly", "Monthly") and start is not None and end is not None:
        last = min(end.date(), now_ist().date())
        ndays = max((last - start.date()).days + 1, 1)
        avg_daily_fresh = fresh_nonref / ndays
        avg_daily_fresh_rel = fresh_rel_nonref / ndays
        avg_rel_pct = (fresh_rel_nonref / fresh_nonref * 100.0) if fresh_nonref else 0.0
        fresh_hex = TXT_GREEN_HEX if avg_daily_fresh > 15 else TXT_RED_HEX
        rel_hex = TXT_GREEN_HEX if avg_rel_pct > 80 else TXT_RED_HEX
        kpis2 = [
            ("Fresh (Non-Referral)", fresh_nonref, fresh_hex),
            ("Fresh-Relevant (Non-Referral)", fresh_rel_nonref, rel_hex),
            ("Avg Daily Fresh (Non-Referral)", f"{avg_daily_fresh:.2f}", fresh_hex),
            ("Avg Daily Fresh-Relevant (Non-Referral)", f"{avg_daily_fresh_rel:.2f}", fresh_hex),
            ("Avg Daily Fresh-Relevant (Non-Referral)", f"{avg_rel_pct:.1f}%", rel_hex),
        ]
    else:
        # Daily (single day): the two counts only (no per-day average), colour-
        # coded directly — Fresh (Non-Referral) green when the count > 15;
        # Fresh-Relevant (Non-Referral) green when its % of that count > 80.
        rel_pct = (fresh_rel_nonref / fresh_nonref * 100.0) if fresh_nonref else 0.0
        fresh_hex = TXT_GREEN_HEX if fresh_nonref > 15 else TXT_RED_HEX
        rel_hex = TXT_GREEN_HEX if rel_pct > 80 else TXT_RED_HEX
        kpis2 = [("Fresh (Non-Referral)", fresh_nonref, fresh_hex),
                 ("Fresh-Relevant (Non-Referral)", fresh_rel_nonref, rel_hex)]

    def _cards(items):
        return "".join(
            "<td style='padding:6px'>"
            "<div style='background:#f4f8fd;border:1px solid #e2e8f0;border-radius:8px;"
            "padding:14px 10px;text-align:center'>"
            f"<div style='font-size:24px;font-weight:700;color:#{color}'>{v}</div>"
            f"<div style='font-size:12px;color:#5b6b86;margin-top:2px'>{lbl}</div>"
            "</div></td>"
            for lbl, v, color in items)
    cards = _cards(kpis)
    cards2 = _cards(kpis2)

    return f"""<html><body style="margin:0;padding:24px;background:#eef2f8;
  font-family:'Segoe UI',Roboto,Arial,sans-serif;color:#1a2a48">
  <div style="max-width:640px;margin:0 auto;background:#ffffff;border-radius:10px;
    overflow:hidden;border:1px solid #e2e8f0">
    <div style="background:#1B355E;padding:22px 28px;color:#ffffff">
      <div style="font-size:20px;font-weight:700">IntelliBI &nbsp;&middot;&nbsp; {report_type} Lead Report</div>
      <div style="font-size:13px;opacity:.85;margin-top:4px">Reporting Period: {period_range}</div>
    </div>
    <div style="padding:24px 28px">
      <p style="margin:0 0 14px">Hello Team,</p>
      <p style="margin:0 0 18px;line-height:1.5">
        Please find the <b>{report_type}</b> lead-performance report for
        <b>{period_range}</b>. Here is a quick snapshot:</p>
      <table role="presentation" width="100%" style="border-collapse:separate;
        margin:0 -6px 10px"><tr>{cards}</tr></table>
      <table role="presentation" width="100%" style="border-collapse:separate;
        margin:0 -6px 20px"><tr>{cards2}</tr></table>
      <p style="margin:0 0 22px;line-height:1.5">The full report covers lead-source
        &amp; counsellor performance, fresh vs. repeat leads, and each lead's complete
        journey from first enquiry to the latest interaction.</p>
      <p style="text-align:center;margin:0 0 24px">
        <a href="{url}" style="background:#2B547E;color:#ffffff;text-decoration:none;
          padding:13px 30px;border-radius:6px;font-weight:600;font-size:15px;
          display:inline-block">Open {link_name} &nbsp;&rsaquo;</a></p>
      <p style="margin:0;color:#5b6b86;font-size:13px">
        Or open it here: <a href="{url}" style="color:#2B547E;font-weight:600;
        text-decoration:none">{link_name}</a></p>
      <p style="margin:26px 0 0;line-height:1.5">
        Thanks &amp; Regards,<br><b>IntelliBI Automation Team</b></p>
    </div>
    <div style="background:#f4f8fd;padding:12px 28px;font-size:12px;color:#8494ad;
      border-top:1px solid #e8edf5">
      Automated report &middot; Generated {gen_stamp}</div>
  </div>
</body></html>"""


# ============================================================
# MAIN
# ============================================================
def run():
    today = now_ist().date()
    sheets = drive = None
    # Read the master as the service account directly (no impersonation needed);
    # upload the reports via an impersonated Drive client (drive scope only).
    if not LOCAL_MASTER_CSV:
        sheets = get_read_service()
    if not DRY_RUN:
        drive = get_drive_service()

    df = read_master_df(sheets)
    print(f"Master rows loaded: {len(df)}")

    # --- exclude already-enrolled students from the counted population --------
    # A lead whose phone appears in "IntelliBI — Student Admission Responses" is
    # an enrolled student following up post-admission; their re-appearances must
    # not inflate Repeat Leads / Total Leads or any dependent metric, tab, chart,
    # percentage or counsellor rollup. Rows are filtered ONLY from this in-memory
    # report DataFrame — the source Google Sheet is never modified. Every
    # downstream section (including the masked copy) derives from this df, so all
    # counts stay reconciled after the exclusion.
    enrolled = load_enrolled_mobiles(sheets)
    if enrolled and C_MOBILE in df.columns:
        before = len(df)
        keep = ~df[C_MOBILE].map(lambda v: norm_phone(v) in enrolled)
        df = df[keep].reset_index(drop=True)
        print(f"Enrolled-student exclusion: {len(enrolled)} enrolled mobile(s); "
              f"removed {before - len(df)} lead row(s); {len(df)} remain.")
    else:
        print(f"Enrolled-student exclusion: {len(enrolled)} enrolled mobile(s) "
              f"loaded — no rows removed.")

    # each job: (label, rng, start, end, folder_id, filename, subject, link_name)
    jobs = []
    if GENERATE_DAILY_REPORT:
        d = datetime.strptime(DAILY_REPORT_DATE, "%Y-%m-%d").date() if DAILY_REPORT_DATE else today
        st, en = day_bounds(d)
        jobs.append(("Daily", d.strftime("%d-%b-%Y"), st, en, DAILY_FOLDER_ID,
                     f"Lead Performance - Daily - {d.strftime('%d-%b-%Y')}",
                     f"Daily Lead Report - {d.strftime('%d-%b-%Y')}",
                     "Daily Lead Report"))
    if GENERATE_WEEKLY_REPORT:
        ref = datetime.strptime(WEEKLY_REPORT_REFERENCE_DATE, "%Y-%m-%d").date() if WEEKLY_REPORT_REFERENCE_DATE else today
        st, en, mon, sun = week_bounds(ref)
        rng = f"{mon.strftime('%d-%b-%Y')} to {sun.strftime('%d-%b-%Y')}"
        iso_year, iso_week, _ = mon.isocalendar()
        jobs.append(("Weekly", rng, st, en, WEEKLY_FOLDER_ID,
                     f"Lead Performance - Weekly - {mon.strftime('%d-%b-%Y')} to {sun.strftime('%d-%b-%Y')}",
                     f"Weekly Lead Report - Week {iso_week} ({iso_year})",
                     "Weekly Lead Report"))
    if GENERATE_MONTHLY_REPORT:
        yr = MONTHLY_REPORT_YEAR or today.year
        mo = MONTHLY_REPORT_MONTH or today.month
        st, en, ms, me = month_bounds(yr, mo)
        rng = f"{ms.strftime('%d-%b-%Y')} to {me.strftime('%d-%b-%Y')}"
        jobs.append(("Monthly", rng, st, en, MONTHLY_FOLDER_ID,
                     f"Lead Performance - Monthly - {ms.strftime('%b-%Y')}",
                     f"Monthly Lead Report - {ms.strftime('%b %Y')}",
                     "Monthly Lead Report"))

    for label, rng, st, en, folder, fname, subject, link_name in jobs:
        tabs, active = build_report(label, rng, df, st, en)
        print(f"\n{label} report | {rng} | active leads: {len(active)} | tabs: {len(tabs)}")

        # styled .xlsx is both the local backup AND the file uploaded to Drive.
        xlsx_path = os.path.join(OUTPUT_DIR, fname + ".xlsx")
        xlsx_ok = False
        try:
            write_local_xlsx(xlsx_path, tabs)
            xlsx_ok = True
            print(f"  local xlsx: {xlsx_path}")
        except Exception as e:
            print("  local xlsx FAILED:", e)

        # Split recipients: those who must get a masked copy vs. everyone else.
        masked_recips = [r for r in EMAIL_RECIPIENTS
                         if s(r).lower() in {m.lower() for m in MASK_RECIPIENTS}]
        normal_recips = [r for r in EMAIL_RECIPIENTS if r not in masked_recips]

        # Build a SEPARATE masked report (from a masked COPY of the data) only
        # when a masking recipient is configured. The original `df`/report is
        # never modified, so authorised recipients still get the full data.
        masked_xlsx_path = None
        tabs_m = active_m = None
        if masked_recips:
            df_masked = mask_dataframe(df)
            tabs_m, active_m = build_report(label, rng, df_masked, st, en)
            masked_xlsx_path = os.path.join(OUTPUT_DIR, fname + " (Masked).xlsx")
            try:
                write_local_xlsx(masked_xlsx_path, tabs_m)
                print(f"  local xlsx (masked): {masked_xlsx_path}")
            except Exception as e:
                print("  local xlsx (masked) FAILED:", e)
                masked_xlsx_path = None

        if DRY_RUN:
            continue
        if not (xlsx_ok and os.path.exists(xlsx_path)):
            print("  [drive] no xlsx to upload — skipping this report.")
            continue
        # Upload the styled .xlsx to Drive as a Google Sheet (drive scope only).
        url, created, _fid = upload_report_to_drive(drive, folder, fname, xlsx_path)
        print(f"  {'created' if created else 'replaced'}: {url}")


        # Upload the masked copy (separate file) for the masking recipients, and
        # share THAT file with them directly as Editor so the link works for them
        # regardless of the folder's sharing (other recipients are unaffected).
        url_masked = None
        if masked_recips and masked_xlsx_path and os.path.exists(masked_xlsx_path):
            url_masked, _mc, masked_fid = upload_report_to_drive(
                drive, folder, fname + " (Masked)", masked_xlsx_path)
            print(f"  masked copy: {url_masked}")
            share_file_with(drive, masked_fid, masked_recips, role="writer")

        # One professional email per report, with its own subject + named link.
        # Normal recipients get the full report; masking recipients get the
        # masked copy — same subject/body, only the linked report differs.
        if SEND_EMAIL:
            gen_stamp = now_ist().strftime("%d-%b-%Y %I:%M %p") + " IST"
            if normal_recips:
                send_email(subject,
                           build_email_body(label, rng, url, link_name, active,
                                            gen_stamp, start=st, end=en),
                           normal_recips)
            if masked_recips:
                if url_masked:
                    send_email(subject,
                               build_email_body(label, rng, url_masked, link_name,
                                                active_m if active_m is not None else active,
                                                gen_stamp, start=st, end=en),
                               masked_recips)
                else:
                    # Never send the FULL report to a masking recipient. If the
                    # masked copy couldn't be produced/uploaded, skip them.
                    print("  [email] masked copy unavailable — NOT emailing masked "
                          "recipients (to avoid sending unmasked data):",
                          ", ".join(masked_recips))
    print("\nDone.")


if __name__ == "__main__":
    run()
